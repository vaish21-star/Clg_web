from flask import Flask, render_template, request, redirect, session, send_file, url_for, jsonify
import os
import hashlib
import random
import csv
import io
import time
import smtplib
import zipfile
import re
import secrets
from email.message import EmailMessage
from datetime import date, datetime
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
try:
    from pypdf import PdfReader
except Exception:
    try:
        from PyPDF2 import PdfReader
    except Exception:
        PdfReader = None
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False

from db import get_db
from pdf_utils import (
    generate_admission_letter,
    generate_fee_receipt,
    generate_students_list_pdf,
    generate_results_summary_pdf,
    generate_result_student_pdf,
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev_secret_key")

AUTH_RATE_LIMIT_WINDOW_SECONDS = 300
AUTH_RATE_LIMIT_MAX_ATTEMPTS = 5
_AUTH_FAILED_ATTEMPTS = {}


def load_local_env(env_path=".env"):
    if not os.path.exists(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as env_file:
            for raw_line in env_file:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and not os.environ.get(key):
                    os.environ[key] = value
    except Exception:
        # If .env cannot be read, continue with OS environment variables.
        pass


def hash_password(password):
    return generate_password_hash(password)


def verify_password(stored_hash, raw_password):
    if not stored_hash:
        return False
    try:
        if check_password_hash(stored_hash, raw_password):
            return True
    except Exception:
        pass
    legacy_sha256 = hashlib.sha256(raw_password.encode()).hexdigest()
    return stored_hash == legacy_sha256


def client_ip():
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or "unknown"


def _rate_limit_bucket(action, identity):
    return f"{action}:{identity}"


def is_auth_rate_limited(action, identity):
    now = time.time()
    key = _rate_limit_bucket(action, identity)
    attempts = _AUTH_FAILED_ATTEMPTS.get(key, [])
    attempts = [ts for ts in attempts if now - ts <= AUTH_RATE_LIMIT_WINDOW_SECONDS]
    _AUTH_FAILED_ATTEMPTS[key] = attempts
    return len(attempts) >= AUTH_RATE_LIMIT_MAX_ATTEMPTS


def record_auth_failure(action, identity):
    now = time.time()
    key = _rate_limit_bucket(action, identity)
    attempts = _AUTH_FAILED_ATTEMPTS.get(key, [])
    attempts = [ts for ts in attempts if now - ts <= AUTH_RATE_LIMIT_WINDOW_SECONDS]
    attempts.append(now)
    _AUTH_FAILED_ATTEMPTS[key] = attempts


def clear_auth_failures(action, identity):
    key = _rate_limit_bucket(action, identity)
    _AUTH_FAILED_ATTEMPTS.pop(key, None)


def ensure_csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def validate_csrf():
    expected = session.get("_csrf_token")
    provided = request.form.get("_csrf_token", "")
    if not expected or not provided:
        return False
    return secrets.compare_digest(str(expected), str(provided))


@app.context_processor
def inject_csrf_token():
    return {"csrf_token": ensure_csrf_token}


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_local_env(os.path.join(BASE_DIR, ".env"))


UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs("static/pdfs", exist_ok=True)
SYLLABUS_UPLOAD_FOLDER = "static/syllabus_uploads"
os.makedirs(SYLLABUS_UPLOAD_FOLDER, exist_ok=True)
RESULT_UPLOAD_FOLDER = "static/result_uploads"
os.makedirs(RESULT_UPLOAD_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

BRANCH_OPTIONS = [
    "Automobile Engineering",
    "Computer Science and Engineering",
    "Electronics and Communication Engineering",
    "Mechanical Engineering",
]

SUBJECT_TYPE_OPTIONS = ["THEORY", "PRACTICAL", "AUDIT COURSE"]
DEFAULT_SUBJECT_SERIES_OPTIONS = ["C15 SERIES", "C20 SERIES", "C25 SERIES"]


def ensure_employees_table():
    db = get_db()
    cur = db.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS employee_details (
            id INT AUTO_INCREMENT PRIMARY KEY,
            employee_name VARCHAR(100) NOT NULL,
            department VARCHAR(100) NOT NULL,
            designation VARCHAR(100) NOT NULL,
            mobile_no VARCHAR(15) NOT NULL,
            employee_type VARCHAR(30) NOT NULL DEFAULT 'TEACHING',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    db.commit()
    cur.close()
    db.close()


def ensure_students_rejection_reason_column():
    db = get_db()
    cur = db.cursor()
    try:
        # Safer than querying information_schema in restricted DB setups.
        cur.execute("SHOW COLUMNS FROM students LIKE 'rejection_reason'")
        row = cur.fetchone()
        if not row:
            cur.execute("ALTER TABLE students ADD COLUMN rejection_reason TEXT NULL")
            db.commit()
    except Exception:
        # If column already exists/race condition, continue.
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_students_status_supports_rejected():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SHOW COLUMNS FROM students LIKE 'status'")
        row = cur.fetchone()
        if not row:
            return

        # row format: Field, Type, Null, Key, Default, Extra
        col_type = str(row[1]).lower() if len(row) > 1 else ""
        if col_type.startswith("enum(") and "'rejected'" not in col_type:
            cur.execute("""
                ALTER TABLE students
                MODIFY COLUMN status ENUM('INACTIVE','ACTIVE','REJECTED')
                NOT NULL DEFAULT 'INACTIVE'
            """)
            db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_students_college_reg_no_column():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SHOW COLUMNS FROM students LIKE 'college_reg_no'")
        row = cur.fetchone()
        if not row:
            cur.execute("ALTER TABLE students ADD COLUMN college_reg_no VARCHAR(100) NULL")
            db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_staff_auth_tables():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS staff_accounts (
                id INT AUTO_INCREMENT PRIMARY KEY,
                employee_id INT NULL UNIQUE,
                employee_name VARCHAR(150) NULL,
                department VARCHAR(150) NULL,
                designation VARCHAR(100) NULL,
                email VARCHAR(255) NOT NULL UNIQUE,
                password_hash VARCHAR(255) NOT NULL,
                is_verified TINYINT(1) NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("SHOW COLUMNS FROM staff_accounts LIKE 'employee_id'")
        emp_col = cur.fetchone()
        if emp_col and str(emp_col[2]).upper() == "NO":
            cur.execute("ALTER TABLE staff_accounts MODIFY COLUMN employee_id INT NULL UNIQUE")
        cur.execute("SHOW COLUMNS FROM staff_accounts LIKE 'employee_name'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE staff_accounts ADD COLUMN employee_name VARCHAR(150) NULL")
        cur.execute("SHOW COLUMNS FROM staff_accounts LIKE 'department'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE staff_accounts ADD COLUMN department VARCHAR(150) NULL")
        cur.execute("SHOW COLUMNS FROM staff_accounts LIKE 'designation'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE staff_accounts ADD COLUMN designation VARCHAR(100) NULL")
        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def generate_otp():
    return f"{random.randint(100000, 999999)}"


def send_otp_email(to_email, subject, otp_code):
    sender = os.getenv("SVP_GMAIL_USER") or os.getenv("GMAIL_USER")
    app_password = os.getenv("SVP_GMAIL_APP_PASSWORD") or os.getenv("GMAIL_APP_PASSWORD")
    if not sender or not app_password:
        return False, "Gmail SMTP is not configured on server."

    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = sender
        msg["To"] = to_email
        msg.set_content(
            f"Your OTP is: {otp_code}\n"
            "This OTP is valid for 10 minutes.\n"
            "If you did not request this, ignore this email."
        )
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(sender, app_password)
            smtp.send_message(msg)
        return True, ""
    except Exception:
        return False, "Failed to send OTP email."


def start_otp_flow(flow_key, email, extra_data):
    otp = generate_otp()
    session["otp_flow"] = {
        "flow_key": flow_key,
        "email": email.lower().strip(),
        "otp": otp,
        "expires_at": time.time() + 600,
        "extra": extra_data or {},
    }
    return otp


def verify_otp_flow(flow_key, email, otp):
    flow = session.get("otp_flow") or {}
    if not flow:
        return False, "OTP session not found."
    if flow.get("flow_key") != flow_key:
        return False, "Invalid OTP flow."
    if flow.get("email") != email.lower().strip():
        return False, "Email mismatch."
    if time.time() > float(flow.get("expires_at", 0)):
        return False, "OTP expired."
    if flow.get("otp") != str(otp).strip():
        return False, "Incorrect OTP."
    return True, ""


def get_access_scope():
    if "admin" not in session:
        return {"allowed": False, "is_staff": False, "department": None, "designation": None, "staff_name": None}

    staff_id = session.get("staff_id")
    if not staff_id:
        return {"allowed": True, "is_staff": False, "department": None, "designation": None, "staff_name": None}

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT id, employee_name, department, designation FROM staff_accounts WHERE id=%s", (staff_id,))
    staff = cur.fetchone()
    cur.close()
    db.close()

    if not staff:
        session.clear()
        return {"allowed": False, "is_staff": False, "department": None, "designation": None, "staff_name": None}

    return {
        "allowed": True,
        "is_staff": True,
        "staff_name": (staff.get("employee_name") or "").strip(),
        "department": (staff.get("department") or "").strip(),
        "designation": (staff.get("designation") or "").strip(),
    }


def student_in_department(admission_id, department):
    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT branch FROM students WHERE admission_id=%s", (admission_id,))
    row = cur.fetchone()
    cur.close()
    db.close()
    return bool(row and (row.get("branch") or "").strip().lower() == (department or "").strip().lower())


def can_edit_fees(scope):
    if not scope.get("is_staff"):
        return True
    dept = (scope.get("department") or "").strip().lower()
    desig = (scope.get("designation") or "").strip().lower()
    if "management" in dept:
        return True
    if "hod" in desig or "head of department" in desig:
        return True
    return False


def can_upload_syllabus(scope):
    if not scope.get("is_staff"):
        return True
    designation = (scope.get("designation") or "").strip().lower()
    return "hod" in designation or "head of department" in designation


def can_edit_student_identity_ids(scope):
    if not scope.get("is_staff"):
        return True
    designation = (scope.get("designation") or "").strip().lower()
    return "hod" in designation or "head of department" in designation


def can_set_allotted_category(scope):
    if not scope.get("is_staff"):
        return True
    dept = (scope.get("department") or "").strip().lower()
    desig = (scope.get("designation") or "").strip().lower()
    return "management" in dept or "management" in desig


def ensure_students_admission_year_column():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SHOW COLUMNS FROM students LIKE 'admission_year'")
        row = cur.fetchone()
        if not row:
            cur.execute("ALTER TABLE students ADD COLUMN admission_year INT NULL")
            db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_students_year_sem_column():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SHOW COLUMNS FROM students LIKE 'year_sem'")
        row = cur.fetchone()
        if not row:
            cur.execute("ALTER TABLE students ADD COLUMN year_sem TINYINT NULL")
            db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_student_personal_extra_columns():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SHOW COLUMNS FROM student_personal_details LIKE 'ssp_id'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE student_personal_details ADD COLUMN ssp_id VARCHAR(100) NULL")
        cur.execute("SHOW COLUMNS FROM student_personal_details LIKE 'apaar_id'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE student_personal_details ADD COLUMN apaar_id VARCHAR(100) NULL")
        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_qualifying_exam_support():
    db = get_db()
    cur = db.cursor()
    try:
        required = ("'sslc'", "'puc'", "'iti'", "'cbse'", "'icse'")

        cur.execute("SHOW COLUMNS FROM student_personal_details LIKE 'qualifying_exam'")
        spd = cur.fetchone()
        if spd:
            col_type = str(spd[1]).lower()
            if col_type.startswith("enum(") and not all(v in col_type for v in required):
                cur.execute("""
                    ALTER TABLE student_personal_details
                    MODIFY COLUMN qualifying_exam ENUM('SSLC','PUC','ITI','CBSE','ICSE') NULL
                """)

        cur.execute("SHOW COLUMNS FROM education_details LIKE 'qualifying_exam'")
        ed = cur.fetchone()
        if ed:
            col_type = str(ed[1]).lower()
            if col_type.startswith("enum(") and not all(v in col_type for v in required):
                cur.execute("""
                    ALTER TABLE education_details
                    MODIFY COLUMN qualifying_exam ENUM('SSLC','PUC','ITI','CBSE','ICSE') NULL
                """)

        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_syllabus_documents_table():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS syllabus_documents (
                id INT AUTO_INCREMENT PRIMARY KEY,
                branch VARCHAR(120) NOT NULL,
                semester_no TINYINT NOT NULL,
                series VARCHAR(50) NOT NULL,
                subject_name VARCHAR(150) NOT NULL DEFAULT '',
                year1_pdf VARCHAR(255) NULL,
                year2_pdf VARCHAR(255) NULL,
                year3_pdf VARCHAR(255) NULL,
                uploaded_by_staff_id INT NULL,
                uploaded_by_role VARCHAR(20) NOT NULL DEFAULT 'ADMIN',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_syllabus_period_subject (branch, semester_no, series, subject_name)
            )
        """)
        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_student_results_table():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_results (
                id INT AUTO_INCREMENT PRIMARY KEY,
                register_number VARCHAR(50) NOT NULL,
                student_name VARCHAR(160) NOT NULL,
                branch VARCHAR(120) NOT NULL,
                semester_no TINYINT NULL,
                subject_code VARCHAR(50) NULL,
                subject_name VARCHAR(255) NULL,
                ia_marks VARCHAR(20) NULL,
                theory_marks VARCHAR(20) NULL,
                practical_marks VARCHAR(20) NULL,
                result_status VARCHAR(20) NULL,
                credit VARCHAR(20) NULL,
                grade VARCHAR(20) NULL,
                final_result VARCHAR(50) NULL,
                cgpa VARCHAR(20) NULL,
                percentage VARCHAR(20) NULL,
                credit_earned_total VARCHAR(20) NULL,
                credit_applied_s1 VARCHAR(20) NULL,
                credit_applied_s2 VARCHAR(20) NULL,
                credit_applied_s3 VARCHAR(20) NULL,
                credit_applied_s4 VARCHAR(20) NULL,
                credit_applied_s5 VARCHAR(20) NULL,
                credit_applied_s6 VARCHAR(20) NULL,
                credit_earned_s1 VARCHAR(20) NULL,
                credit_earned_s2 VARCHAR(20) NULL,
                credit_earned_s3 VARCHAR(20) NULL,
                credit_earned_s4 VARCHAR(20) NULL,
                credit_earned_s5 VARCHAR(20) NULL,
                credit_earned_s6 VARCHAR(20) NULL,
                sgpa_s1 VARCHAR(20) NULL,
                sgpa_s2 VARCHAR(20) NULL,
                sgpa_s3 VARCHAR(20) NULL,
                sgpa_s4 VARCHAR(20) NULL,
                sgpa_s5 VARCHAR(20) NULL,
                sgpa_s6 VARCHAR(20) NULL,
                attempts_s1 VARCHAR(20) NULL,
                attempts_s2 VARCHAR(20) NULL,
                attempts_s3 VARCHAR(20) NULL,
                attempts_s4 VARCHAR(20) NULL,
                attempts_s5 VARCHAR(20) NULL,
                attempts_s6 VARCHAR(20) NULL,
                exam_session VARCHAR(80) NULL,
                source_file VARCHAR(255) NULL,
                imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        extra_columns = [
            ("final_result", "VARCHAR(50) NULL"),
            ("cgpa", "VARCHAR(20) NULL"),
            ("percentage", "VARCHAR(20) NULL"),
            ("credit_earned_total", "VARCHAR(20) NULL"),
            ("credit_applied_s1", "VARCHAR(20) NULL"),
            ("credit_applied_s2", "VARCHAR(20) NULL"),
            ("credit_applied_s3", "VARCHAR(20) NULL"),
            ("credit_applied_s4", "VARCHAR(20) NULL"),
            ("credit_applied_s5", "VARCHAR(20) NULL"),
            ("credit_applied_s6", "VARCHAR(20) NULL"),
            ("credit_earned_s1", "VARCHAR(20) NULL"),
            ("credit_earned_s2", "VARCHAR(20) NULL"),
            ("credit_earned_s3", "VARCHAR(20) NULL"),
            ("credit_earned_s4", "VARCHAR(20) NULL"),
            ("credit_earned_s5", "VARCHAR(20) NULL"),
            ("credit_earned_s6", "VARCHAR(20) NULL"),
            ("sgpa_s1", "VARCHAR(20) NULL"),
            ("sgpa_s2", "VARCHAR(20) NULL"),
            ("sgpa_s3", "VARCHAR(20) NULL"),
            ("sgpa_s4", "VARCHAR(20) NULL"),
            ("sgpa_s5", "VARCHAR(20) NULL"),
            ("sgpa_s6", "VARCHAR(20) NULL"),
            ("attempts_s1", "VARCHAR(20) NULL"),
            ("attempts_s2", "VARCHAR(20) NULL"),
            ("attempts_s3", "VARCHAR(20) NULL"),
            ("attempts_s4", "VARCHAR(20) NULL"),
            ("attempts_s5", "VARCHAR(20) NULL"),
            ("attempts_s6", "VARCHAR(20) NULL"),
        ]
        for col_name, col_def in extra_columns:
            cur.execute(f"SHOW COLUMNS FROM student_results LIKE '{col_name}'")
            if not cur.fetchone():
                cur.execute(f"ALTER TABLE student_results ADD COLUMN {col_name} {col_def}")
        cur.execute("SHOW INDEX FROM student_results WHERE Key_name='idx_student_results_reg_no'")
        if not cur.fetchone():
            cur.execute("CREATE INDEX idx_student_results_reg_no ON student_results(register_number)")
        cur.execute("SHOW INDEX FROM student_results WHERE Key_name='idx_student_results_branch_sem'")
        if not cur.fetchone():
            cur.execute("CREATE INDEX idx_student_results_branch_sem ON student_results(branch, semester_no)")
        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_fee_module_tables():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_fee_structure (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admission_id VARCHAR(50) NOT NULL UNIQUE,
                admission_fee_due DECIMAL(10,2) NOT NULL DEFAULT 0,
                tuition_fee_yearly_due DECIMAL(10,2) NOT NULL DEFAULT 0,
                management_fee_yearly_due DECIMAL(10,2) NOT NULL DEFAULT 0,
                exam_fee_per_sem_due DECIMAL(10,2) NOT NULL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fee_payments (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admission_id VARCHAR(50) NOT NULL,
                fee_type ENUM('ADMISSION','TUITION','MANAGEMENT','EXAM') NOT NULL,
                academic_year VARCHAR(20) NULL,
                semester_no TINYINT NULL,
                amount DECIMAL(10,2) NOT NULL,
                payment_date DATE NOT NULL,
                receipt_no VARCHAR(60) NOT NULL UNIQUE,
                remarks TEXT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fee_structure_master (
                id INT AUTO_INCREMENT PRIMARY KEY,
                branch VARCHAR(100) NOT NULL,
                semester_no TINYINT NOT NULL,
                academic_year VARCHAR(20) NOT NULL,
                admission_fee_due DECIMAL(10,2) NOT NULL DEFAULT 0,
                tuition_fee_due DECIMAL(10,2) NOT NULL DEFAULT 0,
                management_fee_due DECIMAL(10,2) NOT NULL DEFAULT 0,
                exam_fee_due DECIMAL(10,2) NOT NULL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_fee_structure_period (branch, semester_no, academic_year)
            )
        """)
        cur.execute("SHOW INDEX FROM fee_payments WHERE Key_name='idx_fee_payments_admission_id'")
        if not cur.fetchone():
            cur.execute("CREATE INDEX idx_fee_payments_admission_id ON fee_payments(admission_id)")
        cur.execute("SHOW INDEX FROM fee_payments WHERE Key_name='idx_fee_payments_payment_date'")
        if not cur.fetchone():
            cur.execute("CREATE INDEX idx_fee_payments_payment_date ON fee_payments(payment_date)")
        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_academic_module_tables():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_subjects (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admission_id VARCHAR(50) NOT NULL,
                semester_no TINYINT NOT NULL DEFAULT 1,
                subject_code VARCHAR(30) NOT NULL,
                subject_name VARCHAR(150) NOT NULL,
                internal_max INT NOT NULL DEFAULT 25,
                external_max INT NOT NULL DEFAULT 75,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_subject_student_sem_code (admission_id, semester_no, subject_code)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_internal_marks (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admission_id VARCHAR(50) NOT NULL,
                semester_no TINYINT NOT NULL DEFAULT 1,
                subject_code VARCHAR(30) NOT NULL,
                internal_marks DECIMAL(6,2) NOT NULL DEFAULT 0,
                external_marks DECIMAL(6,2) NOT NULL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_marks_student_sem_code (admission_id, semester_no, subject_code)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_attendance (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admission_id VARCHAR(50) NOT NULL,
                semester_no TINYINT NOT NULL DEFAULT 1,
                subject_code VARCHAR(30) NOT NULL,
                total_classes INT NOT NULL DEFAULT 0,
                present_classes INT NOT NULL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_attendance_student_sem_code (admission_id, semester_no, subject_code)
            )
        """)
        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_subject_master_table():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS subjects (
                id INT AUTO_INCREMENT PRIMARY KEY,
                branch VARCHAR(120) NOT NULL,
                semester TINYINT NOT NULL,
                series VARCHAR(40) NOT NULL DEFAULT 'C20 SERIES',
                subject_name VARCHAR(180) NOT NULL,
                course_code VARCHAR(50) NOT NULL,
                subject_type ENUM('THEORY','PRACTICAL','AUDIT COURSE') NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_subjects_branch_sem_series_code (branch, semester, series, course_code)
            )
        """)
        cur.execute("SHOW COLUMNS FROM subjects LIKE 'series'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE subjects ADD COLUMN series VARCHAR(40) NOT NULL DEFAULT 'C20 SERIES' AFTER semester")
        cur.execute("SHOW INDEX FROM subjects WHERE Key_name='uq_subjects_branch_sem_code'")
        if cur.fetchone():
            cur.execute("ALTER TABLE subjects DROP INDEX uq_subjects_branch_sem_code")
        cur.execute("SHOW INDEX FROM subjects WHERE Key_name='uq_subjects_branch_sem_series_code'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE subjects ADD UNIQUE KEY uq_subjects_branch_sem_series_code (branch, semester, series, course_code)")
        cur.execute("SHOW INDEX FROM subjects WHERE Key_name='idx_subjects_branch_sem'")
        if not cur.fetchone():
            cur.execute("CREATE INDEX idx_subjects_branch_sem ON subjects(branch, semester)")
        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def ensure_subject_series_table():
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS subject_series (
                id INT AUTO_INCREMENT PRIMARY KEY,
                series_name VARCHAR(40) NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        for series_name in DEFAULT_SUBJECT_SERIES_OPTIONS:
            cur.execute("INSERT IGNORE INTO subject_series (series_name) VALUES (%s)", (series_name,))
        db.commit()
    except Exception:
        db.rollback()
    finally:
        cur.close()
        db.close()


def fetch_subject_series_options():
    ensure_subject_series_table()
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT series_name FROM subject_series ORDER BY series_name ASC")
        rows = cur.fetchall()
    finally:
        cur.close()
        db.close()
    return [str(r[0]).strip().upper() for r in rows if r and str(r[0]).strip()]


def fetch_subject_master_rows(branch=None, semester_no=None, series_name=None):
    ensure_subject_master_table()
    db = get_db()
    cur = db.cursor(dictionary=True)
    try:
        query = """
            SELECT id, branch, semester, series, subject_name, course_code, subject_type
            FROM subjects
            WHERE 1=1
        """
        params = []
        if semester_no is not None:
            query += " AND semester=%s"
            params.append(semester_no)
        if series_name:
            query += " AND UPPER(series)=%s"
            params.append(str(series_name).strip().upper())
        query += " ORDER BY branch ASC, semester ASC, subject_name ASC, course_code ASC"
        cur.execute(query, tuple(params))
        rows = cur.fetchall()
    finally:
        cur.close()
        db.close()

    if branch:
        rows = [
            r for r in rows
            if same_department(r.get("branch"), branch) or normalize_branch_key(r.get("branch")) == normalize_branch_key(branch)
        ]
    return rows


def find_subject_master_by_code(branch, semester_no, course_code, series_name=None):
    if not branch or semester_no is None or not course_code:
        return None
    rows = fetch_subject_master_rows(branch=branch, semester_no=semester_no, series_name=series_name)
    target = (course_code or "").strip().upper()
    for row in rows:
        if (row.get("course_code") or "").strip().upper() == target:
            return row
    return None


def current_academic_year(today=None):
    today = today or datetime.today().date()
    start_year = today.year if today.month >= 6 else today.year - 1
    return f"{start_year}-{str((start_year + 1) % 100).zfill(2)}"


def marks_grade(score):
    try:
        val = float(score or 0)
    except (TypeError, ValueError):
        return "F"
    if val >= 90:
        return "A+"
    if val >= 80:
        return "A"
    if val >= 70:
        return "B+"
    if val >= 60:
        return "B"
    if val >= 50:
        return "C"
    return "F"


def normalize_branch_key(branch_name):
    raw = (branch_name or "").strip().lower().replace("&", " and ")
    cleaned = re.sub(r"[^a-z0-9]+", " ", raw)
    return re.sub(r"\s+", " ", cleaned).strip()


def branch_code_for_admission(branch_name):
    key = normalize_branch_key(branch_name)
    branch_codes = {
        "computer science and engineering": "CS",
        "computer engineering": "CS",
        "automobile engineering": "AT",
        "automobile": "AT",
        "electronics and communication engineering": "EC",
        "electronic and communication engineering": "EC",
        "electronics and communication": "EC",
        "mechanical engineering": "ME",
        "mechanical": "ME",
    }
    code = branch_codes.get(key)
    if code:
        return code

    parts = [p for p in key.split(" ") if p]
    if len(parts) >= 2:
        return (parts[0][0] + parts[1][0]).upper()
    if len(parts) == 1:
        return parts[0][:2].upper()
    return "SV"


def generate_admission_id(branch_name, academic_year=None):
    ay = (academic_year or current_academic_year()).strip()
    match = re.search(r"(\d{4})", ay)
    start_year = int(match.group(1)) if match else datetime.today().year
    yy = str(start_year)[-2:]
    prefix = f"{branch_code_for_admission(branch_name)}{yy}"
    pattern = f"{prefix}%"

    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT admission_id FROM students WHERE admission_id LIKE %s", (pattern,))
        rows = cur.fetchall() or []
    finally:
        cur.close()
        db.close()

    max_seq = 0
    seq_pattern = re.compile(rf"^{re.escape(prefix)}(\d{{3}})$")
    for row in rows:
        value = row[0] if isinstance(row, (list, tuple)) else row.get("admission_id")
        text = str(value or "").strip().upper()
        seq_match = seq_pattern.match(text)
        if seq_match:
            max_seq = max(max_seq, int(seq_match.group(1)))

    next_seq = max_seq + 1
    if next_seq > 999:
        raise ValueError(f"Admission sequence limit reached for {prefix}")
    return f"{prefix}{next_seq:03d}"


def is_duplicate_admission_id_error(exc):
    msg = str(exc or "").lower()
    return "duplicate entry" in msg and "admission_id" in msg


def insert_student_login_with_retry(
    cur,
    *,
    admission_id,
    student_name,
    branch,
    admission_year,
    mobile,
    password_hash,
    status,
    admission_year_text=None,
    year_sem=None,
    max_attempts=5,
):
    final_id = (admission_id or "").strip().upper()
    if not final_id:
        final_id = generate_admission_id(branch, admission_year_text)

    for _ in range(max_attempts):
        try:
            if year_sem is None:
                cur.execute("""
                    INSERT INTO students
                    (admission_id, student_name, branch, admission_year, mobile, password_hash, status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (final_id, student_name, branch, admission_year, mobile, password_hash, status))
            else:
                cur.execute("""
                    INSERT INTO students
                    (admission_id, student_name, branch, admission_year, year_sem, mobile, password_hash, status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (final_id, student_name, branch, admission_year, year_sem, mobile, password_hash, status))
            return final_id
        except Exception as exc:
            if not is_duplicate_admission_id_error(exc):
                raise
            final_id = generate_admission_id(branch, admission_year_text)

    raise RuntimeError("Could not allocate a unique admission ID after multiple attempts.")


def parse_int_prefix(text):
    if text is None:
        return None
    raw = str(text).strip()
    if not raw:
        return None
    token = ""
    for ch in raw:
        if ch.isdigit():
            token += ch
        else:
            break
    if not token:
        return None
    try:
        return int(token)
    except ValueError:
        return None


def infer_current_sem(admission_year=None, year_sem=None, today=None):
    today = today or datetime.today().date()
    parsed_admission_year = parse_int_prefix(admission_year)
    manual_sem = parse_int_prefix(year_sem)

    if manual_sem and 1 <= manual_sem <= 6:
        return manual_sem
    if not parsed_admission_year:
        return 1

    # Academic year starts in June: 2023-24 => Jun-2023..May-2024.
    current_ay_start = today.year if today.month >= 6 else today.year - 1
    years_elapsed = current_ay_start - parsed_admission_year
    sem = (years_elapsed * 2) + (1 if today.month >= 6 else 2)
    return min(max(sem, 1), 6)


def fee_summary_from_row(row, today=None, due=None, paid=None):
    today = today or datetime.today().date()
    current_sem = infer_current_sem(row.get("admission_year"), row.get("year_sem"), today=today)
    due = due or {}
    paid = paid or {}

    admission_due_total = float(due.get("ADMISSION", 0) or 0)
    tuition_due_total = float(due.get("TUITION", 0) or 0)
    management_due_total = float(due.get("MANAGEMENT", 0) or 0)
    exam_due_total = float(due.get("EXAM", 0) or 0)
    total_due = admission_due_total + tuition_due_total + management_due_total + exam_due_total

    admission_paid = float(paid.get("ADMISSION", 0) or 0)
    tuition_paid = float(paid.get("TUITION", 0) or 0)
    management_paid = float(paid.get("MANAGEMENT", 0) or 0)
    exam_paid = float(paid.get("EXAM", 0) or 0)
    total_paid = admission_paid + tuition_paid + management_paid + exam_paid
    balance = round(total_due - total_paid, 2)

    if total_paid <= 0:
        payment_state = "NOT PAID"
    elif balance <= 0:
        payment_state = "PAID"
    else:
        payment_state = "PENDING"

    result = dict(row)
    result.update({
        "current_sem": current_sem,
        "admission_due_total": round(admission_due_total, 2),
        "tuition_due_total": round(tuition_due_total, 2),
        "management_due_total": round(management_due_total, 2),
        "exam_due_total": round(exam_due_total, 2),
        "total_due": round(total_due, 2),
        "admission_paid": round(admission_paid, 2),
        "tuition_paid": round(tuition_paid, 2),
        "management_paid": round(management_paid, 2),
        "exam_paid": round(exam_paid, 2),
        "total_paid": round(total_paid, 2),
        "balance": balance,
        "payment_state": payment_state,
        "academic_year": current_academic_year(today=today),
    })
    return result


def infer_uploaded_doc_files(admission_id, docs):
    resolved = dict(docs or {})
    try:
        all_files = [
            f for f in os.listdir(UPLOAD_FOLDER)
            if f.startswith(f"{admission_id}_")
        ]
    except Exception:
        return resolved

    all_files.sort(
        key=lambda name: os.path.getmtime(os.path.join(UPLOAD_FOLDER, name)),
        reverse=True
    )

    def pick(tokens):
        for fname in all_files:
            low = fname.lower()
            if any(t in low for t in tokens):
                return fname
        return None

    if not resolved.get("student_photo"):
        resolved["student_photo"] = pick(["_photo_", "photo"])
    if not resolved.get("aadhaar_file"):
        resolved["aadhaar_file"] = pick(["aadhaar"])
    if not resolved.get("caste_file"):
        resolved["caste_file"] = pick(["caste"])
    if not resolved.get("income_file"):
        resolved["income_file"] = pick(["income"])
    if not resolved.get("marks_card_file"):
        resolved["marks_card_file"] = pick(["marks_card", "_marks_", "marks"])

    return resolved


def fetch_student_full_bundle(admission_id):
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT * FROM students WHERE admission_id=%s", (admission_id,))
    student = cur.fetchone()

    cur.execute("""
        SELECT * FROM student_personal_details
        WHERE admission_id=%s
    """, (admission_id,))
    personal = cur.fetchone()

    cur.execute("""
        SELECT * FROM education_details
        WHERE admission_id=%s
        LIMIT 1
    """, (admission_id,))
    education = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM student_documents
        WHERE admission_id=%s
    """, (admission_id,))
    documents = cur.fetchone()

    cur.close()
    db.close()

    docs = {
        "student_photo": (documents or {}).get("student_photo") or (personal or {}).get("photo_file"),
        "aadhaar_file": (documents or {}).get("aadhaar_file") or (personal or {}).get("aadhaar_file"),
        "caste_file": (documents or {}).get("caste_file") or (personal or {}).get("caste_certificate_file"),
        "income_file": (documents or {}).get("income_file") or (personal or {}).get("income_certificate_file"),
        "marks_card_file": (documents or {}).get("marks_card_file") or (personal or {}).get("marks_card_file"),
        "aadhaar_number": (documents or {}).get("aadhaar_number") or (personal or {}).get("aadhaar_number"),
        "caste_rd_number": (documents or {}).get("caste_rd_number") or (personal or {}).get("caste_rd_number"),
        "income_rd_number": (documents or {}).get("income_rd_number") or (personal or {}).get("income_rd_number"),
    }
    docs = infer_uploaded_doc_files(admission_id, docs)
    return student, personal, education, docs


def fetch_student_fee_summary(admission_id, student_row):
    if not student_row:
        return None

    ensure_fee_module_tables()
    db = get_db()
    cur = db.cursor(dictionary=True)
    try:
        current_sem = infer_current_sem(student_row.get("admission_year"), student_row.get("year_sem"))
        ay = current_academic_year()

        cur.execute("""
            SELECT
                admission_fee_due,
                tuition_fee_due,
                management_fee_due,
                exam_fee_due
            FROM fee_structure_master
            WHERE branch=%s AND semester_no=%s AND academic_year=%s
        """, (student_row.get("branch"), current_sem, ay))
        structure = cur.fetchone() or {}

        due = {
            "ADMISSION": float(structure.get("admission_fee_due") or 0),
            "TUITION": float(structure.get("tuition_fee_due") or 0),
            "MANAGEMENT": float(structure.get("management_fee_due") or 0),
            "EXAM": float(structure.get("exam_fee_due") or 0),
        }

        cur.execute("""
            SELECT
                fee_type,
                COALESCE(SUM(amount), 0) AS total_amount
            FROM fee_payments
            WHERE admission_id=%s AND academic_year=%s AND semester_no=%s
            GROUP BY fee_type
        """, (admission_id, ay, current_sem))
        paid_rows = cur.fetchall()
        paid = {"ADMISSION": 0, "TUITION": 0, "MANAGEMENT": 0, "EXAM": 0}
        for pr in paid_rows:
            paid[pr["fee_type"]] = float(pr.get("total_amount") or 0)

        summary = fee_summary_from_row(student_row, due=due, paid=paid)
        return summary
    finally:
        cur.close()
        db.close()


def ensure_student_attendance_table():
    db = get_db()
    cur = db.cursor()
    try:
        # Keep this separate from existing student_attendance (academic module).
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_daily_attendance (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admission_id VARCHAR(50) NOT NULL,
                branch VARCHAR(100) NOT NULL,
                semester_no TINYINT NULL,
                attendance_date DATE NOT NULL,
                subject_name VARCHAR(120) NOT NULL,
                period_no TINYINT NOT NULL DEFAULT 1,
                status ENUM('PRESENT','ABSENT','LATE','LEAVE') NOT NULL,
                remarks VARCHAR(255) NULL,
                marked_by VARCHAR(255) NOT NULL,
                marked_by_name VARCHAR(150) NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_student_daily_attendance_session (admission_id, attendance_date, subject_name, period_no)
            )
        """)
        cur.execute("SHOW INDEX FROM student_daily_attendance WHERE Key_name='idx_student_daily_attendance_date'")
        if not cur.fetchone():
            cur.execute("CREATE INDEX idx_student_daily_attendance_date ON student_daily_attendance(attendance_date)")
        cur.execute("SHOW INDEX FROM student_daily_attendance WHERE Key_name='idx_student_daily_attendance_subject'")
        if not cur.fetchone():
            cur.execute("CREATE INDEX idx_student_daily_attendance_subject ON student_daily_attendance(subject_name)")
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        cur.close()
        db.close()


def normalize_subject_name(subject):
    return re.sub(r"\s+", " ", (subject or "").strip())


def same_department(branch_name, department_name):
    branch_key = normalize_branch_key(branch_name)
    dept_key = normalize_branch_key(department_name)
    if not branch_key or not dept_key:
        return False

    canonical_map = {
        "computer science and engineering": "CS",
        "computer engineering": "CS",
        "computer science": "CS",
        "cs": "CS",
        "cse": "CS",
        "electronics and communication engineering": "EC",
        "electronics and communication": "EC",
        "electronic and communication engineering": "EC",
        "electrical and communication": "EC",
        "ec": "EC",
        "e c": "EC",
        "e and c": "EC",
        "mechanical engineering": "ME",
        "mechanical": "ME",
        "me": "ME",
        "automobile engineering": "AT",
        "automobile": "AT",
        "auto": "AT",
        "at": "AT",
    }

    left = canonical_map.get(branch_key, branch_code_for_admission(branch_key))
    right = canonical_map.get(dept_key, branch_code_for_admission(dept_key))
    if left and right and left == right:
        return True
    if branch_key == dept_key:
        return True
    if branch_key in dept_key or dept_key in branch_key:
        return True
    return False


def parse_iso_date_safe(raw_value, fallback=None):
    fallback = fallback or date.today()
    text = (raw_value or "").strip()
    if not text:
        return fallback
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return fallback


def parse_month_safe(raw_value, fallback=None):
    fallback = fallback or date.today().strftime("%Y-%m")
    text = (raw_value or "").strip()
    if not text:
        text = fallback
    try:
        month_date = datetime.strptime(text + "-01", "%Y-%m-%d").date()
        return month_date.strftime("%Y-%m"), month_date
    except ValueError:
        month_date = datetime.strptime(fallback + "-01", "%Y-%m-%d").date()
        return fallback, month_date


def fetch_attendance_register(scope, selected_date_obj, month_value, subject, period_no, branch, semester, q):
    month_value, month_date = parse_month_safe(month_value, selected_date_obj.strftime("%Y-%m"))
    month_start = month_date.replace(day=1)
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1, day=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1, day=1)
    days_in_month = (month_end - month_start).days
    day_numbers = list(range(1, days_in_month + 1))

    if selected_date_obj < month_start or selected_date_obj >= month_end:
        selected_date_obj = month_start

    semester_no_filter = parse_int_prefix(semester)

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT DISTINCT branch FROM students WHERE status='ACTIVE' ORDER BY branch ASC")
    branches = [row["branch"] for row in cur.fetchall() if row.get("branch")]

    student_query = """
        SELECT
            s.admission_id,
            s.student_name,
            s.branch,
            s.year_sem,
            s.admission_year,
            COALESCE(NULLIF(s.college_reg_no, ''), NULLIF(spd.register_number, ''), s.admission_id) AS reg_no,
            COALESCE(sd.student_photo, spd.photo_file, '') AS photo_file
        FROM students s
        LEFT JOIN student_personal_details spd ON spd.admission_id = s.admission_id
        LEFT JOIN student_documents sd ON sd.admission_id = s.admission_id
        WHERE s.status='ACTIVE'
    """
    params = []
    if branch and not scope.get("is_staff"):
        student_query += " AND s.branch=%s"
        params.append(branch)
    if q:
        student_query += " AND (s.student_name LIKE %s OR s.admission_id LIKE %s OR s.college_reg_no LIKE %s OR spd.register_number LIKE %s)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])
    student_query += " ORDER BY s.student_name ASC"
    cur.execute(student_query, tuple(params))

    students = []
    for row in cur.fetchall():
        if scope["is_staff"] and not same_department(row.get("branch"), scope.get("department")):
            continue
        sem_no = parse_int_prefix(row.get("year_sem"))
        if sem_no is None:
            sem_no = infer_current_sem(row.get("admission_year"), row.get("year_sem"))
        if semester_no_filter and sem_no != semester_no_filter:
            continue
        row["semester_no"] = sem_no
        row["semester_display"] = f"SEM {sem_no}"
        students.append(row)

    attendance_by_student_day = {}
    attendance_by_student_selected_day = {}
    if students and subject:
        ids = [row["admission_id"] for row in students]
        placeholders = ",".join(["%s"] * len(ids))
        cur.execute(f"""
            SELECT
                admission_id,
                attendance_date,
                status,
                COALESCE(remarks, '') AS remarks,
                COALESCE(marked_by_name, marked_by, '') AS marked_by
            FROM student_daily_attendance
            WHERE attendance_date >= %s
              AND attendance_date < %s
              AND subject_name=%s
              AND period_no=%s
              AND admission_id IN ({placeholders})
        """, tuple([month_start.isoformat(), month_end.isoformat(), subject, period_no] + ids))
        for row in cur.fetchall():
            sid = row["admission_id"]
            day_no = row["attendance_date"].day
            attendance_by_student_day.setdefault(sid, {})[day_no] = row["status"]
            if row["attendance_date"] == selected_date_obj:
                attendance_by_student_selected_day[sid] = row

    cur.close()
    db.close()

    today_present = 0
    today_absent = 0
    today_marked = 0
    today_leave = 0
    day_wise_stats_map = {
        day: {"day": day, "present": 0, "absent": 0, "leave": 0, "marked": 0, "unmarked": 0, "present_pct": 0}
        for day in day_numbers
    }
    for row in students:
        sid = row["admission_id"]
        day_map = attendance_by_student_day.get(sid, {})
        existing = attendance_by_student_selected_day.get(sid, {})
        row["daily_status"] = day_map
        row["quick_status"] = existing.get("status", "PRESENT")
        row["quick_remarks"] = existing.get("remarks", "")
        row["marked_by"] = existing.get("marked_by", "-")
        row["total_marked_days"] = len(day_map)
        row["present_days"] = sum(1 for s in day_map.values() if s in ("PRESENT", "LATE"))
        row["absent_days"] = sum(1 for s in day_map.values() if s == "ABSENT")
        row["leave_days"] = sum(1 for s in day_map.values() if s == "LEAVE")
        row["attendance_pct"] = round((row["present_days"] / row["total_marked_days"]) * 100, 1) if row["total_marked_days"] else 0

        for day in day_numbers:
            daily_status = day_map.get(day)
            stat = day_wise_stats_map[day]
            if daily_status in ("PRESENT", "LATE"):
                stat["present"] += 1
                stat["marked"] += 1
            elif daily_status == "ABSENT":
                stat["absent"] += 1
                stat["marked"] += 1
            elif daily_status == "LEAVE":
                stat["leave"] += 1
                stat["marked"] += 1
            else:
                stat["unmarked"] += 1

        selected_status = existing.get("status")
        if selected_status:
            today_marked += 1
            if selected_status in ("PRESENT", "LATE"):
                today_present += 1
            if selected_status == "ABSENT":
                today_absent += 1
            if selected_status == "LEAVE":
                today_leave += 1

    day_wise_stats = []
    for day in day_numbers:
        stat = day_wise_stats_map[day]
        stat["present_pct"] = round((stat["present"] / stat["marked"]) * 100, 1) if stat["marked"] else 0
        day_wise_stats.append(stat)

    today_present_pct = round((today_present / today_marked) * 100, 1) if today_marked else 0
    return {
        "students": students,
        "branches": branches,
        "day_numbers": day_numbers,
        "selected_day": selected_date_obj.day,
        "selected_date": selected_date_obj.isoformat(),
        "month": month_value,
        "month_label": month_start.strftime("%B %Y"),
        "summary": {
            "total_students": len(students),
            "marked_count": today_marked,
            "present_count": today_present,
            "absent_count": today_absent,
            "leave_count": today_leave,
            "unmarked_count": max(len(students) - today_marked, 0),
            "present_pct": today_present_pct,
        },
        "day_wise_stats": day_wise_stats,
    }


def _send_attendance_export(filename_base, header, rows):
    if not OPENPYXL_AVAILABLE:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(header)
        writer.writerows(rows)
        csv_data = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        output.close()
        return send_file(csv_data, mimetype="text/csv", as_attachment=True, download_name=f"{filename_base}.csv")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    for col, label in enumerate(header, start=1):
        sheet.cell(row=1, column=col, value=label)
    for ridx, row in enumerate(rows, start=2):
        for cidx, val in enumerate(row, start=1):
            sheet.cell(row=ridx, column=cidx, value=val)
    for col in range(1, len(header) + 1):
        letter = get_column_letter(col)
        max_len = len(str(header[col - 1]))
        for ridx in range(2, sheet.max_row + 1):
            v = sheet.cell(row=ridx, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        sheet.column_dimensions[letter].width = min(40, max(10, max_len + 2))
    excel_data = io.BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)
    return send_file(
        excel_data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{filename_base}.xlsx",
    )


def normalize_series_text(raw):
    text = (raw or "").strip().upper()
    text = re.sub(r"\s+", "", text)
    if not text:
        return ""
    if not re.match(r"^[A-Z0-9-]{2,30}$", text):
        return ""
    return text


def branch_name_from_code(code):
    code_map = {
        "CS": "Computer Science and Engineering",
        "EC": "Electronics and Communication Engineering",
        "ME": "Mechanical Engineering",
        "AT": "Automobile Engineering",
    }
    return code_map.get((code or "").strip().upper(), "")


def branch_name_from_reg_no(reg_no):
    token = (reg_no or "").strip().upper()
    match = re.search(r"(CS|EC|ME|AT)", token)
    if not match:
        return ""
    return branch_name_from_code(match.group(1))


def parse_result_pdf_rows(pdf_path, source_file):
    if PdfReader is None:
        raise RuntimeError("PDF import requires pypdf or PyPDF2. Install with: python -m pip install pypdf")
    reader = PdfReader(pdf_path)
    exam_session = ""
    branch_name = ""
    rows = []
    current_reg = ""
    current_name = ""
    current_row_refs = []
    current_summary = {}

    exam_re = re.compile(r"RESULT LEDGER - DIPLOMA EXAMINATION\s+(.+?)\s*\(\s*\d+[A-Z]?\s*\)", re.IGNORECASE)
    programme_re = re.compile(r"Programme\s*:\s*([A-Z]{2})\s*-\s*(.+)", re.IGNORECASE)
    student_re = re.compile(r"^\d+\s+([0-9A-Z]{8,})\s+(.+?)\s+\[")
    subject_re = re.compile(
        r"^(\d+)\s+(\d+)\s+([0-9A-Z]+)\s*:\s*(.+?)\s+([0-9A-Z-]+)\s*/\s*([0-9A-Z-]+)\s*/\s*([0-9A-Z-]+)\s+([A-Z]+)\s+([0-9A-Z.+-]+)\s+([A-Z0-9+.-]+)$"
    )
    credit_applied_re = re.compile(r"^Credit Applied\s+([0-9 ]+)\s+Credit Earned\s*:\s*([0-9]+)", re.IGNORECASE)
    credit_earned_re = re.compile(r"^Credit Earned\s+([0-9 ]+)\s+CGPA\s*:\s*(.+)$", re.IGNORECASE)
    percent_re = re.compile(r"% Conversion\s*:\s*(.+)$", re.IGNORECASE)
    sgpa_re = re.compile(r"^SGPA\s*\((?:Atempts|Attempts)\)\s+(.+?)\s+Results\s*:\s*(.+)$", re.IGNORECASE)

    def apply_summary_to_refs():
        if not current_row_refs:
            return
        for ref in current_row_refs:
            ref.update({
                "final_result": current_summary.get("final_result", ""),
                "cgpa": current_summary.get("cgpa", ""),
                "percentage": current_summary.get("percentage", ""),
                "credit_earned_total": current_summary.get("credit_earned_total", ""),
                "credit_applied_s1": current_summary.get("credit_applied_s1", ""),
                "credit_applied_s2": current_summary.get("credit_applied_s2", ""),
                "credit_applied_s3": current_summary.get("credit_applied_s3", ""),
                "credit_applied_s4": current_summary.get("credit_applied_s4", ""),
                "credit_applied_s5": current_summary.get("credit_applied_s5", ""),
                "credit_applied_s6": current_summary.get("credit_applied_s6", ""),
                "credit_earned_s1": current_summary.get("credit_earned_s1", ""),
                "credit_earned_s2": current_summary.get("credit_earned_s2", ""),
                "credit_earned_s3": current_summary.get("credit_earned_s3", ""),
                "credit_earned_s4": current_summary.get("credit_earned_s4", ""),
                "credit_earned_s5": current_summary.get("credit_earned_s5", ""),
                "credit_earned_s6": current_summary.get("credit_earned_s6", ""),
                "sgpa_s1": current_summary.get("sgpa_s1", ""),
                "sgpa_s2": current_summary.get("sgpa_s2", ""),
                "sgpa_s3": current_summary.get("sgpa_s3", ""),
                "sgpa_s4": current_summary.get("sgpa_s4", ""),
                "sgpa_s5": current_summary.get("sgpa_s5", ""),
                "sgpa_s6": current_summary.get("sgpa_s6", ""),
                "attempts_s1": current_summary.get("attempts_s1", ""),
                "attempts_s2": current_summary.get("attempts_s2", ""),
                "attempts_s3": current_summary.get("attempts_s3", ""),
                "attempts_s4": current_summary.get("attempts_s4", ""),
                "attempts_s5": current_summary.get("attempts_s5", ""),
                "attempts_s6": current_summary.get("attempts_s6", ""),
            })

    for page in reader.pages:
        page_text = (page.extract_text() or "").encode("ascii", "ignore").decode()
        for raw_line in page_text.splitlines():
            line = re.sub(r"\s+", " ", raw_line).strip()
            if not line:
                continue

            if not exam_session:
                m_exam = exam_re.search(line)
                if m_exam:
                    exam_session = m_exam.group(1).strip()

            if not branch_name:
                m_programme = programme_re.search(line)
                if m_programme:
                    branch_name = branch_name_from_code(m_programme.group(1).strip()) or m_programme.group(2).strip().title()

            m_student = student_re.match(line)
            if m_student:
                apply_summary_to_refs()
                current_reg = m_student.group(1).strip().upper()
                current_name = m_student.group(2).strip()
                current_row_refs = []
                current_summary = {}
                continue

            m_subject = subject_re.match(line)
            if m_subject and current_reg:
                sem_no = parse_int_prefix(m_subject.group(1))
                subject_code = m_subject.group(3).strip().upper()
                subject_name = m_subject.group(4).strip()
                row_branch = branch_name or branch_name_from_reg_no(current_reg) or "Unknown"
                rows.append({
                    "register_number": current_reg,
                    "student_name": current_name,
                    "branch": row_branch,
                    "semester_no": sem_no,
                    "subject_code": subject_code,
                    "subject_name": subject_name,
                    "ia_marks": m_subject.group(5).strip().upper(),
                    "theory_marks": m_subject.group(6).strip().upper(),
                    "practical_marks": m_subject.group(7).strip().upper(),
                    "result_status": m_subject.group(8).strip().upper(),
                    "credit": m_subject.group(9).strip().upper(),
                    "grade": m_subject.group(10).strip().upper(),
                    "final_result": "",
                    "cgpa": "",
                    "percentage": "",
                    "credit_earned_total": "",
                    "credit_applied_s1": "",
                    "credit_applied_s2": "",
                    "credit_applied_s3": "",
                    "credit_applied_s4": "",
                    "credit_applied_s5": "",
                    "credit_applied_s6": "",
                    "credit_earned_s1": "",
                    "credit_earned_s2": "",
                    "credit_earned_s3": "",
                    "credit_earned_s4": "",
                    "credit_earned_s5": "",
                    "credit_earned_s6": "",
                    "sgpa_s1": "",
                    "sgpa_s2": "",
                    "sgpa_s3": "",
                    "sgpa_s4": "",
                    "sgpa_s5": "",
                    "sgpa_s6": "",
                    "attempts_s1": "",
                    "attempts_s2": "",
                    "attempts_s3": "",
                    "attempts_s4": "",
                    "attempts_s5": "",
                    "attempts_s6": "",
                    "exam_session": exam_session,
                    "source_file": source_file,
                })
                current_row_refs.append(rows[-1])

            def map_sem_values(prefix, values):
                for idx in range(6):
                    key = f"{prefix}_s{idx + 1}"
                    current_summary[key] = values[idx] if idx < len(values) else current_summary.get(key, "")

            m_ca = credit_applied_re.match(line)
            if m_ca:
                nums = re.findall(r"\d+", m_ca.group(1))
                map_sem_values("credit_applied", nums)
                current_summary["credit_earned_total"] = m_ca.group(2).strip()
                continue

            m_ce = credit_earned_re.match(line)
            if m_ce:
                nums = re.findall(r"\d+", m_ce.group(1))
                map_sem_values("credit_earned", nums)
                cgpa_text = m_ce.group(2).strip()
                current_summary["cgpa"] = "" if "credit(s) pending" in cgpa_text.lower() else cgpa_text
                continue

            m_percent = percent_re.search(line)
            if m_percent:
                pct = m_percent.group(1).strip()
                current_summary["percentage"] = "" if "not applicable" in pct.lower() else pct
                continue

            m_sgpa = sgpa_re.match(line)
            if m_sgpa:
                sgpa_chunk = m_sgpa.group(1).strip()
                result_label = m_sgpa.group(2).strip()
                current_summary["final_result"] = result_label
                sgpa_pairs = re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*\((\d+)\)", sgpa_chunk)
                sgpa_vals = [pair[0] for pair in sgpa_pairs]
                attempt_vals = [pair[1] for pair in sgpa_pairs]
                map_sem_values("sgpa", sgpa_vals)
                map_sem_values("attempts", attempt_vals)
                apply_summary_to_refs()
                continue

    apply_summary_to_refs()

    return rows


def _norm_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def parse_result_excel_rows(excel_path, source_file):
    if load_workbook is None:
        raise RuntimeError("Excel import requires openpyxl. Install it with: python -m pip install openpyxl")
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    header_row = all_rows[0]
    idx_map = {}
    for idx, col in enumerate(header_row):
        key = _norm_header(col)
        if key:
            idx_map[key] = idx

    def pick_index(*keys):
        for key in keys:
            if key in idx_map:
                return idx_map[key]
        return None

    reg_idx = pick_index("regno", "registerno", "registernumber")
    if reg_idx is None:
        return []

    name_idx = pick_index("studentname", "name")
    branch_idx = pick_index("branch", "department")
    sem_idx = pick_index("sem", "semester", "semno", "semesterno")
    sub_code_idx = pick_index("subjectcode", "qpcode", "code")
    sub_name_idx = pick_index("subjectname", "subject")
    ia_idx = pick_index("ia", "iamarks")
    th_idx = pick_index("theory", "tr", "theorymarks")
    pr_idx = pick_index("practical", "pr", "prmarks")
    result_idx = pick_index("result", "status")
    credit_idx = pick_index("credit", "credits")
    grade_idx = pick_index("grade")
    exam_idx = pick_index("examsession", "exam", "exammonthyear")
    final_result_idx = pick_index("finalresult", "overallresult", "resultclass")
    cgpa_idx = pick_index("cgpa")
    pct_idx = pick_index("percentage", "percent")
    ce_total_idx = pick_index("creditearnedtotal", "totalcreditearned")
    ca1_idx = pick_index("creditapplieds1", "creditapplied1")
    ca2_idx = pick_index("creditapplieds2", "creditapplied2")
    ca3_idx = pick_index("creditapplieds3", "creditapplied3")
    ca4_idx = pick_index("creditapplieds4", "creditapplied4")
    ca5_idx = pick_index("creditapplieds5", "creditapplied5")
    ca6_idx = pick_index("creditapplieds6", "creditapplied6")
    ce1_idx = pick_index("creditearneds1", "creditearned1")
    ce2_idx = pick_index("creditearneds2", "creditearned2")
    ce3_idx = pick_index("creditearneds3", "creditearned3")
    ce4_idx = pick_index("creditearneds4", "creditearned4")
    ce5_idx = pick_index("creditearneds5", "creditearned5")
    ce6_idx = pick_index("creditearneds6", "creditearned6")
    sg1_idx = pick_index("sgpas1", "sgpa1")
    sg2_idx = pick_index("sgpas2", "sgpa2")
    sg3_idx = pick_index("sgpas3", "sgpa3")
    sg4_idx = pick_index("sgpas4", "sgpa4")
    sg5_idx = pick_index("sgpas5", "sgpa5")
    sg6_idx = pick_index("sgpas6", "sgpa6")
    at1_idx = pick_index("attemptss1", "attempts1")
    at2_idx = pick_index("attemptss2", "attempts2")
    at3_idx = pick_index("attemptss3", "attempts3")
    at4_idx = pick_index("attemptss4", "attempts4")
    at5_idx = pick_index("attemptss5", "attempts5")
    at6_idx = pick_index("attemptss6", "attempts6")

    parsed = []
    for row in all_rows[1:]:
        reg_no = str(row[reg_idx] or "").strip().upper() if reg_idx is not None else ""
        if not reg_no:
            continue
        student_name = str(row[name_idx] or "").strip() if name_idx is not None else ""
        branch_val = str(row[branch_idx] or "").strip() if branch_idx is not None else ""
        sem_raw = str(row[sem_idx] or "").strip() if sem_idx is not None else ""
        semester_no = parse_int_prefix(sem_raw)
        parsed.append({
            "register_number": reg_no,
            "student_name": student_name or "-",
            "branch": branch_val or branch_name_from_reg_no(reg_no) or "Unknown",
            "semester_no": semester_no,
            "subject_code": str(row[sub_code_idx] or "").strip().upper() if sub_code_idx is not None else "",
            "subject_name": str(row[sub_name_idx] or "").strip() if sub_name_idx is not None else "",
            "ia_marks": str(row[ia_idx] or "").strip().upper() if ia_idx is not None else "",
            "theory_marks": str(row[th_idx] or "").strip().upper() if th_idx is not None else "",
            "practical_marks": str(row[pr_idx] or "").strip().upper() if pr_idx is not None else "",
            "result_status": str(row[result_idx] or "").strip().upper() if result_idx is not None else "",
            "credit": str(row[credit_idx] or "").strip().upper() if credit_idx is not None else "",
            "grade": str(row[grade_idx] or "").strip().upper() if grade_idx is not None else "",
            "final_result": str(row[final_result_idx] or "").strip() if final_result_idx is not None else "",
            "cgpa": str(row[cgpa_idx] or "").strip() if cgpa_idx is not None else "",
            "percentage": str(row[pct_idx] or "").strip() if pct_idx is not None else "",
            "credit_earned_total": str(row[ce_total_idx] or "").strip() if ce_total_idx is not None else "",
            "credit_applied_s1": str(row[ca1_idx] or "").strip() if ca1_idx is not None else "",
            "credit_applied_s2": str(row[ca2_idx] or "").strip() if ca2_idx is not None else "",
            "credit_applied_s3": str(row[ca3_idx] or "").strip() if ca3_idx is not None else "",
            "credit_applied_s4": str(row[ca4_idx] or "").strip() if ca4_idx is not None else "",
            "credit_applied_s5": str(row[ca5_idx] or "").strip() if ca5_idx is not None else "",
            "credit_applied_s6": str(row[ca6_idx] or "").strip() if ca6_idx is not None else "",
            "credit_earned_s1": str(row[ce1_idx] or "").strip() if ce1_idx is not None else "",
            "credit_earned_s2": str(row[ce2_idx] or "").strip() if ce2_idx is not None else "",
            "credit_earned_s3": str(row[ce3_idx] or "").strip() if ce3_idx is not None else "",
            "credit_earned_s4": str(row[ce4_idx] or "").strip() if ce4_idx is not None else "",
            "credit_earned_s5": str(row[ce5_idx] or "").strip() if ce5_idx is not None else "",
            "credit_earned_s6": str(row[ce6_idx] or "").strip() if ce6_idx is not None else "",
            "sgpa_s1": str(row[sg1_idx] or "").strip() if sg1_idx is not None else "",
            "sgpa_s2": str(row[sg2_idx] or "").strip() if sg2_idx is not None else "",
            "sgpa_s3": str(row[sg3_idx] or "").strip() if sg3_idx is not None else "",
            "sgpa_s4": str(row[sg4_idx] or "").strip() if sg4_idx is not None else "",
            "sgpa_s5": str(row[sg5_idx] or "").strip() if sg5_idx is not None else "",
            "sgpa_s6": str(row[sg6_idx] or "").strip() if sg6_idx is not None else "",
            "attempts_s1": str(row[at1_idx] or "").strip() if at1_idx is not None else "",
            "attempts_s2": str(row[at2_idx] or "").strip() if at2_idx is not None else "",
            "attempts_s3": str(row[at3_idx] or "").strip() if at3_idx is not None else "",
            "attempts_s4": str(row[at4_idx] or "").strip() if at4_idx is not None else "",
            "attempts_s5": str(row[at5_idx] or "").strip() if at5_idx is not None else "",
            "attempts_s6": str(row[at6_idx] or "").strip() if at6_idx is not None else "",
            "exam_session": str(row[exam_idx] or "").strip() if exam_idx is not None else "",
            "source_file": source_file,
        })

    return parsed


def parse_result_file_rows(file_path, source_file, ext):
    ext = (ext or "").strip().lower()
    if ext == "pdf":
        return parse_result_pdf_rows(file_path, source_file)
    if ext in {"xlsx", "xlsm", "xltx", "xltm"}:
        return parse_result_excel_rows(file_path, source_file)
    raise RuntimeError("Unsupported result file type")


def verify_password_compat(stored_hash, raw_password):
    stored = str(stored_hash or "").strip()
    if not stored:
        return False

    try:
        if check_password_hash(stored, raw_password):
            return True
    except Exception:
        pass

    return stored == hashlib.sha256((raw_password or "").encode()).hexdigest()


@app.route("/")
def root():
    return redirect("/home")

# =========================
# HOME PAGE (INDEX)
# =========================

@app.route("/home")
def home():
    return render_template("index.html")


# =========================
# LOGIN / AUTH
# =========================
@app.route("/login")
def login():
    return render_template("login_choice.html")


@app.route("/login/student", methods=["GET", "POST"])
def login_student():
    error = ""
    if request.method == "POST":
        if not validate_csrf():
            error = "Session expired. Refresh and try again."
            return render_template("login_student.html", error=error)

        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        raw_password = request.form.get("password", "")
        identity = f"{client_ip()}:{username.lower()}"
        if is_auth_rate_limited("student_login", identity):
            error = "Too many failed attempts. Try again in 5 minutes."
            return render_template("login_student.html", error=error)

        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("""
            SELECT * FROM students
            WHERE UPPER(admission_id)=UPPER(%s)
        """, (username,))
        student = cur.fetchone()
        cur.close()
        db.close()

        if not student or not verify_password_compat(student.get("password_hash"), password):
            error = "Invalid student credentials."
            return render_template("login_student.html", error=error)

        clear_auth_failures("student_login", identity)
        if student["status"] == "ACTIVE":
            session.clear()
            session["student"] = student["admission_id"]
            return redirect("/student")
        elif student["status"] == "REJECTED":
            return render_template("student_rejected.html", reason=student.get("rejection_reason"))
        return render_template("student_pending.html")

    return render_template("login_student.html", error=error)


@app.route("/login/staff", methods=["GET", "POST"])
def login_staff_admin():
    ensure_staff_auth_tables()
    error = ""
    if request.method == "POST":
        if not validate_csrf():
            error = "Session expired. Refresh and try again."
            return render_template("login_staff_admin.html", error=error)

        login_type = request.form.get("login_type", "staff").strip().lower()
        if login_type not in {"admin", "hod", "staff", "management_staff"}:
            login_type = "staff"
        login_id = request.form.get("login_id", "").strip()
        raw_password = request.form.get("password", "")
        identity = f"{client_ip()}:{login_type}:{login_id.lower()}"
        if is_auth_rate_limited("staff_admin_login", identity):
            error = "Too many failed attempts. Try again in 5 minutes."
            return render_template("login_staff_admin.html", error=error)

        db = get_db()
        cur = db.cursor(dictionary=True)

        if login_type == "admin":
            cur.execute(
                "SELECT * FROM admins WHERE LOWER(username)=LOWER(%s)",
                (login_id,)
            )
            admin = cur.fetchone()
            if admin and verify_password(admin.get("password_hash", ""), raw_password):
                clear_auth_failures("staff_admin_login", identity)
                session.clear()
                session["admin"] = admin["username"]
                cur.close()
                db.close()
                return redirect("/admin")
            record_auth_failure("staff_admin_login", identity)
            error = "Invalid admin credentials."
        else:
            cur.execute("""
                SELECT * FROM staff_accounts
                WHERE LOWER(email)=LOWER(%s) AND is_verified=1
            """, (login_id,))
            staff = cur.fetchone()
            if staff and verify_password(staff.get("password_hash", ""), raw_password):
                designation = (staff.get("designation") or "").strip().lower()
                department = (staff.get("department") or "").strip().lower()
                is_hod = "hod" in designation or "head of department" in designation
                is_management = "management" in department

                if login_type == "hod" and not is_hod:
                    record_auth_failure("staff_admin_login", identity)
                    error = "This account is not a HOD account."
                    cur.close()
                    db.close()
                    return render_template("login_staff_admin.html", error=error)
                if login_type == "management_staff" and not is_management:
                    record_auth_failure("staff_admin_login", identity)
                    error = "This account is not a Management Staff account."
                    cur.close()
                    db.close()
                    return render_template("login_staff_admin.html", error=error)
                if login_type == "staff" and (is_hod or is_management):
                    record_auth_failure("staff_admin_login", identity)
                    error = "Use HOD or Management Staff login type for this account."
                    cur.close()
                    db.close()
                    return render_template("login_staff_admin.html", error=error)

                clear_auth_failures("staff_admin_login", identity)
                session.clear()
                session["admin"] = staff["email"]
                session["staff_id"] = staff["id"]
                cur.close()
                db.close()
                return redirect("/admin")
            record_auth_failure("staff_admin_login", identity)
            error = "Invalid staff credentials."

        cur.close()
        db.close()

    return render_template("login_staff_admin.html", error=error)


@app.route("/staff/register", methods=["GET", "POST"])
def staff_register():
    ensure_staff_auth_tables()
    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT department FROM employee_details ORDER BY department ASC")
    department_rows = cur.fetchall()
    departments = [row["department"] for row in department_rows if row.get("department")]
    departments = [d for d in departments if (d or "").strip().lower() != "management"]
    if "Management Department" not in departments:
        departments.append("Management Department")
    departments.sort()
    cur.close()
    db.close()

    error = ""
    message = ""
    if request.method == "POST":
        if not validate_csrf():
            error = "Session expired. Refresh and try again."
            return render_template("staff_register.html", error=error, message=message, departments=departments)

        employee_name = request.form.get("employee_name", "").strip()
        department = request.form.get("department", "").strip()
        designation = request.form.get("designation", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not employee_name:
            error = "Employee name is required."
        elif not department:
            error = "Department is required."
        elif not designation:
            error = "Designation is required."
        elif not email.endswith("@gmail.com"):
            error = "Use a valid Gmail address."
        elif len(password) < 6:
            error = "Password must be at least 6 characters."
        elif password != confirm_password:
            error = "Passwords do not match."
        else:
            db = get_db()
            cur = db.cursor(dictionary=True)
            cur.execute("SELECT id FROM staff_accounts WHERE LOWER(email)=LOWER(%s)", (email,))
            existing = cur.fetchone()
            cur.close()
            db.close()

            if existing:
                error = "Staff account already exists for this Gmail."
            else:
                pending = {
                    "employee_id": None,
                    "employee_name": employee_name,
                    "department": department,
                    "designation": designation,
                    "email": email,
                    "password_hash": hash_password(password),
                }
                otp = start_otp_flow("staff_register", email, pending)
                sent, msg = send_otp_email(email, "SVP Staff Registration OTP", otp)
                if not sent:
                    error = msg
                else:
                    message = "OTP sent to your Gmail. Verify to complete registration."
                    return redirect(url_for("staff_register_verify", email=email, msg=message))

    return render_template("staff_register.html", error=error, message=message, departments=departments)


@app.route("/staff/register/verify", methods=["GET", "POST"])
def staff_register_verify():
    ensure_staff_auth_tables()
    email = request.args.get("email", "").strip().lower() or request.form.get("email", "").strip().lower()
    error = ""
    message = request.args.get("msg", "")
    if request.method == "POST":
        if not validate_csrf():
            error = "Session expired. Refresh and try again."
            return render_template("staff_register_verify.html", email=email, error=error, message=message)

        otp = request.form.get("otp", "").strip()
        ok, msg = verify_otp_flow("staff_register", email, otp)
        if not ok:
            error = msg
        else:
            flow = session.get("otp_flow") or {}
            extra = flow.get("extra") or {}
            db = get_db()
            cur = db.cursor()
            cur.execute("""
                INSERT INTO staff_accounts (employee_id, employee_name, department, designation, email, password_hash, is_verified)
                VALUES (%s,%s,%s,%s,%s,%s,1)
            """, (
                extra.get("employee_id"),
                extra.get("employee_name"),
                extra.get("department"),
                extra.get("designation"),
                extra.get("email"),
                extra.get("password_hash"),
            ))
            db.commit()
            cur.close()
            db.close()
            session.pop("otp_flow", None)
            return redirect(url_for("login_staff_admin"))

    return render_template("staff_register_verify.html", email=email, error=error, message=message)


@app.route("/forgot-password/student", methods=["GET", "POST"])
def forgot_password_student():
    error = ""
    message = ""
    if request.method == "POST":
        if not validate_csrf():
            error = "Session expired. Refresh and try again."
            return render_template("forgot_password_student.html", error=error, message=message)

        email = request.form.get("email", "").strip().lower()
        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("""
            SELECT spd.admission_id
            FROM student_personal_details spd
            JOIN students s ON s.admission_id = spd.admission_id
            WHERE LOWER(spd.student_email)=LOWER(%s)
            LIMIT 1
        """, (email,))
        row = cur.fetchone()
        cur.close()
        db.close()
        if not row:
            error = "Student email not found."
        else:
            otp = start_otp_flow("student_forgot_password", email, {"admission_id": row["admission_id"]})
            sent, msg = send_otp_email(email, "SVP Student Password Reset OTP", otp)
            if not sent:
                error = msg
            else:
                return redirect(url_for("forgot_password_student_verify"))

    return render_template("forgot_password_student.html", error=error, message=message)


@app.route("/forgot-password/student/verify", methods=["GET", "POST"])
def forgot_password_student_verify():
    email = request.args.get("email", "").strip().lower() or request.form.get("email", "").strip().lower()
    flow = session.get("otp_flow") or {}
    if (not email) and flow.get("flow_key") == "student_forgot_password":
        email = str(flow.get("email") or "").strip().lower()
    error = ""
    if request.method == "POST":
        if not validate_csrf():
            error = "Session expired. Refresh and try again."
            return render_template("forgot_password_verify.html", email=email, error=error, role_name="Student")

        otp = request.form.get("otp", "").strip()
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        if len(new_password) < 6:
            error = "Password must be at least 6 characters."
        elif new_password != confirm_password:
            error = "Passwords do not match."
        elif not email:
            error = "OTP session not found. Please request OTP again."
        else:
            ok, msg = verify_otp_flow("student_forgot_password", email, otp)
            if not ok:
                error = msg
            else:
                flow = session.get("otp_flow") or {}
                admission_id = (flow.get("extra") or {}).get("admission_id")
                db = get_db()
                cur = db.cursor()
                cur.execute(
                    "UPDATE students SET password_hash=%s WHERE admission_id=%s",
                    (generate_password_hash(new_password), admission_id)
                )
                db.commit()
                cur.close()
                db.close()
                session.pop("otp_flow", None)
                return redirect(url_for("login_student"))

    return render_template("forgot_password_verify.html", email=email, error=error, role_name="Student")


@app.route("/forgot-password/staff", methods=["GET", "POST"])
def forgot_password_staff():
    ensure_staff_auth_tables()
    error = ""
    if request.method == "POST":
        if not validate_csrf():
            error = "Session expired. Refresh and try again."
            return render_template("forgot_password_staff.html", error=error)

        email = request.form.get("email", "").strip().lower()
        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT id FROM staff_accounts WHERE LOWER(email)=LOWER(%s)", (email,))
        staff = cur.fetchone()
        cur.close()
        db.close()
        if not staff:
            error = "Staff email not found."
        else:
            otp = start_otp_flow("staff_forgot_password", email, {"staff_id": staff["id"]})
            sent, msg = send_otp_email(email, "SVP Staff Password Reset OTP", otp)
            if not sent:
                error = msg
            else:
                return redirect(url_for("forgot_password_staff_verify", email=email))

    return render_template("forgot_password_staff.html", error=error)


@app.route("/forgot-password/staff/verify", methods=["GET", "POST"])
def forgot_password_staff_verify():
    ensure_staff_auth_tables()
    email = request.args.get("email", "").strip().lower() or request.form.get("email", "").strip().lower()
    error = ""
    if request.method == "POST":
        if not validate_csrf():
            error = "Session expired. Refresh and try again."
            return render_template("forgot_password_verify.html", email=email, error=error, role_name="Staff")

        otp = request.form.get("otp", "").strip()
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        if len(new_password) < 6:
            error = "Password must be at least 6 characters."
        elif new_password != confirm_password:
            error = "Passwords do not match."
        else:
            ok, msg = verify_otp_flow("staff_forgot_password", email, otp)
            if not ok:
                error = msg
            else:
                flow = session.get("otp_flow") or {}
                staff_id = (flow.get("extra") or {}).get("staff_id")
                db = get_db()
                cur = db.cursor()
                cur.execute(
                    "UPDATE staff_accounts SET password_hash=%s WHERE id=%s",
                    (hash_password(new_password), staff_id)
                )
                db.commit()
                cur.close()
                db.close()
                session.pop("otp_flow", None)
                return redirect(url_for("login_staff_admin"))

    return render_template("forgot_password_verify.html", email=email, error=error, role_name="Staff")


# =========================
# LOGOUT
# =========================
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# =========================
# ADMIN DASHBOARD
# =========================
@app.route("/admin")
def admin_dashboard():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    is_management_staff = scope["is_staff"] and "management" in (scope.get("department") or "").strip().lower()
    is_hod_staff = scope["is_staff"] and (
        "hod" in (scope.get("designation") or "").strip().lower()
        or "head of department" in (scope.get("designation") or "").strip().lower()
    )
    return render_template(
        "admin_dashboard.html",
        is_staff=scope["is_staff"],
        staff_name=scope.get("staff_name"),
        staff_department=scope["department"],
        is_management_staff=is_management_staff,
        is_hod_staff=is_hod_staff,
        can_manage_fees=can_edit_fees(scope),
        can_manage_syllabus=can_upload_syllabus(scope)
    )


@app.route("/admin/subject-master", methods=["GET", "POST"])
def admin_subject_master():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    is_hod_staff = scope["is_staff"] and (
        "hod" in (scope.get("designation") or "").strip().lower()
        or "head of department" in (scope.get("designation") or "").strip().lower()
    )
    if scope["is_staff"] and not is_hod_staff:
        return "Forbidden: Only Admin or HOD can access Subject Master.", 403

    ensure_subject_master_table()
    ensure_subject_series_table()
    msg = request.args.get("msg", "").strip()
    branch_filter = (request.args.get("branch") or "").strip()
    if scope["is_staff"]:
        branch_filter = scope.get("department") or branch_filter
    semester_filter = (request.args.get("semester") or "").strip()
    series_filter = (request.args.get("series") or "").strip().upper()

    if request.method == "POST":
        if not validate_csrf():
            return redirect(url_for("admin_subject_master", msg="Session expired. Refresh and try again."))

        action = (request.form.get("action") or "").strip().lower()
        subject_id = parse_int_prefix(request.form.get("subject_id"))
        branch = (request.form.get("branch") or "").strip()
        semester_no = parse_int_prefix(request.form.get("semester"))
        series_name = (request.form.get("series") or "").strip().upper()
        custom_series = (request.form.get("custom_series") or "").strip().upper()
        subject_name = (request.form.get("subject_name") or "").strip()
        course_code = (request.form.get("course_code") or "").strip().upper()
        subject_type = (request.form.get("subject_type") or "").strip().upper()

        if action not in {"add", "update", "delete", "add_series"}:
            return redirect(url_for("admin_subject_master", msg="Invalid action."))

        db = get_db()
        cur = db.cursor()
        try:
            if action == "add_series":
                if scope["is_staff"]:
                    return "Forbidden: Only Admin can add new series.", 403
                new_series = custom_series or series_name
                if not new_series:
                    return redirect(url_for("admin_subject_master", msg="Series name is required."))
                cur.execute("INSERT IGNORE INTO subject_series (series_name) VALUES (%s)", (new_series,))
                db.commit()
                return redirect(url_for("admin_subject_master", msg="Series added successfully.", series=new_series))

            if action == "delete":
                if scope["is_staff"]:
                    return "Forbidden: Only Admin can delete subjects.", 403
                if not subject_id:
                    return redirect(url_for("admin_subject_master", msg="Subject id is required for delete."))
                cur.execute("DELETE FROM subjects WHERE id=%s", (subject_id,))
                db.commit()
                return redirect(url_for("admin_subject_master", msg="Subject deleted."))

            if branch not in BRANCH_OPTIONS:
                return redirect(url_for("admin_subject_master", msg="Please select a valid branch."))
            if scope["is_staff"] and not same_department(branch, scope.get("department")):
                return "Forbidden: HOD can add/edit only their department subjects.", 403
            if semester_no is None or semester_no < 1 or semester_no > 6:
                return redirect(url_for("admin_subject_master", msg="Semester must be between 1 and 6."))
            series_value = custom_series or series_name
            if not series_value:
                return redirect(url_for("admin_subject_master", msg="Series is required."))
            cur.execute("INSERT IGNORE INTO subject_series (series_name) VALUES (%s)", (series_value,))
            if not subject_name or not course_code:
                return redirect(url_for("admin_subject_master", msg="Subject name and course code are required."))
            if subject_type not in SUBJECT_TYPE_OPTIONS:
                return redirect(url_for("admin_subject_master", msg="Invalid subject type."))

            if action == "add":
                cur.execute("""
                    INSERT INTO subjects (branch, semester, series, subject_name, course_code, subject_type)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (branch, semester_no, series_value, subject_name, course_code, subject_type))
                db.commit()
                return redirect(url_for("admin_subject_master", msg="Subject added successfully.", branch=branch, semester=semester_no, series=series_value))

            if not subject_id:
                return redirect(url_for("admin_subject_master", msg="Subject id is required for update."))
            cur.execute("""
                UPDATE subjects
                SET branch=%s, semester=%s, series=%s, subject_name=%s, course_code=%s, subject_type=%s
                WHERE id=%s
            """, (branch, semester_no, series_value, subject_name, course_code, subject_type, subject_id))
            db.commit()
            return redirect(url_for("admin_subject_master", msg="Subject updated successfully.", branch=branch, semester=semester_no, series=series_value))
        except Exception as exc:
            db.rollback()
            return redirect(url_for("admin_subject_master", msg=f"Unable to save subject: {exc}"))
        finally:
            cur.close()
            db.close()

    subjects = fetch_subject_master_rows(
        branch=branch_filter or None,
        semester_no=parse_int_prefix(semester_filter) if semester_filter else None,
        series_name=series_filter or None
    )
    series_options = fetch_subject_series_options()
    return render_template(
        "admin_subject_master.html",
        msg=msg,
        subjects=subjects,
        branch_options=BRANCH_OPTIONS,
        subject_type_options=SUBJECT_TYPE_OPTIONS,
        series_options=series_options,
        can_manage_series=not scope["is_staff"],
        can_delete_subject=not scope["is_staff"],
        branch_filter=branch_filter,
        semester_filter=semester_filter,
        series_filter=series_filter
    )


@app.route("/admin/api/subjects")
def admin_api_subjects():
    scope = get_access_scope()
    if not scope["allowed"]:
        return jsonify({"subjects": []}), 403

    ensure_subject_master_table()
    branch = (request.args.get("branch") or "").strip()
    if scope["is_staff"]:
        branch = scope.get("department") or branch
    semester_no = parse_int_prefix(request.args.get("semester"))
    series_name = (request.args.get("series") or "").strip().upper()

    if not branch or semester_no is None or semester_no < 1 or semester_no > 6 or not series_name:
        return jsonify({"subjects": []})

    rows = fetch_subject_master_rows(branch=branch, semester_no=semester_no, series_name=series_name)
    payload = []
    for row in rows:
        payload.append({
            "id": row.get("id"),
            "branch": row.get("branch"),
            "semester": row.get("semester"),
            "series": row.get("series"),
            "subject_name": row.get("subject_name"),
            "course_code": row.get("course_code"),
            "subject_type": row.get("subject_type"),
            "label": f"{row.get('subject_name')} ({row.get('course_code')})",
        })
    return jsonify({"subjects": payload})


@app.route("/admin/tools/backfill-personal-details", methods=["GET", "POST"])
def backfill_personal_details():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"]:
        return "Forbidden: Only admin can run backfill.", 403

    if request.method == "POST" and not validate_csrf():
        return "Session expired. Refresh and try again.", 400

    db = get_db()
    cur = db.cursor()
    try:
        # Fill missing personal mobile numbers from students table.
        cur.execute("""
            UPDATE student_personal_details spd
            JOIN students s ON s.admission_id = spd.admission_id
            SET spd.student_mobile = s.mobile
            WHERE (spd.student_mobile IS NULL OR TRIM(spd.student_mobile) = '')
              AND s.mobile IS NOT NULL
              AND TRIM(s.mobile) <> ''
        """)
        mobile_backfilled = cur.rowcount

        # Normalize existing emails.
        cur.execute("""
            UPDATE student_personal_details
            SET student_email = LOWER(TRIM(student_email))
            WHERE student_email IS NOT NULL
              AND TRIM(student_email) <> ''
              AND student_email <> LOWER(TRIM(student_email))
        """)
        emails_normalized = cur.rowcount

        # Fill missing personal document fields from student_documents.
        cur.execute("""
            UPDATE student_personal_details spd
            JOIN student_documents sd ON sd.admission_id = spd.admission_id
            SET
                spd.aadhaar_number = COALESCE(NULLIF(spd.aadhaar_number, ''), NULLIF(sd.aadhaar_number, '')),
                spd.caste_rd_number = COALESCE(NULLIF(spd.caste_rd_number, ''), NULLIF(sd.caste_rd_number, '')),
                spd.income_rd_number = COALESCE(NULLIF(spd.income_rd_number, ''), NULLIF(sd.income_rd_number, '')),
                spd.photo_file = COALESCE(NULLIF(spd.photo_file, ''), NULLIF(sd.student_photo, '')),
                spd.caste_certificate_file = COALESCE(NULLIF(spd.caste_certificate_file, ''), NULLIF(sd.caste_file, '')),
                spd.income_certificate_file = COALESCE(NULLIF(spd.income_certificate_file, ''), NULLIF(sd.income_file, '')),
                spd.marks_card_file = COALESCE(NULLIF(spd.marks_card_file, ''), NULLIF(sd.marks_card_file, ''))
        """)
        documents_backfilled = cur.rowcount

        cur.execute("""
            SELECT COUNT(*)
            FROM student_personal_details
            WHERE student_mobile IS NULL OR TRIM(student_mobile) = ''
        """)
        missing_mobile = cur.fetchone()[0]

        cur.execute("""
            SELECT COUNT(*)
            FROM student_personal_details
            WHERE student_email IS NULL OR TRIM(student_email) = ''
        """)
        missing_email = cur.fetchone()[0]

        cur.execute("""
            SELECT COUNT(*)
            FROM student_personal_details
            WHERE aadhaar_number IS NULL OR TRIM(aadhaar_number) = ''
               OR caste_rd_number IS NULL OR TRIM(caste_rd_number) = ''
               OR income_rd_number IS NULL OR TRIM(income_rd_number) = ''
               OR marks_card_file IS NULL OR TRIM(marks_card_file) = ''
        """)
        missing_docs = cur.fetchone()[0]

        db.commit()
    except Exception as exc:
        db.rollback()
        return f"Backfill failed: {exc}", 500
    finally:
        cur.close()
        db.close()

    return (
        "Backfill complete.<br>"
        f"Mobiles backfilled: {mobile_backfilled}<br>"
        f"Emails normalized: {emails_normalized}<br>"
        f"Document fields backfilled: {documents_backfilled}<br>"
        f"Remaining missing mobile: {missing_mobile}<br>"
        f"Remaining missing email: {missing_email}<br>"
        f"Remaining rows with missing Aadhaar/Caste/Income/Marks fields: {missing_docs}"
    )


@app.route("/admin/tools/backfill-education-details", methods=["GET", "POST"])
def backfill_education_details():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"]:
        return "Forbidden: Only admin can run backfill.", 403

    if request.method == "POST" and not validate_csrf():
        return "Session expired. Refresh and try again.", 400

    db = get_db()
    cur = db.cursor(dictionary=True)
    try:
        cur.execute("SHOW COLUMNS FROM student_personal_details")
        spd_cols = {row["Field"] for row in cur.fetchall()}

        def pick_expr(candidates):
            for c in candidates:
                if c in spd_cols:
                    return f"NULLIF(TRIM(spd.{c}), '')"
            return "NULL"

        def pick_int_expr(candidates):
            for c in candidates:
                if c in spd_cols:
                    return f"NULLIF(TRIM(spd.{c}), '')"
            return "NULL"

        qualifying_exam_expr = pick_expr(["qualifying_exam"])
        register_number_expr = pick_expr(["register_number"])
        year_of_passing_expr = pick_expr(["year_of_passing"])

        total_max_expr = pick_int_expr(["total_max_marks", "total_marks"])
        total_obt_expr = pick_int_expr(["total_marks_obtained", "marks_obtained"])
        sci_max_expr = pick_int_expr(["science_max_marks"])
        sci_obt_expr = pick_int_expr(["science_marks_obtained", "science_marks"])
        mat_max_expr = pick_int_expr(["maths_max_marks"])
        mat_obt_expr = pick_int_expr(["maths_marks_obtained", "maths_marks"])

        # 1) Insert missing education rows.
        insert_sql = f"""
            INSERT INTO education_details (
                admission_id,
                qualifying_exam,
                register_number,
                year_of_passing,
                total_max_marks,
                total_marks_obtained,
                science_max_marks,
                science_marks_obtained,
                maths_max_marks,
                maths_marks_obtained
            )
            SELECT
                spd.admission_id,
                {qualifying_exam_expr},
                {register_number_expr},
                {year_of_passing_expr},
                COALESCE({total_max_expr}, 0),
                COALESCE({total_obt_expr}, 0),
                COALESCE({sci_max_expr}, 0),
                COALESCE({sci_obt_expr}, 0),
                COALESCE({mat_max_expr}, 0),
                COALESCE({mat_obt_expr}, 0)
            FROM student_personal_details spd
            LEFT JOIN education_details ed ON ed.admission_id = spd.admission_id
            WHERE ed.admission_id IS NULL
        """
        cur.execute(insert_sql)
        inserted_rows = cur.rowcount

        # 2) Fill missing values in existing education rows.
        update_sql = f"""
            UPDATE education_details ed
            JOIN student_personal_details spd ON spd.admission_id = ed.admission_id
            SET
                ed.qualifying_exam = COALESCE(NULLIF(TRIM(ed.qualifying_exam), ''), {qualifying_exam_expr}),
                ed.register_number = COALESCE(NULLIF(TRIM(ed.register_number), ''), {register_number_expr}),
                ed.year_of_passing = COALESCE(NULLIF(TRIM(ed.year_of_passing), ''), {year_of_passing_expr}),
                ed.total_max_marks = COALESCE(ed.total_max_marks, {total_max_expr}, 0),
                ed.total_marks_obtained = COALESCE(ed.total_marks_obtained, {total_obt_expr}, 0),
                ed.science_max_marks = COALESCE(ed.science_max_marks, {sci_max_expr}, 0),
                ed.science_marks_obtained = COALESCE(ed.science_marks_obtained, {sci_obt_expr}, 0),
                ed.maths_max_marks = COALESCE(ed.maths_max_marks, {mat_max_expr}, 0),
                ed.maths_marks_obtained = COALESCE(ed.maths_marks_obtained, {mat_obt_expr}, 0)
        """
        cur.execute(update_sql)
        updated_rows = cur.rowcount

        cur.execute("""
            SELECT COUNT(*)
            FROM education_details
            WHERE total_max_marks IS NULL
               OR total_marks_obtained IS NULL
               OR science_marks_obtained IS NULL
               OR maths_marks_obtained IS NULL
        """)
        remaining_missing_marks = cur.fetchone()["COUNT(*)"]

        db.commit()
    except Exception as exc:
        db.rollback()
        return f"Education backfill failed: {exc}", 500
    finally:
        cur.close()
        db.close()

    return (
        "Education backfill complete.<br>"
        f"Inserted missing education rows: {inserted_rows}<br>"
        f"Updated existing education rows: {updated_rows}<br>"
        f"Remaining rows with missing marks fields: {remaining_missing_marks}"
    )


@app.route("/admin/staff-accounts")
def admin_staff_accounts():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"]:
        return "Forbidden: Only admin can view staff login accounts.", 403

    ensure_staff_auth_tables()

    q = request.args.get("q", "").strip()
    department = request.args.get("department", "").strip()
    designation = request.args.get("designation", "").strip()
    verified = request.args.get("verified", "").strip()
    msg = request.args.get("msg", "").strip()

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT
            id,
            COALESCE(employee_name, '-') AS employee_name,
            COALESCE(department, '-') AS department,
            COALESCE(designation, '-') AS designation,
            email,
            is_verified,
            created_at
        FROM staff_accounts
        WHERE 1=1
    """
    params = []

    if q:
        query += """
            AND (
                employee_name LIKE %s
                OR email LIKE %s
            )
        """
        like_q = f"%{q}%"
        params.extend([like_q, like_q])

    if department:
        query += " AND department=%s"
        params.append(department)

    if designation:
        query += " AND designation=%s"
        params.append(designation)

    if verified == "yes":
        query += " AND is_verified=1"
    elif verified == "no":
        query += " AND is_verified=0"

    query += " ORDER BY created_at DESC, id DESC"
    cur.execute(query, tuple(params))
    accounts = cur.fetchall()

    cur.execute("SELECT DISTINCT department FROM staff_accounts WHERE department IS NOT NULL AND department<>'' ORDER BY department ASC")
    departments = [row["department"] for row in cur.fetchall() if row.get("department")]

    cur.execute("SELECT DISTINCT designation FROM staff_accounts WHERE designation IS NOT NULL AND designation<>'' ORDER BY designation ASC")
    designations = [row["designation"] for row in cur.fetchall() if row.get("designation")]

    cur.close()
    db.close()

    return render_template(
        "admin_staff_accounts.html",
        accounts=accounts,
        departments=departments,
        designations=designations,
        msg=msg,
        filters={
            "q": q,
            "department": department,
            "designation": designation,
            "verified": verified,
        }
    )


@app.route("/admin/staff-accounts/edit/<int:account_id>", methods=["GET", "POST"])
def admin_edit_staff_account(account_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"]:
        return "Forbidden: Only admin can edit staff login accounts.", 403

    ensure_staff_auth_tables()

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT id, employee_name, department, designation, email, is_verified
        FROM staff_accounts
        WHERE id=%s
    """, (account_id,))
    account = cur.fetchone()

    if not account:
        cur.close()
        db.close()
        return "Staff account not found", 404

    if request.method == "POST":
        employee_name = request.form.get("employee_name", "").strip()
        department = request.form.get("department", "").strip()
        designation = request.form.get("designation", "").strip()
        email = request.form.get("email", "").strip().lower()
        verified_raw = request.form.get("is_verified", "").strip().lower()
        is_verified = 1 if verified_raw in {"1", "yes", "true"} else 0

        if not email:
            cur.close()
            db.close()
            return "Email is required", 400

        cur.execute("SELECT id FROM staff_accounts WHERE LOWER(email)=LOWER(%s) AND id<>%s", (email, account_id))
        email_conflict = cur.fetchone()
        if email_conflict:
            cur.close()
            db.close()
            return "Another staff account already uses this email", 400

        update_cur = db.cursor()
        update_cur.execute("""
            UPDATE staff_accounts
            SET employee_name=%s,
                department=%s,
                designation=%s,
                email=%s,
                is_verified=%s
            WHERE id=%s
        """, (employee_name or None, department or None, designation or None, email, is_verified, account_id))
        db.commit()
        update_cur.close()
        cur.close()
        db.close()
        return redirect(url_for("admin_staff_accounts", msg="Staff account updated successfully"))

    cur.close()
    db.close()
    return render_template("admin_edit_staff_account.html", account=account)


@app.route("/admin/staff-accounts/delete/<int:account_id>", methods=["POST"])
def admin_delete_staff_account(account_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"]:
        return "Forbidden: Only admin can delete staff login accounts.", 403

    if session.get("staff_id") and int(session.get("staff_id")) == int(account_id):
        return redirect(url_for("admin_staff_accounts", msg="You cannot delete your own active account"))

    ensure_staff_auth_tables()
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT id FROM staff_accounts WHERE id=%s", (account_id,))
    exists = cur.fetchone()
    if not exists:
        cur.close()
        db.close()
        return redirect(url_for("admin_staff_accounts", msg="Staff account not found"))

    cur.execute("DELETE FROM staff_accounts WHERE id=%s", (account_id,))
    db.commit()
    cur.close()
    db.close()
    return redirect(url_for("admin_staff_accounts", msg="Staff account deleted successfully"))


@app.route("/admin/results", methods=["GET", "POST"])
def admin_results():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_student_results_table()
    can_manage_results = can_edit_fees(scope)
    is_management_staff = scope["is_staff"] and "management" in (scope.get("department") or "").strip().lower()
    msg = request.args.get("msg", "").strip()

    if request.method == "POST":
        if not can_manage_results:
            return "Forbidden: Only Admin, HOD, or Management staff can import results.", 403
        import_type = (request.form.get("import_type", "auto") or "auto").strip().lower()
        result_file = request.files.get("result_file")
        if not result_file or not result_file.filename:
            return redirect(url_for("admin_results", msg="Please select an Excel/PDF file"))

        filename = secure_filename(result_file.filename)
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in {"pdf", "xlsx", "xlsm", "xltx", "xltm"}:
            return redirect(url_for("admin_results", msg="Allowed file types: PDF, XLSX"))

        if import_type == "pdf" and ext != "pdf":
            return redirect(url_for("admin_results", msg="Selected import type is PDF but uploaded file is not PDF"))
        if import_type == "excel" and ext == "pdf":
            return redirect(url_for("admin_results", msg="Selected import type is Excel but uploaded file is PDF"))

        save_name = f"results_{datetime.now().strftime('%Y%m%d%H%M%S')}_{random.randint(1000,9999)}_{filename}"
        save_path = os.path.join(RESULT_UPLOAD_FOLDER, save_name)
        result_file.save(save_path)

        try:
            rows = parse_result_file_rows(save_path, save_name, ext)
        except Exception as exc:
            return redirect(url_for("admin_results", msg=f"Failed to parse file: {exc}"))

        if not rows:
            return redirect(url_for("admin_results", msg="No result rows found in uploaded file"))

        if scope["is_staff"] and not is_management_staff:
            dept = (scope.get("department") or "").strip().lower()
            rows = [r for r in rows if (r.get("branch") or "").strip().lower() == dept]
            if not rows:
                return redirect(url_for("admin_results", msg="No rows matched your department in this file"))

        db = get_db()
        cur = db.cursor()
        try:
            cur.executemany("""
                INSERT INTO student_results (
                    register_number, student_name, branch, semester_no,
                    subject_code, subject_name, ia_marks, theory_marks, practical_marks,
                    result_status, credit, grade,
                    final_result, cgpa, percentage, credit_earned_total,
                    credit_applied_s1, credit_applied_s2, credit_applied_s3, credit_applied_s4, credit_applied_s5, credit_applied_s6,
                    credit_earned_s1, credit_earned_s2, credit_earned_s3, credit_earned_s4, credit_earned_s5, credit_earned_s6,
                    sgpa_s1, sgpa_s2, sgpa_s3, sgpa_s4, sgpa_s5, sgpa_s6,
                    attempts_s1, attempts_s2, attempts_s3, attempts_s4, attempts_s5, attempts_s6,
                    exam_session, source_file
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, [(
                r.get("register_number"),
                r.get("student_name"),
                r.get("branch"),
                r.get("semester_no"),
                r.get("subject_code"),
                r.get("subject_name"),
                r.get("ia_marks"),
                r.get("theory_marks"),
                r.get("practical_marks"),
                r.get("result_status"),
                r.get("credit"),
                r.get("grade"),
                r.get("final_result"),
                r.get("cgpa"),
                r.get("percentage"),
                r.get("credit_earned_total"),
                r.get("credit_applied_s1"),
                r.get("credit_applied_s2"),
                r.get("credit_applied_s3"),
                r.get("credit_applied_s4"),
                r.get("credit_applied_s5"),
                r.get("credit_applied_s6"),
                r.get("credit_earned_s1"),
                r.get("credit_earned_s2"),
                r.get("credit_earned_s3"),
                r.get("credit_earned_s4"),
                r.get("credit_earned_s5"),
                r.get("credit_earned_s6"),
                r.get("sgpa_s1"),
                r.get("sgpa_s2"),
                r.get("sgpa_s3"),
                r.get("sgpa_s4"),
                r.get("sgpa_s5"),
                r.get("sgpa_s6"),
                r.get("attempts_s1"),
                r.get("attempts_s2"),
                r.get("attempts_s3"),
                r.get("attempts_s4"),
                r.get("attempts_s5"),
                r.get("attempts_s6"),
                r.get("exam_session"),
                r.get("source_file"),
            ) for r in rows])
            db.commit()
        finally:
            cur.close()
            db.close()

        return redirect(url_for("admin_results", msg=f"Imported {len(rows)} result rows successfully"))

    q_reg_no = request.args.get("reg_no", "").strip().upper()
    branch = request.args.get("branch", "").strip()
    sem = request.args.get("sem", "").strip()
    sort_by = request.args.get("sort_by", "recent").strip()
    if scope["is_staff"] and not is_management_staff:
        branch = scope["department"]

    order_map = {
        "recent": "MAX(sr.imported_at) DESC, sr.register_number ASC",
        "branch": "MAX(sr.branch) ASC, MAX(sr.semester_no) ASC, sr.register_number ASC",
        "sem": "MAX(sr.semester_no) ASC, MAX(sr.branch) ASC, sr.register_number ASC",
        "reg_no": "sr.register_number ASC",
    }
    order_clause = order_map.get(sort_by, order_map["recent"])

    db = get_db()
    cur = db.cursor(dictionary=True)
    query = """
        SELECT
            sr.register_number,
            MAX(sr.student_name) AS student_name,
            MAX(sr.branch) AS branch,
            MAX(sr.semester_no) AS semester_no,
            MAX(sr.final_result) AS final_result,
            MAX(sr.percentage) AS percentage,
            MAX(sr.cgpa) AS cgpa
        FROM student_results sr
        WHERE 1=1
    """
    params = []

    if q_reg_no:
        query += " AND UPPER(sr.register_number) LIKE UPPER(%s)"
        params.append(f"%{q_reg_no}%")
    if branch:
        query += " AND sr.branch=%s"
        params.append(branch)
    sem_no = parse_int_prefix(sem)
    if sem_no is not None and 1 <= sem_no <= 6:
        query += " AND sr.semester_no=%s"
        params.append(sem_no)

    query += " GROUP BY sr.register_number"
    query += f" ORDER BY {order_clause}"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    cur.execute("""
        SELECT
            source_file,
            COUNT(*) AS row_count,
            MAX(imported_at) AS imported_at
        FROM student_results
        WHERE source_file IS NOT NULL AND source_file<>''
        GROUP BY source_file
        ORDER BY MAX(imported_at) DESC
    """)
    imports = cur.fetchall()

    if scope["is_staff"] and not is_management_staff:
        dept = (scope.get("department") or "").strip().lower()
        filtered_imports = []
        for imp in imports:
            cur.execute("""
                SELECT COUNT(*) AS cnt
                FROM student_results
                WHERE source_file=%s AND LOWER(branch)=%s
            """, (imp.get("source_file"), dept))
            c = cur.fetchone() or {}
            if int(c.get("cnt") or 0) > 0:
                filtered_imports.append(imp)
        imports = filtered_imports

    cur.execute("SELECT DISTINCT branch FROM student_results ORDER BY branch ASC")
    branches = [r["branch"] for r in cur.fetchall() if r.get("branch")]
    cur.close()
    db.close()

    return render_template(
        "admin_results.html",
        rows=rows,
        imports=imports,
        branches=branches,
        is_staff=scope["is_staff"],
        is_management_staff=is_management_staff,
        staff_department=scope["department"],
        can_manage_results=can_manage_results,
        msg=msg,
        filters={
            "reg_no": q_reg_no,
            "branch": branch,
            "sem": sem,
            "sort_by": sort_by,
        }
    )


@app.route("/admin/results/pdf")
def admin_results_pdf():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_student_results_table()
    is_management_staff = scope["is_staff"] and "management" in (scope.get("department") or "").strip().lower()
    q_reg_no = request.args.get("reg_no", "").strip().upper()
    branch = request.args.get("branch", "").strip()
    sem = request.args.get("sem", "").strip()
    sort_by = request.args.get("sort_by", "recent").strip()
    if scope["is_staff"] and not is_management_staff:
        branch = scope["department"]

    order_map = {
        "recent": "MAX(sr.imported_at) DESC, sr.register_number ASC",
        "branch": "MAX(sr.branch) ASC, MAX(sr.semester_no) ASC, sr.register_number ASC",
        "sem": "MAX(sr.semester_no) ASC, MAX(sr.branch) ASC, sr.register_number ASC",
        "reg_no": "sr.register_number ASC",
    }
    order_clause = order_map.get(sort_by, order_map["recent"])

    db = get_db()
    cur = db.cursor(dictionary=True)
    query = """
        SELECT
            sr.register_number,
            MAX(sr.student_name) AS student_name,
            MAX(sr.branch) AS branch,
            MAX(sr.semester_no) AS semester_no,
            MAX(sr.final_result) AS final_result,
            MAX(sr.percentage) AS percentage,
            MAX(sr.cgpa) AS cgpa
        FROM student_results sr
        WHERE 1=1
    """
    params = []
    if q_reg_no:
        query += " AND UPPER(sr.register_number) LIKE UPPER(%s)"
        params.append(f"%{q_reg_no}%")
    if branch:
        query += " AND sr.branch=%s"
        params.append(branch)
    sem_no = parse_int_prefix(sem)
    if sem_no is not None and 1 <= sem_no <= 6:
        query += " AND sr.semester_no=%s"
        params.append(sem_no)
    query += " GROUP BY sr.register_number"
    query += f" ORDER BY {order_clause}"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    db.close()

    filter_bits = []
    if q_reg_no:
        filter_bits.append(f"Reg No: {q_reg_no}")
    if branch:
        filter_bits.append(f"Branch: {branch}")
    if sem_no is not None and 1 <= sem_no <= 6:
        filter_bits.append(f"Sem: {sem_no}")
    filter_text = " | ".join(filter_bits)

    pdf = generate_results_summary_pdf(rows, filter_text=filter_text)
    return send_file(pdf, as_attachment=True)


@app.route("/admin/results/import/delete", methods=["POST"])
def admin_results_delete_import():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if not can_edit_fees(scope):
        return "Forbidden: Only Admin, HOD, or Management staff can delete imported result files.", 403

    source_file = (request.form.get("source_file") or "").strip()
    if not source_file:
        return redirect(url_for("admin_results", msg="Invalid source file"))

    db = get_db()
    cur = db.cursor(dictionary=True)
    try:
        if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower():
            cur.execute("""
                SELECT COUNT(*) AS cnt
                FROM student_results
                WHERE source_file=%s AND LOWER(branch)<>%s
            """, (source_file, (scope.get("department") or "").strip().lower()))
            out_scope = cur.fetchone() or {}
            if int(out_scope.get("cnt") or 0) > 0:
                return "Forbidden: You can delete only your department import.", 403

        cur.execute("DELETE FROM student_results WHERE source_file=%s", (source_file,))
        db.commit()
    finally:
        cur.close()
        db.close()

    file_path = os.path.join(RESULT_UPLOAD_FOLDER, source_file)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception:
            pass

    return redirect(url_for("admin_results", msg="Imported file records deleted successfully"))


@app.route("/admin/results/import/replace", methods=["POST"])
def admin_results_replace_import():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if not can_edit_fees(scope):
        return "Forbidden: Only Admin, HOD, or Management staff can replace imported result files.", 403

    source_file = (request.form.get("source_file") or "").strip()
    replace_file = request.files.get("replace_file")
    if not source_file or not replace_file or not replace_file.filename:
        return redirect(url_for("admin_results", msg="Source and replacement file are required"))

    old_ext = source_file.rsplit(".", 1)[-1].lower() if "." in source_file else ""
    new_name = secure_filename(replace_file.filename)
    new_ext = new_name.rsplit(".", 1)[-1].lower() if "." in new_name else ""
    if old_ext != new_ext:
        return redirect(url_for("admin_results", msg="Replacement file must have same extension as original import"))
    if old_ext not in {"pdf", "xlsx", "xlsm", "xltx", "xltm"}:
        return redirect(url_for("admin_results", msg="Unsupported import file type"))

    db = get_db()
    cur = db.cursor(dictionary=True)
    try:
        if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower():
            cur.execute("""
                SELECT COUNT(*) AS cnt
                FROM student_results
                WHERE source_file=%s AND LOWER(branch)<>%s
            """, (source_file, (scope.get("department") or "").strip().lower()))
            out_scope = cur.fetchone() or {}
            if int(out_scope.get("cnt") or 0) > 0:
                return "Forbidden: You can replace only your department import.", 403
    finally:
        cur.close()
        db.close()

    save_path = os.path.join(RESULT_UPLOAD_FOLDER, source_file)
    replace_file.save(save_path)

    try:
        rows = parse_result_file_rows(save_path, source_file, old_ext)
    except Exception as exc:
        return redirect(url_for("admin_results", msg=f"Failed to parse replacement file: {exc}"))
    if not rows:
        return redirect(url_for("admin_results", msg="No result rows found in replacement file"))

    if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower():
        dept = (scope.get("department") or "").strip().lower()
        rows = [r for r in rows if (r.get("branch") or "").strip().lower() == dept]
        if not rows:
            return redirect(url_for("admin_results", msg="No rows matched your department in replacement file"))

    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("DELETE FROM student_results WHERE source_file=%s", (source_file,))
        cur.executemany("""
            INSERT INTO student_results (
                register_number, student_name, branch, semester_no,
                subject_code, subject_name, ia_marks, theory_marks, practical_marks,
                result_status, credit, grade,
                final_result, cgpa, percentage, credit_earned_total,
                credit_applied_s1, credit_applied_s2, credit_applied_s3, credit_applied_s4, credit_applied_s5, credit_applied_s6,
                credit_earned_s1, credit_earned_s2, credit_earned_s3, credit_earned_s4, credit_earned_s5, credit_earned_s6,
                sgpa_s1, sgpa_s2, sgpa_s3, sgpa_s4, sgpa_s5, sgpa_s6,
                attempts_s1, attempts_s2, attempts_s3, attempts_s4, attempts_s5, attempts_s6,
                exam_session, source_file
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            r.get("register_number"),
            r.get("student_name"),
            r.get("branch"),
            r.get("semester_no"),
            r.get("subject_code"),
            r.get("subject_name"),
            r.get("ia_marks"),
            r.get("theory_marks"),
            r.get("practical_marks"),
            r.get("result_status"),
            r.get("credit"),
            r.get("grade"),
            r.get("final_result"),
            r.get("cgpa"),
            r.get("percentage"),
            r.get("credit_earned_total"),
            r.get("credit_applied_s1"),
            r.get("credit_applied_s2"),
            r.get("credit_applied_s3"),
            r.get("credit_applied_s4"),
            r.get("credit_applied_s5"),
            r.get("credit_applied_s6"),
            r.get("credit_earned_s1"),
            r.get("credit_earned_s2"),
            r.get("credit_earned_s3"),
            r.get("credit_earned_s4"),
            r.get("credit_earned_s5"),
            r.get("credit_earned_s6"),
            r.get("sgpa_s1"),
            r.get("sgpa_s2"),
            r.get("sgpa_s3"),
            r.get("sgpa_s4"),
            r.get("sgpa_s5"),
            r.get("sgpa_s6"),
            r.get("attempts_s1"),
            r.get("attempts_s2"),
            r.get("attempts_s3"),
            r.get("attempts_s4"),
            r.get("attempts_s5"),
            r.get("attempts_s6"),
            r.get("exam_session"),
            source_file,
        ) for r in rows])
        db.commit()
    finally:
        cur.close()
        db.close()

    return redirect(url_for("admin_results", msg=f"Replaced import successfully. Rows: {len(rows)}"))


@app.route("/admin/results/student/<reg_no>")
def admin_result_student(reg_no):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_student_results_table()
    is_management_staff = scope["is_staff"] and "management" in (scope.get("department") or "").strip().lower()

    target_reg = (reg_no or "").strip().upper()
    if not target_reg:
        return redirect(url_for("admin_results"))

    db = get_db()
    cur = db.cursor(dictionary=True)
    if scope["is_staff"] and not is_management_staff:
        cur.execute("""
            SELECT register_number FROM student_results
            WHERE UPPER(register_number)=UPPER(%s) AND branch=%s
            LIMIT 1
        """, (target_reg, scope.get("department")))
        allowed_row = cur.fetchone()
        if not allowed_row:
            cur.close()
            db.close()
            return "Forbidden: You can view only your department results.", 403

    cur.execute("""
        SELECT
            register_number,
            MAX(student_name) AS student_name,
            MAX(branch) AS branch,
            MAX(semester_no) AS semester_no,
            MAX(final_result) AS final_result,
            MAX(percentage) AS percentage,
            MAX(cgpa) AS cgpa,
            MAX(credit_earned_total) AS credit_earned_total,
            MAX(sgpa_s1) AS sgpa_s1, MAX(sgpa_s2) AS sgpa_s2, MAX(sgpa_s3) AS sgpa_s3,
            MAX(sgpa_s4) AS sgpa_s4, MAX(sgpa_s5) AS sgpa_s5, MAX(sgpa_s6) AS sgpa_s6,
            MAX(attempts_s1) AS attempts_s1, MAX(attempts_s2) AS attempts_s2, MAX(attempts_s3) AS attempts_s3,
            MAX(attempts_s4) AS attempts_s4, MAX(attempts_s5) AS attempts_s5, MAX(attempts_s6) AS attempts_s6,
            MAX(credit_applied_s1) AS credit_applied_s1, MAX(credit_applied_s2) AS credit_applied_s2, MAX(credit_applied_s3) AS credit_applied_s3,
            MAX(credit_applied_s4) AS credit_applied_s4, MAX(credit_applied_s5) AS credit_applied_s5, MAX(credit_applied_s6) AS credit_applied_s6,
            MAX(credit_earned_s1) AS credit_earned_s1, MAX(credit_earned_s2) AS credit_earned_s2, MAX(credit_earned_s3) AS credit_earned_s3,
            MAX(credit_earned_s4) AS credit_earned_s4, MAX(credit_earned_s5) AS credit_earned_s5, MAX(credit_earned_s6) AS credit_earned_s6,
            MAX(exam_session) AS exam_session
        FROM student_results
        WHERE UPPER(register_number)=UPPER(%s)
        GROUP BY register_number
    """, (target_reg,))
    summary = cur.fetchone()
    if not summary:
        cur.close()
        db.close()
        return "Result not found for this register number", 404

    cur.execute("""
        SELECT
            semester_no, subject_code, subject_name,
            ia_marks, theory_marks, practical_marks,
            result_status, credit, grade
        FROM student_results
        WHERE UPPER(register_number)=UPPER(%s)
        ORDER BY semester_no ASC, subject_code ASC, id ASC
    """, (target_reg,))
    subject_rows = cur.fetchall()
    cur.close()
    db.close()

    return render_template("admin_result_student.html", summary=summary, subject_rows=subject_rows)


@app.route("/admin/results/student/<reg_no>/pdf")
def admin_result_student_pdf(reg_no):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_student_results_table()
    is_management_staff = scope["is_staff"] and "management" in (scope.get("department") or "").strip().lower()

    target_reg = (reg_no or "").strip().upper()
    if not target_reg:
        return redirect(url_for("admin_results"))

    db = get_db()
    cur = db.cursor(dictionary=True)
    if scope["is_staff"] and not is_management_staff:
        cur.execute("""
            SELECT register_number FROM student_results
            WHERE UPPER(register_number)=UPPER(%s) AND branch=%s
            LIMIT 1
        """, (target_reg, scope.get("department")))
        allowed_row = cur.fetchone()
        if not allowed_row:
            cur.close()
            db.close()
            return "Forbidden: You can view only your department results.", 403

    cur.execute("""
        SELECT
            register_number,
            MAX(student_name) AS student_name,
            MAX(branch) AS branch,
            MAX(semester_no) AS semester_no,
            MAX(final_result) AS final_result,
            MAX(percentage) AS percentage,
            MAX(cgpa) AS cgpa,
            MAX(credit_earned_total) AS credit_earned_total,
            MAX(exam_session) AS exam_session
        FROM student_results
        WHERE UPPER(register_number)=UPPER(%s)
        GROUP BY register_number
    """, (target_reg,))
    summary = cur.fetchone()
    if not summary:
        cur.close()
        db.close()
        return "Result not found for this register number", 404

    cur.execute("""
        SELECT
            semester_no, subject_code, subject_name,
            ia_marks, theory_marks, practical_marks,
            result_status, credit, grade
        FROM student_results
        WHERE UPPER(register_number)=UPPER(%s)
        ORDER BY semester_no ASC, subject_code ASC, id ASC
    """, (target_reg,))
    subject_rows = cur.fetchall()
    cur.close()
    db.close()

    pdf = generate_result_student_pdf(summary, subject_rows)
    return send_file(pdf, as_attachment=True)


@app.route("/admin/syllabus", methods=["GET", "POST"])
def admin_syllabus():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_syllabus_documents_table()

    is_staff = scope["is_staff"]
    staff_department = (scope.get("department") or "").strip()
    can_manage = can_upload_syllabus(scope)
    default_series = ["C15", "C20", "C25"]
    if is_staff and not staff_department:
        return "Forbidden: Staff department is not configured.", 403

    if request.method == "POST":
        if not can_manage:
            return "Forbidden: Only Admin and HOD can upload syllabus PDFs.", 403

        branch = request.form.get("branch", "").strip()
        if is_staff:
            branch = staff_department
        if branch not in BRANCH_OPTIONS:
            return "Invalid branch selected", 400

        year_no = parse_int_prefix(request.form.get("year_no", "").strip())
        if year_no is None or year_no < 1 or year_no > 3:
            return "Year must be 1, 2, or 3", 400

        series_choice = normalize_series_text(request.form.get("series", ""))
        custom_series = normalize_series_text(request.form.get("custom_series", ""))
        if series_choice == "ADD":
            series = custom_series
        else:
            series = series_choice
        if not series:
            return "Series is required", 400

        file_obj = request.files.get("syllabus_pdf")
        if not file_obj or file_obj.filename == "":
            return "Please upload a syllabus PDF", 400

        def save_syllabus_pdf(file_obj):
            filename = secure_filename(file_obj.filename)
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            if ext != "pdf":
                return "INVALID_FILE"
            stamp = datetime.now().strftime("%Y%m%d%H%M%S")
            branch_code = branch_code_for_admission(branch)
            unique = f"{branch_code}_{series}_Y{year_no}_{stamp}_{random.randint(1000,9999)}.pdf"
            file_obj.save(os.path.join(SYLLABUS_UPLOAD_FOLDER, unique))
            return unique

        uploaded_pdf = save_syllabus_pdf(file_obj)
        if uploaded_pdf == "INVALID_FILE":
            return "Only PDF files are allowed for syllabus upload", 400

        db = get_db()
        cur = db.cursor(dictionary=True)
        try:
            cur.execute("""
                SELECT id, year1_pdf, year2_pdf, year3_pdf
                FROM syllabus_documents
                WHERE branch=%s AND series=%s AND subject_name=''
                ORDER BY id DESC
                LIMIT 1
            """, (branch, series))
            existing = cur.fetchone() or {}

            year1_pdf = existing.get("year1_pdf")
            year2_pdf = existing.get("year2_pdf")
            year3_pdf = existing.get("year3_pdf")
            if year_no == 1:
                year1_pdf = uploaded_pdf
            elif year_no == 2:
                year2_pdf = uploaded_pdf
            else:
                year3_pdf = uploaded_pdf

            if not year1_pdf and not year2_pdf and not year3_pdf:
                return "Upload at least one PDF (1st year / 2nd year / 3rd year)", 400

            if existing.get("id"):
                cur.execute("""
                    UPDATE syllabus_documents
                    SET year1_pdf=%s,
                        year2_pdf=%s,
                        year3_pdf=%s,
                        semester_no=%s,
                        uploaded_by_staff_id=%s,
                        uploaded_by_role=%s
                    WHERE id=%s
                """, (
                    year1_pdf,
                    year2_pdf,
                    year3_pdf,
                    year_no,
                    session.get("staff_id"),
                    "STAFF" if is_staff else "ADMIN",
                    existing["id"],
                ))
            else:
                cur.execute("""
                    INSERT INTO syllabus_documents (
                        branch, semester_no, series, subject_name,
                        year1_pdf, year2_pdf, year3_pdf,
                        uploaded_by_staff_id, uploaded_by_role
                    )
                    VALUES (%s,%s,%s,'',%s,%s,%s,%s,%s)
                """, (
                    branch,
                    year_no,
                    series,
                    year1_pdf,
                    year2_pdf,
                    year3_pdf,
                    session.get("staff_id"),
                    "STAFF" if is_staff else "ADMIN",
                ))
            db.commit()
        finally:
            cur.close()
            db.close()

        return redirect(url_for(
            "admin_syllabus",
            department=(staff_department if is_staff else branch),
            series=series,
            sort_by=request.form.get("sort_by", "branch_asc"),
            msg="Syllabus saved successfully"
        ))

    selected_department = request.args.get("department", "").strip()
    if is_staff:
        selected_department = staff_department
    elif selected_department and selected_department not in BRANCH_OPTIONS:
        selected_department = ""

    selected_series = normalize_series_text(request.args.get("series", ""))
    year_filter = parse_int_prefix(request.args.get("year_no", "").strip())
    sort_by = (request.args.get("sort_by", "branch_asc") or "branch_asc").strip()
    msg = (request.args.get("msg", "") or "").strip()

    sort_map = {
        "year_asc": "sd.semester_no ASC, sd.branch ASC, sd.series ASC",
        "year_desc": "sd.semester_no DESC, sd.branch ASC, sd.series ASC",
        "branch_asc": "sd.branch ASC, sd.series ASC, sd.semester_no ASC",
        "branch_desc": "sd.branch DESC, sd.series ASC, sd.semester_no ASC",
    }
    order_clause = sort_map.get(sort_by, sort_map["branch_asc"])

    db = get_db()
    cur = db.cursor(dictionary=True)
    query = """
        SELECT
            sd.id,
            sd.semester_no,
            sd.branch,
            sd.series,
            sd.year1_pdf,
            sd.year2_pdf,
            sd.year3_pdf
        FROM syllabus_documents sd
        WHERE 1=1
    """
    params = []

    if selected_department:
        query += " AND sd.branch=%s"
        params.append(selected_department)
    if selected_series:
        query += " AND sd.series=%s"
        params.append(selected_series)
    if year_filter and 1 <= year_filter <= 3:
        if year_filter == 1:
            query += " AND COALESCE(sd.year1_pdf,'')<>''"
        elif year_filter == 2:
            query += " AND COALESCE(sd.year2_pdf,'')<>''"
        else:
            query += " AND COALESCE(sd.year3_pdf,'')<>''"

    query += f" ORDER BY {order_clause}"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    series_query = "SELECT DISTINCT series FROM syllabus_documents WHERE 1=1"
    series_params = []
    if is_staff:
        series_query += " AND branch=%s"
        series_params.append(staff_department)
    series_query += " ORDER BY series ASC"
    cur.execute(series_query, tuple(series_params))
    db_series = [str(r.get("series") or "").strip().upper() for r in cur.fetchall() if r.get("series")]

    cur.close()
    db.close()

    series_options = list(default_series)
    for series_name in db_series:
        if series_name and series_name not in series_options:
            series_options.append(series_name)

    return render_template(
        "admin_syllabus.html",
        syllabus_rows=rows,
        branches=BRANCH_OPTIONS,
        is_staff=is_staff,
        staff_department=staff_department,
        can_manage_syllabus=can_manage,
        series_options=series_options,
        filters={
            "department": selected_department,
            "series": selected_series,
            "year_no": str(year_filter) if year_filter else "",
            "sort_by": sort_by,
        },
        msg=msg
    )


@app.route("/admin/student-details")
def admin_student_details():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    from datetime import date

    q = request.args.get("q", "").strip()
    status = request.args.get("status", "ACTIVE").strip()
    branch = request.args.get("branch", "").strip()

    ensure_students_college_reg_no_column()
    ensure_student_personal_extra_columns()
    students, branches, branch = fetch_admin_student_rows(scope, q=q, status=status, branch=branch)

    current_year = date.today().year
    academic_year = f"{current_year}-{str((current_year + 1) % 100).zfill(2)}"

    return render_template(
        "admin_student_details.html",
        students=students,
        branches=branches,
        academic_year=academic_year,
        is_staff=scope["is_staff"],
        staff_department=scope["department"],
        filters={
            "q": q,
            "status": status,
            "branch": branch
        }
    )


def fetch_admin_student_rows(scope, q="", status="ACTIVE", branch=""):
    if scope.get("is_staff"):
        branch = scope.get("department") or ""

    db = get_db()
    cur = db.cursor(dictionary=True)
    query = """
        SELECT
            s.admission_id,
            s.student_name,
            s.branch,
            s.mobile,
            s.status,
            COALESCE(s.college_reg_no, '') AS college_reg_no,
            COALESCE(spd.dob, '-') AS dob,
            COALESCE(spd.gender, '-') AS gender,
            COALESCE(NULLIF(spd.student_email, ''), '-') AS student_email,
            COALESCE(NULLIF(spd.ssp_id, ''), '-') AS ssp_id,
            COALESCE(NULLIF(spd.apaar_id, ''), '-') AS apaar_id,
            COALESCE(spd.caste_category, '-') AS caste_category,
            COALESCE(spd.alloted_category, '-') AS alloted_category,
            COALESCE(sd.student_photo, spd.photo_file, '') AS photo_file
        FROM students s
        LEFT JOIN student_personal_details spd
            ON spd.admission_id = s.admission_id
        LEFT JOIN student_documents sd
            ON sd.admission_id = s.admission_id
        WHERE 1=1
    """
    params = []

    if q:
        query += """
            AND (
                s.student_name LIKE %s
                OR s.admission_id LIKE %s
                OR s.college_reg_no LIKE %s
                OR spd.register_number LIKE %s
                OR spd.student_email LIKE %s
                OR spd.ssp_id LIKE %s
                OR spd.apaar_id LIKE %s
            )
        """
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])

    if status:
        query += " AND s.status=%s"
        params.append(status)

    if branch:
        query += " AND s.branch=%s"
        params.append(branch)

    query += " ORDER BY s.student_name ASC"
    cur.execute(query, tuple(params))
    students = cur.fetchall()

    cur.execute("SELECT DISTINCT branch FROM students ORDER BY branch ASC")
    branch_rows = cur.fetchall()
    branches = [row["branch"] for row in branch_rows if row.get("branch")]

    cur.close()
    db.close()
    return students, branches, branch


@app.route("/admin/student-details/pdf")
def admin_student_details_pdf():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    q = request.args.get("q", "").strip()
    status = request.args.get("status", "ACTIVE").strip()
    branch = request.args.get("branch", "").strip()
    students, _, branch = fetch_admin_student_rows(scope, q=q, status=status, branch=branch)

    filter_parts = []
    if q:
        filter_parts.append(f"Search: {q}")
    if status:
        filter_parts.append(f"Status: {status}")
    if branch:
        filter_parts.append(f"Branch: {branch}")
    filter_text = " | ".join(filter_parts)

    pdf_path = generate_students_list_pdf(
        students,
        title_text="Student Records - Consolidated",
        filter_text=filter_text
    )
    return send_file(pdf_path, as_attachment=True)


@app.route("/admin/attendance", methods=["GET", "POST"])
def admin_attendance():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_student_attendance_table()
    ensure_students_college_reg_no_column()
    ensure_subject_master_table()
    ensure_subject_series_table()

    today = date.today()
    today_value = today.isoformat()
    msg = request.args.get("msg", "").strip()

    if request.method == "POST":
        if not validate_csrf():
            return redirect(url_for("admin_attendance", msg="Session expired. Refresh and try again."))

        selected_date_obj = parse_iso_date_safe(request.form.get("attendance_date", ""), today)
        selected_date = selected_date_obj.isoformat()
        month_value, _ = parse_month_safe(request.form.get("month", ""), selected_date[:7])
        subject = normalize_subject_name(request.form.get("subject", ""))
        period_no = parse_int_prefix(request.form.get("period_no", "1")) or 1
        branch = (request.form.get("branch", "") or "").strip()
        semester = (request.form.get("semester", "") or "").strip()
        series = (request.form.get("series", "") or "").strip().upper()
        q = (request.form.get("q", "") or "").strip()
        bulk_status = (request.form.get("bulk_status", "") or "").strip().upper()
        if scope["is_staff"]:
            branch = scope["department"]

        if period_no < 1 or period_no > 12:
            period_no = 1
        if not subject:
            return redirect(url_for(
                "admin_attendance",
                attendance_date=selected_date,
                month=month_value,
                period_no=period_no,
                branch=branch,
                semester=semester,
                series=series,
                q=q,
                msg="Subject is required."
            ))

        marked_by = (session.get("admin") or "admin").strip()
        marked_by_name = marked_by
        if scope["is_staff"] and session.get("staff_id"):
            db_staff = get_db()
            cur_staff = db_staff.cursor(dictionary=True)
            cur_staff.execute("SELECT employee_name FROM staff_accounts WHERE id=%s", (session.get("staff_id"),))
            staff_row = cur_staff.fetchone()
            cur_staff.close()
            db_staff.close()
            if staff_row and staff_row.get("employee_name"):
                marked_by_name = staff_row.get("employee_name").strip()

        student_ids = [s.strip() for s in request.form.getlist("student_ids") if (s or "").strip()]
        allowed_status = {"PRESENT", "ABSENT", "LATE", "LEAVE"}
        if bulk_status not in allowed_status:
            bulk_status = ""
        saved_count = 0

        student_meta = {}
        if student_ids:
            db_meta = get_db()
            cur_meta = db_meta.cursor(dictionary=True)
            placeholders = ",".join(["%s"] * len(student_ids))
            cur_meta.execute(f"""
                SELECT admission_id, branch, year_sem, admission_year
                FROM students
                WHERE status='ACTIVE' AND admission_id IN ({placeholders})
            """, tuple(student_ids))
            for row in cur_meta.fetchall():
                if scope["is_staff"] and not same_department(row.get("branch"), scope.get("department")):
                    continue
                row_sem = parse_int_prefix(row.get("year_sem"))
                if row_sem is None:
                    row_sem = infer_current_sem(row.get("admission_year"), row.get("year_sem"))
                student_meta[row["admission_id"]] = {
                    "branch": row.get("branch") or "",
                    "semester_no": row_sem
                }
            cur_meta.close()
            db_meta.close()

        db = get_db()
        cur = db.cursor()
        save_error = ""
        try:
            for admission_id in student_ids:
                sid = (admission_id or "").strip()
                if not sid or sid not in student_meta:
                    continue
                status = bulk_status or (request.form.get(f"quick_status_{sid}", "") or "").strip().upper()
                if status not in allowed_status:
                    continue
                remarks = (request.form.get(f"quick_remarks_{sid}", "") or "").strip()
                row_branch = student_meta[sid]["branch"]
                row_semester = student_meta[sid]["semester_no"]
                if not row_branch:
                    continue
                cur.execute("""
                    INSERT INTO student_daily_attendance (
                        admission_id, branch, semester_no, attendance_date,
                        subject_name, period_no, status, remarks, marked_by, marked_by_name
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                        branch=VALUES(branch),
                        semester_no=VALUES(semester_no),
                        status=VALUES(status),
                        remarks=VALUES(remarks),
                        marked_by=VALUES(marked_by),
                        marked_by_name=VALUES(marked_by_name),
                        updated_at=CURRENT_TIMESTAMP
                """, (
                    sid, row_branch, row_semester, selected_date, subject,
                    period_no, status, remarks, marked_by, marked_by_name
                ))
                saved_count += 1
            db.commit()
        except Exception as exc:
            db.rollback()
            save_error = f"Attendance save failed: {exc}"
        finally:
            cur.close()
            db.close()

        if save_error:
            save_msg = save_error
        elif saved_count == 0:
            save_msg = "No rows saved. Check subject/date and choose at least one status."
        else:
            save_msg = f"Attendance saved for {saved_count} student(s)."
        return redirect(url_for(
            "admin_attendance",
            attendance_date=selected_date,
            month=month_value,
            subject=subject,
            period_no=period_no,
            branch=branch,
            semester=semester,
            series=series,
            q=q,
            msg=save_msg
        ))

    selected_date_obj = parse_iso_date_safe(request.args.get("attendance_date", today_value).strip(), today)
    subject = normalize_subject_name(request.args.get("subject", ""))
    period_no = parse_int_prefix(request.args.get("period_no", "1")) or 1
    if period_no < 1 or period_no > 12:
        period_no = 1
    branch = request.args.get("branch", "").strip()
    if scope["is_staff"]:
        branch = scope["department"]
    semester = request.args.get("semester", "").strip()
    series = (request.args.get("series", "") or "").strip().upper()
    q = request.args.get("q", "").strip()
    month_value = request.args.get("month", "").strip() or selected_date_obj.strftime("%Y-%m")

    register = fetch_attendance_register(
        scope=scope,
        selected_date_obj=selected_date_obj,
        month_value=month_value,
        subject=subject,
        period_no=period_no,
        branch=branch,
        semester=semester,
        q=q
    )

    return render_template(
        "admin_attendance.html",
        students=register["students"],
        day_numbers=register["day_numbers"],
        day_wise_stats=register["day_wise_stats"],
        selected_day=register["selected_day"],
        month_label=register["month_label"],
        summary=register["summary"],
        branches=register["branches"],
        is_staff=scope["is_staff"],
        staff_department=scope["department"],
        series_options=fetch_subject_series_options(),
        msg=msg,
        filters={
            "attendance_date": register["selected_date"],
            "month": register["month"],
            "subject": subject,
            "period_no": period_no,
            "branch": branch,
            "semester": semester,
            "series": series,
            "q": q
        }
    )


@app.route("/admin/attendance/export")
def admin_attendance_export():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    selected_date_obj = parse_iso_date_safe(request.args.get("attendance_date", ""), date.today())
    month_value = request.args.get("month", "").strip() or selected_date_obj.strftime("%Y-%m")
    subject = normalize_subject_name(request.args.get("subject", ""))
    period_no = parse_int_prefix(request.args.get("period_no", "1")) or 1
    branch = request.args.get("branch", "").strip()
    if scope["is_staff"]:
        branch = scope["department"]
    semester = request.args.get("semester", "").strip()
    q = request.args.get("q", "").strip()
    register = fetch_attendance_register(scope, selected_date_obj, month_value, subject, period_no, branch, semester, q)

    header = ["Admission ID", "Reg No", "Student Name", "Branch", "Semester"] + [f"D{d}" for d in register["day_numbers"]] + [
        "Total Marked Days", "Present Days", "Absent Days", "Leave Days", "Attendance %",
    ]
    rows = []
    for row in register["students"]:
        day_values = []
        for day in register["day_numbers"]:
            st = row["daily_status"].get(day)
            if st in ("PRESENT", "LATE"):
                day_values.append("P")
            elif st == "ABSENT":
                day_values.append("A")
            elif st == "LEAVE":
                day_values.append("L")
            else:
                day_values.append("")
        rows.append([
            row["admission_id"], row.get("reg_no", ""), row.get("student_name", ""),
            row.get("branch", ""), row.get("semester_display", ""), *day_values,
            row.get("total_marked_days", 0), row.get("present_days", 0),
            row.get("absent_days", 0), row.get("leave_days", 0), row.get("attendance_pct", 0),
        ])
    subject_token = (subject or "all-subjects").replace(" ", "_")
    return _send_attendance_export(f"attendance_{register['month']}_p{period_no}_{subject_token}", header, rows)


@app.route("/admin/attendance/daywise/export")
def admin_attendance_daywise_export():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    selected_date_obj = parse_iso_date_safe(request.args.get("attendance_date", ""), date.today())
    month_value = request.args.get("month", "").strip() or selected_date_obj.strftime("%Y-%m")
    subject = normalize_subject_name(request.args.get("subject", ""))
    period_no = parse_int_prefix(request.args.get("period_no", "1")) or 1
    branch = request.args.get("branch", "").strip()
    if scope["is_staff"]:
        branch = scope["department"]
    semester = request.args.get("semester", "").strip()
    q = request.args.get("q", "").strip()
    register = fetch_attendance_register(scope, selected_date_obj, month_value, subject, period_no, branch, semester, q)

    header = ["Day", "Present", "Absent", "Leave", "Unmarked", "Marked", "Present %"]
    rows = [[r["day"], r["present"], r["absent"], r["leave"], r["unmarked"], r["marked"], r["present_pct"]] for r in register["day_wise_stats"]]
    subject_token = (subject or "all-subjects").replace(" ", "_")
    return _send_attendance_export(f"attendance_daywise_{register['month']}_p{period_no}_{subject_token}", header, rows)


@app.route("/admin/attendance/blank-sheet/export")
def admin_attendance_blank_sheet_export():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    selected_date_obj = parse_iso_date_safe(request.args.get("attendance_date", ""), date.today())
    month_value = request.args.get("month", "").strip() or selected_date_obj.strftime("%Y-%m")
    subject = normalize_subject_name(request.args.get("subject", ""))
    period_no = parse_int_prefix(request.args.get("period_no", "1")) or 1
    branch = request.args.get("branch", "").strip()
    if scope["is_staff"]:
        branch = scope["department"]
    semester = request.args.get("semester", "").strip()
    q = request.args.get("q", "").strip()
    register = fetch_attendance_register(scope, selected_date_obj, month_value, subject, period_no, branch, semester, q)

    header = ["Admission ID", "Reg No", "Student Name", "Branch", "Semester"] + [f"D{d}" for d in register["day_numbers"]]
    rows = []
    for row in register["students"]:
        rows.append([
            row.get("admission_id", ""), row.get("reg_no", ""), row.get("student_name", ""),
            row.get("branch", ""), row.get("semester_display", ""), *([""] * len(register["day_numbers"]))
        ])
    subject_token = (subject or "all-subjects").replace(" ", "_")
    return _send_attendance_export(f"attendance_blank_sheet_{register['month']}_p{period_no}_{subject_token}", header, rows)


@app.route("/admin/student-details/edit/<admission_id>", methods=["GET", "POST"])
def edit_student_details(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"] and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can edit only your department students.", 403
    can_set_alloted = can_set_allotted_category(scope)
    can_edit_identity_ids = can_edit_student_identity_ids(scope)

    ensure_students_college_reg_no_column()
    ensure_students_admission_year_column()
    ensure_students_year_sem_column()
    ensure_student_personal_extra_columns()

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT
            s.admission_id,
            s.student_name,
            s.branch,
            s.mobile,
            s.admission_year,
            s.year_sem,
            COALESCE(s.college_reg_no, '') AS college_reg_no,
            COALESCE(spd.gender, '') AS gender,
            COALESCE(spd.caste_category, '') AS caste_category,
            COALESCE(spd.alloted_category, '') AS alloted_category,
            COALESCE(spd.register_number, '') AS register_number,
            COALESCE(spd.student_email, '') AS student_email,
            COALESCE(spd.ssp_id, '') AS ssp_id,
            COALESCE(spd.apaar_id, '') AS apaar_id,
            ed.total_max_marks,
            ed.total_marks_obtained,
            ed.science_max_marks,
            ed.science_marks_obtained,
            ed.maths_max_marks,
            ed.maths_marks_obtained
        FROM students s
        LEFT JOIN student_personal_details spd
            ON spd.admission_id = s.admission_id
        LEFT JOIN education_details ed
            ON ed.admission_id = s.admission_id
        WHERE s.admission_id=%s
        LIMIT 1
    """, (admission_id,))
    student = cur.fetchone()

    if not student:
        cur.close()
        db.close()
        return "Student not found", 404

    if request.method == "POST":
        student_name = request.form.get("student_name", "").strip()
        branch = request.form.get("branch", "").strip()
        mobile = request.form.get("mobile", "").strip()
        college_reg_no = request.form.get("college_reg_no", "").strip()
        gender = request.form.get("gender", "").strip()
        caste_category = request.form.get("caste_category", "").strip()
        alloted_category = request.form.get("alloted_category", "").strip()
        allowed_alloted_categories = {"1G", "SCG", "STG", "2AG", "2BG", "3AG", "3BG", "GM", "PENDING", ""}
        student_email = request.form.get("student_email", "").strip().lower()
        ssp_id = request.form.get("ssp_id", "").strip().upper()
        apaar_id = request.form.get("apaar_id", "").strip().upper()
        if not can_set_alloted:
            current_value = (student.get("alloted_category") or "").strip()
            if alloted_category != current_value:
                cur.close()
                db.close()
                return "Forbidden: Only Admin or Management can update Allotted Category.", 403
            alloted_category = current_value
        elif alloted_category not in allowed_alloted_categories:
            cur.close()
            db.close()
            return "Invalid allotted category", 400
        if scope.get("is_staff") and not can_edit_identity_ids:
            current_email = (student.get("student_email") or "").strip().lower()
            current_ssp = (student.get("ssp_id") or "").strip().upper()
            current_apaar = (student.get("apaar_id") or "").strip().upper()
            if student_email != current_email or ssp_id != current_ssp or apaar_id != current_apaar:
                cur.close()
                db.close()
                return "Forbidden: Only Admin/HOD can update Email/SSP/APAAR.", 403
            student_email = current_email
            ssp_id = current_ssp
            apaar_id = current_apaar
        register_number = request.form.get("register_number", "").strip()
        total_max_marks_raw = request.form.get("total_max_marks", "").strip()
        total_marks_obtained_raw = request.form.get("total_marks_obtained", "").strip()
        science_max_marks_raw = request.form.get("science_max_marks", "").strip()
        science_marks_obtained_raw = request.form.get("science_marks_obtained", "").strip()
        maths_max_marks_raw = request.form.get("maths_max_marks", "").strip()
        maths_marks_obtained_raw = request.form.get("maths_marks_obtained", "").strip()
        manual_sem_raw = request.form.get("year_sem", "").strip()
        manual_sem = None
        if manual_sem_raw:
            manual_sem = parse_int_prefix(manual_sem_raw)
            if manual_sem is None or manual_sem < 1 or manual_sem > 6:
                cur.close()
                db.close()
                return "Manual semester must be between 1 and 6", 400

        if not student_name or not branch or not mobile:
            cur.close()
            db.close()
            return "Student name, branch, and mobile are required"

        def to_int_or_none(val):
            text = str(val or "").strip()
            if text == "":
                return None
            try:
                return int(float(text))
            except ValueError:
                return None

        total_max_marks = to_int_or_none(total_max_marks_raw)
        total_marks_obtained = to_int_or_none(total_marks_obtained_raw)
        science_max_marks = to_int_or_none(science_max_marks_raw)
        science_marks_obtained = to_int_or_none(science_marks_obtained_raw)
        maths_max_marks = to_int_or_none(maths_max_marks_raw)
        maths_marks_obtained = to_int_or_none(maths_marks_obtained_raw)

        update_cur = db.cursor()
        update_cur.execute("""
            UPDATE students
            SET student_name=%s, branch=%s, mobile=%s, college_reg_no=%s, year_sem=%s
            WHERE admission_id=%s
        """, (student_name, branch, mobile, college_reg_no, manual_sem, admission_id))

        update_cur.execute("""
            SELECT id FROM student_personal_details WHERE admission_id=%s
        """, (admission_id,))
        personal_exists = update_cur.fetchone()

        if personal_exists:
            update_cur.execute("""
                UPDATE student_personal_details
                SET gender=%s,
                    student_email=%s,
                    ssp_id=%s,
                    apaar_id=%s,
                    caste_category=%s,
                    alloted_category=%s,
                    register_number=%s
                WHERE admission_id=%s
            """, (gender, student_email or None, ssp_id or None, apaar_id or None, caste_category, alloted_category, register_number, admission_id))
        else:
            update_cur.execute("""
                INSERT INTO student_personal_details
                (admission_id, gender, student_email, ssp_id, apaar_id, caste_category, alloted_category, register_number)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (admission_id, gender, student_email or None, ssp_id or None, apaar_id or None, caste_category, alloted_category, register_number))

        update_cur.execute(
            "SELECT admission_id FROM education_details WHERE admission_id=%s LIMIT 1",
            (admission_id,)
        )
        education_exists = update_cur.fetchone()
        if education_exists:
            update_cur.execute("""
                UPDATE education_details
                SET total_max_marks=%s,
                    total_marks_obtained=%s,
                    science_max_marks=%s,
                    science_marks_obtained=%s,
                    maths_max_marks=%s,
                    maths_marks_obtained=%s
                WHERE admission_id=%s
            """, (
                total_max_marks,
                total_marks_obtained,
                science_max_marks,
                science_marks_obtained,
                maths_max_marks,
                maths_marks_obtained,
                admission_id,
            ))
        else:
            update_cur.execute("""
                INSERT INTO education_details (
                    admission_id,
                    total_max_marks,
                    total_marks_obtained,
                    science_max_marks,
                    science_marks_obtained,
                    maths_max_marks,
                    maths_marks_obtained
                ) VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (
                admission_id,
                total_max_marks if total_max_marks is not None else 0,
                total_marks_obtained if total_marks_obtained is not None else 0,
                science_max_marks if science_max_marks is not None else 0,
                science_marks_obtained if science_marks_obtained is not None else 0,
                maths_max_marks if maths_max_marks is not None else 0,
                maths_marks_obtained if maths_marks_obtained is not None else 0,
            ))

        db.commit()
        update_cur.close()
        cur.close()
        db.close()
        return redirect("/admin/student-details")

    cur.close()
    db.close()
    return render_template(
        "admin_edit_student.html",
        student=student,
        can_set_alloted_category=can_set_alloted,
        can_edit_identity_ids=can_edit_identity_ids
    )


@app.route("/admin/employees")
def admin_employees():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_employees_table()

    employee_type = request.args.get("employee_type", "").strip()
    department = request.args.get("department", "").strip()
    if scope["is_staff"]:
        department = scope["department"]
    hod_only = request.args.get("hod", "").strip().lower()
    sort_by = request.args.get("sort_by", "employee_name").strip()
    name_query = request.args.get("name", "").strip()

    sort_columns = {
        "employee_name": "employee_name",
        "department": "department",
        "designation": "designation"
    }
    sort_column = sort_columns.get(sort_by, "employee_name")

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT id, employee_name, department, designation, mobile_no, employee_type
        FROM employee_details
        WHERE 1=1
    """
    params = []

    if employee_type:
        query += " AND employee_type = %s"
        params.append(employee_type)

    if department:
        query += " AND department = %s"
        params.append(department)

    if name_query:
        query += " AND LOWER(employee_name) LIKE LOWER(%s)"
        params.append(f"%{name_query}%")

    if hod_only == "yes":
        query += " AND (UPPER(designation) LIKE %s OR UPPER(designation) = %s)"
        params.append("%HOD%")
        params.append("HEAD OF DEPARTMENT")

    query += f" ORDER BY {sort_column} ASC"
    cur.execute(query, tuple(params))
    employees = cur.fetchall()

    cur.execute("SELECT DISTINCT department FROM employee_details ORDER BY department ASC")
    department_rows = cur.fetchall()
    departments = [row["department"] for row in department_rows if row.get("department")]

    cur.execute("SELECT DISTINCT employee_name FROM employee_details ORDER BY employee_name ASC")
    name_rows = cur.fetchall()
    employee_names = [row["employee_name"] for row in name_rows if row.get("employee_name")]

    cur.close()
    db.close()

    return render_template(
        "admin_employees.html",
        employees=employees,
        departments=departments,
        employee_names=employee_names,
        is_staff=scope["is_staff"],
        staff_department=scope["department"],
        filters={
            "employee_type": employee_type,
            "department": department,
            "hod": hod_only,
            "sort_by": sort_by,
            "name": name_query
        }
    )


@app.route("/admin/employees/add", methods=["GET", "POST"])
def add_employee():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_employees_table()

    if request.method == "POST":
        employee_name = request.form.get("employee_name", "").strip()
        department = request.form.get("department", "").strip()
        if scope["is_staff"]:
            department = scope["department"]
        designation = request.form.get("designation", "").strip()
        mobile_no = request.form.get("mobile_no", "").strip()

        if not employee_name or not department or not designation or not mobile_no:
            return "All fields are required"

        if not mobile_no.isdigit() or len(mobile_no) != 10:
            return "PLEASE ENTER 10 DIGIT MOBILE NUMBER"

        designation_type_map = {
            "HEAD OF DEPARTMENT": "TEACHING",
            "TEACHING STAFF": "TEACHING",
            "NON TEACHING STAFF": "NON-TEACHING",
            "HELPER": "HELPER"
        }
        employee_type = designation_type_map.get(designation.upper(), "TEACHING")

        db = get_db()
        cur = db.cursor()
        cur.execute("""
            INSERT INTO employee_details
            (employee_name, department, designation, mobile_no, employee_type)
            VALUES (%s, %s, %s, %s, %s)
        """, (employee_name, department, designation, mobile_no, employee_type))
        db.commit()
        cur.close()
        db.close()

        return redirect("/admin/employees")

    return render_template(
        "add_employee.html",
        edit_mode=False,
        employee={},
        is_staff=scope["is_staff"],
        staff_department=scope["department"]
    )


@app.route("/admin/employees/edit/<int:employee_id>", methods=["GET", "POST"])
def edit_employee(employee_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_employees_table()

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT id, employee_name, department, designation, mobile_no
        FROM employee_details
        WHERE id=%s
    """, (employee_id,))
    employee = cur.fetchone()
    if scope["is_staff"] and employee and (employee.get("department") or "").strip().lower() != (scope["department"] or "").strip().lower():
        cur.close()
        db.close()
        return "Forbidden: You can edit only your department.", 403

    if not employee:
        cur.close()
        db.close()
        return "Employee not found", 404

    if request.method == "POST":
        employee_name = request.form.get("employee_name", "").strip()
        department = request.form.get("department", "").strip()
        if scope["is_staff"]:
            department = scope["department"]
        designation = request.form.get("designation", "").strip()
        mobile_no = request.form.get("mobile_no", "").strip()

        if not employee_name or not department or not designation or not mobile_no:
            cur.close()
            db.close()
            return "All fields are required"

        if not mobile_no.isdigit() or len(mobile_no) != 10:
            cur.close()
            db.close()
            return "PLEASE ENTER 10 DIGIT MOBILE NUMBER"

        designation_type_map = {
            "HEAD OF DEPARTMENT": "TEACHING",
            "TEACHING STAFF": "TEACHING",
            "NON TEACHING STAFF": "NON-TEACHING",
            "HELPER": "HELPER"
        }
        employee_type = designation_type_map.get(designation.upper(), "TEACHING")

        update_cur = db.cursor()
        update_cur.execute("""
            UPDATE employee_details
            SET employee_name=%s,
                department=%s,
                designation=%s,
                mobile_no=%s,
                employee_type=%s
            WHERE id=%s
        """, (employee_name, department, designation, mobile_no, employee_type, employee_id))
        db.commit()
        update_cur.close()
        cur.close()
        db.close()
        return redirect("/admin/employees")

    cur.close()
    db.close()
    return render_template(
        "add_employee.html",
        edit_mode=True,
        employee=employee,
        is_staff=scope["is_staff"],
        staff_department=scope["department"]
    )


# =========================
# ONLINE ADMISSION (PUBLIC)
# =========================

import os
import random
import hashlib
from werkzeug.utils import secure_filename
from flask import request, render_template
from db import get_db
UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


@app.route("/admission", methods=["GET", "POST"])
def admission():
    if request.method == "POST":
        admission_year_text = request.form.get("admission_year", "").strip() or current_academic_year()
        if not re.match(r"^\d{4}-\d{2}$", admission_year_text):
            return "Admission year must be in YYYY-YY format (example: 2026-27)", 400
        admission_id = generate_admission_id(
            request.form.get("branch", ""),
            admission_year_text
        )
        admission_year = int(admission_year_text[:4])

        # PASSWORD (student creates)
        password_hash = generate_password_hash(request.form["password"])


        # =========================
        # FILE UPLOADS
        # =========================
        photo = request.files["photo"]
        marks_card = request.files["marks_card"]
        caste_cert = request.files["caste_certificate"]
        income_cert = request.files["income_certificate"]

        photo_name = admission_id + "_photo_" + secure_filename(photo.filename)
        marks_name = admission_id + "_marks_" + secure_filename(marks_card.filename)
        caste_name = admission_id + "_caste_" + secure_filename(caste_cert.filename)
        income_name = admission_id + "_income_" + secure_filename(income_cert.filename)

        photo.save(os.path.join(UPLOAD_FOLDER, photo_name))
        marks_card.save(os.path.join(UPLOAD_FOLDER, marks_name))
        caste_cert.save(os.path.join(UPLOAD_FOLDER, caste_name))
        income_cert.save(os.path.join(UPLOAD_FOLDER, income_name))

        # =========================
        # DATABASE INSERTS
        # =========================
        db = get_db()
        cur = db.cursor()
        ensure_students_admission_year_column()

        # 1️⃣ STUDENTS TABLE (LOGIN)
        admission_id = insert_student_login_with_retry(
            cur,
            admission_id=admission_id,
            student_name=request.form["student_name"],
            branch=request.form["branch"],
            admission_year=admission_year,
            mobile=request.form["student_mobile"],
            password_hash=password_hash,
            status="INACTIVE",
            admission_year_text=admission_year_text,
        )

        # 2️⃣ STUDENT PERSONAL DETAILS
        cur.execute("""
    INSERT INTO student_personal_details
    (admission_id, student_mobile, student_email,
     indian_nationality, religion, disability,
     aadhaar_number,
     caste_rd_number, caste_certificate_file,
     income_rd_number, income_certificate_file,
     annual_income,
     mother_name, mother_mobile,
     father_name, father_mobile,
     residential_address, permanent_address)
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
""", (
    admission_id,
    request.form["student_mobile"].strip().upper(),
    request.form["student_email"].strip(),   # ← comma added here
    request.form["indian_nationality"].strip().upper(),
    request.form["religion"].strip().upper(),
    request.form["disability"].strip().upper(),
    request.form["aadhaar_number"].strip().upper(),
    request.form["caste_rd_number"].strip().upper(),
    caste_name,
    request.form["income_rd_number"].strip().upper(),
    income_name,
    request.form["annual_income"],
    request.form["mother_name"].strip().upper(),
    request.form["mother_mobile"],
    request.form["father_name"].strip().upper(),
    request.form["father_mobile"],
    request.form["residential_address"],
    request.form["permanent_address"]
))

        # 3️⃣ EDUCATION DETAILS (for admin/student view pages)
        def _to_int_or_none(v):
            text = str(v or "").strip()
            if not text:
                return None
            try:
                return int(float(text))
            except ValueError:
                return None

        total_max_marks = _to_int_or_none(request.form.get("total_marks") or request.form.get("total_max_marks"))
        total_marks_obtained = _to_int_or_none(request.form.get("marks_obtained") or request.form.get("total_marks_obtained"))
        science_marks_obtained = _to_int_or_none(request.form.get("science_marks") or request.form.get("science_marks_obtained"))
        maths_marks_obtained = _to_int_or_none(request.form.get("maths_marks") or request.form.get("maths_marks_obtained"))

        cur.execute("""
            INSERT INTO education_details
            (admission_id, qualifying_exam, register_number, year_of_passing,
             total_max_marks, total_marks_obtained,
             science_max_marks, science_marks_obtained,
             maths_max_marks, maths_marks_obtained)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            admission_id,
            request.form.get("qualifying_exam"),
            request.form.get("register_number"),
            request.form.get("year_of_passing"),
            total_max_marks,
            total_marks_obtained,
            None,
            science_marks_obtained,
            None,
            maths_marks_obtained
        ))

        db.commit()

        return render_template(
            "admission_success.html",
            admission_id=admission_id
        )

    return render_template("admission_form.html")

@app.route("/admission/step-1", methods=["GET", "POST"])
def admission_step1():
    if request.method == "POST":
        admission_year = request.form.get("admission_year", "").strip() or current_academic_year()
        if not re.match(r"^\d{4}-\d{2}$", admission_year):
            return "Admission year must be in YYYY-YY format (example: 2026-27)", 400

        session["admission"] = {
            "admission_id": None,
            "admission_year": admission_year,

            # Student Personal
            "student_name": request.form["student_name"],
            "student_mobile": request.form["student_mobile"],
            "student_email": request.form["student_email"],
            "dob": request.form["dob"],
            "gender": request.form["gender"],
            "indian_nationality": request.form["indian_nationality"],
            "religion": request.form.get("religion"),
            "caste_category": request.form["caste_category"],
            # Student cannot set allotted category in online admission.
            "alloted_category": "PENDING",

            # Academic
            # Academic (Dynamic)
"qualifying_exam": request.form["qualifying_exam"],
"year_of_passing": request.form["year_of_passing"],
"register_number": request.form["register_number"],

# SSLC / PUC Marks
"maths_marks": request.form.get("maths_marks"),
"science_marks": request.form.get("science_marks"),
"total_marks": request.form.get("total_marks"),
"marks_obtained": request.form.get("marks_obtained"),
"percentage": request.form.get("percentage"),




            # Admission
            "admission_quota": request.form["admission_quota"],
            "branch": request.form["branch"],
            "password": request.form["password"]
        }

        session.modified = True
        return redirect("/admission/step-2")

    return render_template("admission_step1.html", admission_year=current_academic_year())






@app.route("/admission/step-2", methods=["GET", "POST"])
def admission_step2():

    # get admission data created in step-1
    admission = session.get("admission")
    if not admission:
        return redirect("/admission/step-1")

    if request.method == "POST":
        admission.update({
            "father_name": request.form.get("father_name").upper(),
            "father_mobile": request.form.get("father_mobile"),
            "mother_name": request.form.get("mother_name").upper(),
            "mother_mobile": request.form.get("mother_mobile"),
            "residential_address": request.form.get("residential_address").upper(),
            "permanent_address": request.form.get("permanent_address").upper(),
        })

        # VERY IMPORTANT
        session["admission"] = admission
        session.modified = True

        return redirect("/admission/step-3")

    return render_template("admission_step2.html")



from datetime import date
import hashlib
import os
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = "static/uploads"
ALLOWED_EXTENSIONS = {"pdf", "jpg", "jpeg", "png"}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


from datetime import date
import hashlib
import os
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = "static/uploads"
ALLOWED_EXTENSIONS = {"pdf", "jpg", "jpeg", "png"}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


from datetime import date
from werkzeug.utils import secure_filename

from datetime import date
import hashlib

import re
from datetime import date
import hashlib

@app.route("/admission/step-3", methods=["GET", "POST"])
def admission_step3():

    admission = session.get("admission")
    if not admission:
        return redirect("/admission/step-1")

    # =========================
    # POST → VALIDATION & SUBMIT
    # =========================
    if request.method == "POST":
        if not admission.get("admission_id"):
            admission["admission_id"] = generate_admission_id(
                admission.get("branch", ""),
                admission.get("admission_year") or current_academic_year()
            )
            session["admission"] = admission
            session.modified = True

        # ===== READ STEP-3 FORM DATA =====
        aadhaar_number = request.form.get("aadhaar_number", "").strip().upper()
        caste_rd_number = request.form.get("caste_rd_number", "").strip().upper()
        income_rd_number = request.form.get("income_rd_number", "").strip().upper()

        # ===== AADHAAR VALIDATION =====
        if not aadhaar_number.isdigit() or len(aadhaar_number) != 12:
            return "❌ Aadhaar number must be exactly 12 digits"

        # ===== CASTE RD VALIDATION =====
        if not re.match(r"^[A-Za-z0-9]{6,20}$", caste_rd_number):
            return "❌ Invalid Caste Certificate RD Number"

        # ===== INCOME RD VALIDATION =====
        if not re.match(r"^[A-Za-z0-9]{6,20}$", income_rd_number):
            return "❌ Invalid Income Certificate RD Number"

        # ===== AGE VALIDATION =====
        dob = date.fromisoformat(admission["dob"])
        age = (date.today() - dob).days // 365
        if age < 14:
            return "❌ Student must be at least 14 years old"

        # ===== YEAR VALIDATION =====
        year = int(admission["year_of_passing"])
        current_year = date.today().year
        if year > current_year or year < current_year - 10:
            return "❌ Invalid Year of Passing"

        # ===== QUALIFYING EXAM =====
        if admission["qualifying_exam"] not in ["SSLC","CBSE","ICSE", "PUC", "ITI"]:
            return "❌ Invalid Qualifying Exam"

        # ===== FILES (SAVE IF PROVIDED) =====
        def save_step3_file(file_obj, prefix):
            if not file_obj or file_obj.filename == "":
                return None
            if not allowed_file(file_obj.filename):
                return "INVALID_FILE"
            filename = f"{admission['admission_id']}_{prefix}_{secure_filename(file_obj.filename)}"
            file_obj.save(os.path.join(UPLOAD_FOLDER, filename))
            return filename

        photo_name = save_step3_file(request.files.get("student_photo"), "photo")
        aadhaar_file_name = save_step3_file(request.files.get("aadhaar_file"), "aadhaar")
        caste_file_name = save_step3_file(request.files.get("caste_certificate_file"), "caste")
        income_file_name = save_step3_file(request.files.get("income_certificate_file"), "income")
        marks_file_name = save_step3_file(request.files.get("marks_card_file"), "marks")

        if "INVALID_FILE" in [photo_name, aadhaar_file_name, caste_file_name, income_file_name, marks_file_name]:
            return "❌ Invalid file type. Allowed: PDF, JPG, JPEG, PNG"

        

        # ================= DATABASE =================
        db = get_db()
        cur = db.cursor()
        ensure_students_admission_year_column()

        # ===== CREATE STUDENT LOGIN =====
        final_admission_id = insert_student_login_with_retry(
            cur,
            admission_id=admission.get("admission_id"),
            student_name=admission["student_name"],
            branch=admission["branch"],
            admission_year=parse_int_prefix(admission.get("admission_year")) or datetime.today().year,
            mobile=admission["student_mobile"],
            password_hash=generate_password_hash(admission["password"]),
            status="INACTIVE",
            admission_year_text=admission.get("admission_year") or current_academic_year(),
        )
        admission["admission_id"] = final_admission_id
        session["admission"] = admission
        session.modified = True

        # ===== PERSONAL DETAILS =====
        cur.execute("""
            INSERT INTO student_personal_details (
                admission_id,
                student_mobile, student_email, disability,
                dob, gender, indian_nationality, religion,
                caste_category, alloted_category,
                qualifying_exam, year_of_passing, register_number,
                admission_quota,
                father_name, father_mobile,
                mother_name, mother_mobile,
                residential_address, permanent_address,
                aadhaar_number, caste_rd_number, income_rd_number
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                      %s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            admission["admission_id"],
            admission.get("student_mobile"),
            admission.get("student_email"),
            admission.get("disability", "NO"),
            admission["dob"],
            admission["gender"],
            admission["indian_nationality"],
            admission["religion"],
            admission["caste_category"],
            admission["alloted_category"],
            admission["qualifying_exam"],
            admission["year_of_passing"],
            admission["register_number"],
            admission["admission_quota"],
            admission["father_name"],
            admission["father_mobile"],
            admission["mother_name"],
            admission["mother_mobile"],
            admission["residential_address"],
            admission["permanent_address"],
            aadhaar_number,
            caste_rd_number,
            income_rd_number
        ))

        def _to_int_or_none(value):
            text = str(value or "").strip()
            if text == "":
                return None
            try:
                return int(float(text))
            except ValueError:
                return None

        total_max_marks = _to_int_or_none(
            admission.get("total_max_marks") or admission.get("total_marks")
        )
        total_marks_obtained = _to_int_or_none(
            admission.get("marks_obtained") or admission.get("total_marks_obtained") or admission.get("total_marks")
        )
        science_max_marks = _to_int_or_none(admission.get("science_max_marks"))
        science_marks_obtained = _to_int_or_none(
            admission.get("science_marks_obtained") or admission.get("science_marks")
        )
        maths_max_marks = _to_int_or_none(admission.get("maths_max_marks"))
        maths_marks_obtained = _to_int_or_none(
            admission.get("maths_marks_obtained") or admission.get("maths_marks")
        )

        cur.execute("""
            INSERT INTO education_details (
                admission_id, qualifying_exam, register_number, year_of_passing,
                total_max_marks, total_marks_obtained,
                science_max_marks, science_marks_obtained,
                maths_max_marks, maths_marks_obtained
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            admission["admission_id"],
            admission.get("qualifying_exam"),
            admission.get("register_number"),
            admission.get("year_of_passing"),
            total_max_marks if total_max_marks is not None else 0,
            total_marks_obtained if total_marks_obtained is not None else 0,
            science_max_marks if science_max_marks is not None else 0,
            science_marks_obtained if science_marks_obtained is not None else 0,
            maths_max_marks if maths_max_marks is not None else 0,
            maths_marks_obtained if maths_marks_obtained is not None else 0,
        ))

        # ===== STEP-3 DETAILS (AADHAAR / RD NUMBERS + FILES) =====
        cur.execute(
            "SELECT id FROM student_documents WHERE admission_id=%s",
            (admission["admission_id"],)
        )
        existing_doc = cur.fetchone()

        if existing_doc:
            cur.execute("""
                UPDATE student_documents
                SET aadhaar_number=%s,
                    caste_rd_number=%s,
                    income_rd_number=%s,
                    student_photo=COALESCE(%s, student_photo),
                    aadhaar_file=COALESCE(%s, aadhaar_file),
                    caste_file=COALESCE(%s, caste_file),
                    income_file=COALESCE(%s, income_file),
                    marks_card_file=COALESCE(%s, marks_card_file)
                WHERE admission_id=%s
            """, (
                aadhaar_number,
                caste_rd_number,
                income_rd_number,
                photo_name,
                aadhaar_file_name,
                caste_file_name,
                income_file_name,
                marks_file_name,
                admission["admission_id"]
            ))
        else:
            cur.execute("""
                INSERT INTO student_documents
                (admission_id, aadhaar_number, caste_rd_number, income_rd_number,
                 student_photo, aadhaar_file, caste_file, income_file, marks_card_file)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                admission["admission_id"],
                aadhaar_number,
                caste_rd_number,
                income_rd_number,
                photo_name,
                aadhaar_file_name,
                caste_file_name,
                income_file_name,
                marks_file_name
            ))

        cur.execute("""
            UPDATE student_personal_details
            SET aadhaar_number = %s,
                caste_rd_number = %s,
                income_rd_number = %s,
                photo_file = COALESCE(%s, photo_file),
                caste_certificate_file = COALESCE(%s, caste_certificate_file),
                income_certificate_file = COALESCE(%s, income_certificate_file),
                marks_card_file = COALESCE(%s, marks_card_file)
            WHERE admission_id = %s
        """, (
            aadhaar_number,
            caste_rd_number,
            income_rd_number,
            photo_name,
            caste_file_name,
            income_file_name,
            marks_file_name,
            admission["admission_id"]
        ))

        db.commit()
        cur.close()
        db.close()

        # CLEAR SESSION
        session.pop("admission", None)

        return render_template(
            "admission_success.html",
            admission_id=admission["admission_id"]
        )

    # =========================
    # GET → SHOW PAGE
    # =========================
    return render_template(
        "admission_step3.html",
        admission_id=admission["admission_id"]
    )

# =========================
# ADMIN: VIEW APPLICATIONS
# =========================
@app.route("/admin/applications")
def admin_applications():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"]:
        return "Forbidden: Admissions is available only for admin.", 403

    ensure_students_status_supports_rejected()
    ensure_students_rejection_reason_column()

    q = request.args.get("q", "").strip()
    status = request.args.get("status", "INACTIVE")
    branch = request.args.get("branch", "").strip()
    if scope["is_staff"]:
        branch = scope["department"]

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT
            s.admission_id,
            s.student_name,
            s.branch,
            s.status,
            COALESCE(spd.student_mobile, s.mobile) AS mobile,
            COALESCE(spd.dob, '-') AS dob,
            COALESCE(spd.gender, '-') AS gender,
            COALESCE(spd.caste_category, '-') AS caste_category,
            COALESCE(spd.alloted_category, '-') AS alloted_category,
            COALESCE(spd.register_number, '-') AS register_number,
            COALESCE(s.rejection_reason, '-') AS rejection_reason
        FROM students s
        LEFT JOIN student_personal_details spd
            ON spd.admission_id = s.admission_id
        WHERE 1=1
    """
    params = []

    if status and status != "ALL":
        query += " AND s.status=%s"
        params.append(status)

    if branch and branch != "ALL":
        query += " AND s.branch=%s"
        params.append(branch)

    if q:
        query += """
            AND (
                s.student_name LIKE %s
                OR s.admission_id LIKE %s
                OR s.college_reg_no LIKE %s
                OR spd.register_number LIKE %s
            )
        """
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])

    query += """
        ORDER BY s.student_name ASC
    """

    cur.execute(query, tuple(params))

    students = cur.fetchall()

    cur.execute("SELECT DISTINCT branch FROM students ORDER BY branch ASC")
    branch_rows = cur.fetchall()
    branches = [row["branch"] for row in branch_rows if row.get("branch")]

    cur.close()
    db.close()

    return render_template(
        "admin_applications.html",
        students=students,
        branches=branches,
        is_staff=scope["is_staff"],
        staff_department=scope["department"],
        can_set_alloted_category=can_set_allotted_category(scope),
        q=q,
        status=status,
        branch=branch
    )

# =========================
# ADMIN: APPROVE STUDENT
# =========================
@app.route("/approve/<admission_id>", methods=["GET", "POST"])

def approve_student(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"] and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can approve only your department students.", 403

    ensure_students_status_supports_rejected()
    ensure_students_rejection_reason_column()

    can_set_alloted = can_set_allotted_category(scope)
    if request.method == "POST":
        next_status = request.form.get("status", "INACTIVE")
        q = request.form.get("q", "")
        branch = request.form.get("branch", "")
        alloted_category = request.form.get("alloted_category", "").strip().upper()
    else:
        next_status = request.args.get("status", "INACTIVE")
        q = request.args.get("q", "")
        branch = request.args.get("branch", "")
        alloted_category = request.args.get("alloted_category", "").strip().upper()

    allowed_alloted_categories = {"1G", "SCG", "STG", "2AG", "2BG", "3AG", "3BG", "GM", "PENDING"}
    if alloted_category:
        if not can_set_alloted:
            return "Forbidden: Only Admin/Management can update Allotted Category.", 403
        if alloted_category not in allowed_alloted_categories:
            return "Invalid allotted category", 400

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        UPDATE students
        SET status='ACTIVE'
        WHERE admission_id=%s
    """, (admission_id,))

    if alloted_category and can_set_alloted:
        cur.execute("SELECT id FROM student_personal_details WHERE admission_id=%s LIMIT 1", (admission_id,))
        row = cur.fetchone()
        if row:
            cur.execute(
                "UPDATE student_personal_details SET alloted_category=%s WHERE admission_id=%s",
                (alloted_category, admission_id)
            )
        else:
            cur.execute(
                "INSERT INTO student_personal_details (admission_id, alloted_category) VALUES (%s, %s)",
                (admission_id, alloted_category)
            )

    db.commit()
    cur.close()
    db.close()   
    return redirect(url_for("admin_applications", status=next_status, q=q, branch=branch))

@app.route("/reject/<admission_id>", methods=["POST"])
def reject_student(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"] and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can reject only your department students.", 403

    ensure_students_status_supports_rejected()
    ensure_students_rejection_reason_column()

    reason = request.form.get("reason", "").strip()
    next_status = request.form.get("status", "REJECTED")
    q = request.form.get("q", "")
    branch = request.form.get("branch", "")

    if not reason:
        return "Reason is required"

    db = get_db()
    cur = db.cursor()

    cur.execute("""
        UPDATE students
        SET status='REJECTED', rejection_reason=%s
        WHERE admission_id=%s
    """, (reason, admission_id))

    db.commit()
    cur.close()
    db.close()   

    return redirect(url_for("admin_applications", status=next_status, q=q, branch=branch))


# ADMIN: ACADEMIC RECORDS
# =========================
@app.route("/admin/academic-records", methods=["GET", "POST"])
def admin_academic_records():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_academic_module_tables()
    ensure_subject_master_table()
    ensure_subject_series_table()

    q = (request.args.get("admission_id") or "").strip().upper()
    msg = request.args.get("msg", "")

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        admission_id = (request.form.get("admission_id") or "").strip().upper()
        semester_no = parse_int_prefix(request.form.get("semester_no")) or 1
        series_name = (request.form.get("series") or "").strip().upper()

        if not admission_id:
            return redirect(url_for("admin_academic_records", msg="Admission ID is required."))
        if scope["is_staff"] and not student_in_department(admission_id, scope["department"]):
            return "Forbidden: You can edit only your department students.", 403

        db = get_db()
        cur = db.cursor()
        try:
            cur.execute("SELECT branch FROM students WHERE admission_id=%s LIMIT 1", (admission_id,))
            student_row = cur.fetchone()
            student_branch = (student_row[0] if student_row else "") or ""
            if action == "subject":
                subject_code = (request.form.get("subject_code") or "").strip().upper()
                subject_name = (request.form.get("subject_name") or "").strip()
                internal_max = parse_int_prefix(request.form.get("internal_max")) or 25
                external_max = parse_int_prefix(request.form.get("external_max")) or 75
                master = find_subject_master_by_code(student_branch, semester_no, subject_code, series_name=series_name)
                if not master:
                    return redirect(url_for("admin_academic_records", admission_id=admission_id, msg="Select a valid subject from Subject Master."))
                subject_name = (master.get("subject_name") or "").strip()
                subject_code = (master.get("course_code") or "").strip().upper()
                if not subject_code or not subject_name:
                    return redirect(url_for("admin_academic_records", admission_id=admission_id, msg="Subject code and name are required."))
                cur.execute("""
                    INSERT INTO student_subjects
                    (admission_id, semester_no, subject_code, subject_name, internal_max, external_max)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                        subject_name=VALUES(subject_name),
                        internal_max=VALUES(internal_max),
                        external_max=VALUES(external_max)
                """, (admission_id, semester_no, subject_code, subject_name, internal_max, external_max))
                message = "Subject saved."
            elif action == "marks":
                subject_code = (request.form.get("subject_code") or "").strip().upper()
                internal_marks = float(request.form.get("internal_marks") or 0)
                external_marks = float(request.form.get("external_marks") or 0)
                if not subject_code:
                    return redirect(url_for("admin_academic_records", admission_id=admission_id, msg="Subject code is required for marks."))
                master = find_subject_master_by_code(student_branch, semester_no, subject_code, series_name=series_name)
                if not master:
                    return redirect(url_for("admin_academic_records", admission_id=admission_id, msg="Select a valid subject from Subject Master."))
                cur.execute("""
                    INSERT INTO student_internal_marks
                    (admission_id, semester_no, subject_code, internal_marks, external_marks)
                    VALUES (%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                        internal_marks=VALUES(internal_marks),
                        external_marks=VALUES(external_marks)
                """, (admission_id, semester_no, subject_code, internal_marks, external_marks))
                message = "Marks saved."
            elif action == "attendance":
                subject_code = (request.form.get("subject_code") or "").strip().upper()
                total_classes = parse_int_prefix(request.form.get("total_classes")) or 0
                present_classes = parse_int_prefix(request.form.get("present_classes")) or 0
                if not subject_code:
                    return redirect(url_for("admin_academic_records", admission_id=admission_id, msg="Subject code is required for attendance."))
                master = find_subject_master_by_code(student_branch, semester_no, subject_code, series_name=series_name)
                if not master:
                    return redirect(url_for("admin_academic_records", admission_id=admission_id, msg="Select a valid subject from Subject Master."))
                if present_classes > total_classes:
                    return redirect(url_for("admin_academic_records", admission_id=admission_id, msg="Present classes cannot be greater than total classes."))
                cur.execute("""
                    INSERT INTO student_attendance
                    (admission_id, semester_no, subject_code, total_classes, present_classes)
                    VALUES (%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                        total_classes=VALUES(total_classes),
                        present_classes=VALUES(present_classes)
                """, (admission_id, semester_no, subject_code, total_classes, present_classes))
                message = "Attendance saved."
            else:
                return redirect(url_for("admin_academic_records", admission_id=admission_id, msg="Invalid academic action."))
            db.commit()
        finally:
            cur.close()
            db.close()

        return redirect(url_for("admin_academic_records", admission_id=admission_id, msg=message))

    student = None
    records = []
    default_semester = 1
    if q:
        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT admission_id, student_name, branch, year_sem, admission_year FROM students WHERE admission_id=%s", (q,))
        student = cur.fetchone()
        if student and scope["is_staff"] and (student.get("branch") or "").strip().lower() != (scope["department"] or "").strip().lower():
            cur.close()
            db.close()
            return "Forbidden: You can view only your department students.", 403

        if student:
            semester_no = infer_current_sem(student.get("admission_year"), student.get("year_sem"))
            default_semester = semester_no
            cur.execute("""
                SELECT
                    ss.semester_no,
                    ss.subject_code,
                    ss.subject_name,
                    ss.internal_max,
                    ss.external_max,
                    COALESCE(sm.internal_marks, 0) AS internal_marks,
                    COALESCE(sm.external_marks, 0) AS external_marks,
                    COALESCE(sa.total_classes, 0) AS total_classes,
                    COALESCE(sa.present_classes, 0) AS present_classes
                FROM student_subjects ss
                LEFT JOIN student_internal_marks sm
                    ON sm.admission_id=ss.admission_id
                    AND sm.semester_no=ss.semester_no
                    AND sm.subject_code=ss.subject_code
                LEFT JOIN student_attendance sa
                    ON sa.admission_id=ss.admission_id
                    AND sa.semester_no=ss.semester_no
                    AND sa.subject_code=ss.subject_code
                WHERE ss.admission_id=%s AND ss.semester_no=%s
                ORDER BY ss.subject_name ASC
            """, (q, semester_no))
            records = cur.fetchall()

        cur.close()
        db.close()

    return render_template(
        "admin_academic_records.html",
        q=q,
        student=student,
        records=records,
        default_semester=default_semester,
        branch_options=BRANCH_OPTIONS,
        series_options=fetch_subject_series_options(),
        msg=msg,
        is_staff=scope["is_staff"],
        staff_department=scope["department"],
    )


# =========================
# ADMIN: ADD STUDENT
# =========================
@app.route("/add-student", methods=["GET", "POST"])
def add_student():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    can_set_alloted = can_set_allotted_category(scope)

    if request.method == "POST":
        name = request.form.get("student_name", "").strip().upper()
        branch = request.form.get("branch", "").strip()
        admission_year_text = request.form.get("admission_year", "").strip() or current_academic_year()
        if scope["is_staff"]:
            branch = scope["department"]
        if not name or not branch:
            return "Student name and branch are required", 400

        if not re.match(r"^\d{4}-\d{2}$", admission_year_text):
            return "Admission year must be in YYYY-YY format (example: 2026-27)", 400
        admission_year = int(admission_year_text[:4])
        admission_id = generate_admission_id(branch, admission_year_text)
        manual_sem_raw = request.form.get("year_sem", "").strip()
        manual_sem = None
        if manual_sem_raw:
            manual_sem = parse_int_prefix(manual_sem_raw)
            if manual_sem is None or manual_sem < 1 or manual_sem > 6:
                return "Manual semester must be between 1 and 6", 400
        student_mobile = request.form.get("student_mobile", "").strip()
        student_email = request.form.get("student_email", "").strip().lower()
        password_raw = request.form.get("password", "").strip()
        password_hash = generate_password_hash(password_raw)

        # Online step-1/2 equivalent fields
        dob = request.form.get("dob", "").strip()
        gender = request.form.get("gender", "").strip()
        indian_nationality = request.form.get("indian_nationality", "").strip().upper()
        religion = request.form.get("religion", "").strip().upper()
        caste_category = request.form.get("caste_category", "").strip().upper()
        alloted_category = request.form.get("alloted_category", "").strip().upper()
        if not can_set_alloted:
            alloted_category = "PENDING"
        admission_quota = request.form.get("admission_quota", "").strip().upper()
        register_number = request.form.get("register_number", "").strip().upper()
        year_of_passing = request.form.get("year_of_passing", "").strip()
        qualifying_exam = request.form.get("qualifying_exam", "").strip().upper()

        father_name = request.form.get("father_name", "").strip().upper()
        father_mobile = request.form.get("father_mobile", "").strip()
        mother_name = request.form.get("mother_name", "").strip().upper()
        mother_mobile = request.form.get("mother_mobile", "").strip()
        residential_address = request.form.get("residential_address", "").strip().upper()
        permanent_address = request.form.get("permanent_address", "").strip().upper()

        # Education details table mapping
        def first_non_empty(*keys):
            for key in keys:
                for raw_val in request.form.getlist(key):
                    val = str(raw_val or "").strip()
                    if val:
                        return val
            return ""

        def to_int_or_none(val):
            text = str(val or "").strip()
            if not text:
                return None
            try:
                return int(float(text))
            except ValueError:
                return None

        total_max_marks = to_int_or_none(first_non_empty("total_marks", "total_max_marks"))
        total_marks_obtained = to_int_or_none(first_non_empty("marks_obtained", "total_marks_obtained"))
        science_max_marks = to_int_or_none(first_non_empty("science_max_marks"))
        science_marks_obtained = to_int_or_none(first_non_empty("science_marks", "science_marks_obtained"))
        maths_max_marks = to_int_or_none(first_non_empty("maths_max_marks"))
        maths_marks_obtained = to_int_or_none(first_non_empty("maths_marks", "maths_marks_obtained"))

        aadhaar_number = request.form.get("aadhaar_number", "").strip().upper()
        caste_rd_number = request.form.get("caste_rd_number", "").strip().upper()
        income_rd_number = request.form.get("income_rd_number", "").strip().upper()

        if qualifying_exam not in ["SSLC", "PUC", "ITI", "CBSE", "ICSE"]:
            return "Invalid qualifying exam", 400
        if not student_mobile or not student_email or not password_raw:
            return "Student mobile, email and password are required", 400
        if not dob or not gender or not caste_category:
            return "Missing mandatory personal details", 400
        if can_set_alloted and not alloted_category:
            return "Alloted category is required for Admin/Management admission entry", 400
        if not register_number or not year_of_passing:
            return "Register number and year of passing are required", 400
        if not admission_quota:
            return "Admission quota is required", 400
        if not father_name or not father_mobile or not mother_name:
            return "Parent details are required", 400
        if not residential_address or not permanent_address:
            return "Address details are required", 400
        if not aadhaar_number or not caste_rd_number or not income_rd_number:
            return "Document numbers are required", 400

        def save_offline_doc(file_obj, prefix):
            if not file_obj or file_obj.filename == "":
                return None
            if not allowed_file(file_obj.filename):
                return "INVALID_FILE"
            filename = f"{admission_id}_{prefix}_{secure_filename(file_obj.filename)}"
            file_obj.save(os.path.join(UPLOAD_FOLDER, filename))
            return filename

        photo_name = save_offline_doc(request.files.get("student_photo"), "photo")
        aadhaar_file_name = save_offline_doc(request.files.get("aadhaar_file"), "aadhaar")
        caste_file_name = save_offline_doc(request.files.get("caste_certificate_file") or request.files.get("caste_file"), "caste")
        income_file_name = save_offline_doc(request.files.get("income_certificate_file") or request.files.get("income_file"), "income")
        marks_file_name = save_offline_doc(request.files.get("marks_card_file"), "marks")

        if "INVALID_FILE" in [photo_name, aadhaar_file_name, caste_file_name, income_file_name, marks_file_name]:
            return "Invalid file type. Allowed: pdf, jpg, jpeg, png", 400

        db = get_db()
        cur = db.cursor(dictionary=True)
        try:
            ensure_qualifying_exam_support()
            ensure_students_admission_year_column()
            ensure_students_year_sem_column()
            admission_id = insert_student_login_with_retry(
                cur,
                admission_id=admission_id,
                student_name=name,
                branch=branch,
                admission_year=admission_year,
                year_sem=manual_sem,
                mobile=student_mobile,
                password_hash=password_hash,
                status="ACTIVE",
                admission_year_text=admission_year_text,
            )

            cur.execute("""
                INSERT INTO student_personal_details (
                    admission_id,
                    student_mobile,
                    student_email,
                    dob, gender, indian_nationality, religion,
                    caste_category, alloted_category,
                    qualifying_exam, year_of_passing, register_number,
                    admission_quota,
                    father_name, father_mobile,
                    mother_name, mother_mobile,
                    residential_address, permanent_address,
                    aadhaar_number, caste_rd_number, income_rd_number,
                    photo_file, caste_certificate_file, income_certificate_file, marks_card_file
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                admission_id,
                student_mobile,
                student_email,
                dob,
                gender,
                indian_nationality,
                religion,
                caste_category,
                alloted_category,
                qualifying_exam,
                year_of_passing,
                register_number,
                admission_quota,
                father_name,
                father_mobile,
                mother_name,
                mother_mobile,
                residential_address,
                permanent_address,
                aadhaar_number,
                caste_rd_number,
                income_rd_number,
                photo_name,
                caste_file_name,
                income_file_name,
                marks_file_name
            ))

            cur.execute("""
                INSERT INTO education_details
                (admission_id, qualifying_exam, register_number, year_of_passing,
                 total_max_marks, total_marks_obtained,
                 science_max_marks, science_marks_obtained,
                 maths_max_marks, maths_marks_obtained)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                admission_id,
                qualifying_exam,
                register_number,
                year_of_passing,
                total_max_marks,
                total_marks_obtained,
                science_max_marks,
                science_marks_obtained,
                maths_max_marks,
                maths_marks_obtained
            ))

            cur.execute("""
                INSERT INTO student_documents
                (admission_id, aadhaar_number, caste_rd_number, income_rd_number,
                 student_photo, aadhaar_file, caste_file, income_file, marks_card_file)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                admission_id,
                aadhaar_number,
                caste_rd_number,
                income_rd_number,
                photo_name,
                aadhaar_file_name,
                caste_file_name,
                income_file_name,
                marks_file_name
            ))

            db.commit()
        except Exception as exc:
            db.rollback()
            return f"Unable to save offline admission: {exc}", 400
        finally:
            cur.close()
            db.close()

        return redirect(url_for("admin_view_student", admission_id=admission_id))

    return render_template(
        "add_student.html",
        is_staff=scope["is_staff"],
        staff_department=scope["department"],
        can_set_alloted_category=can_set_alloted,
        admission_year=current_academic_year()
    )


# =========================
# ADMIN: ADD EDUCATION
# =========================
@app.route("/add-education", methods=["GET", "POST"])
def add_education():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    return redirect("/add-student")


# =========================
# ADMIN: FEES MANAGEMENT
# =========================
def fetch_fee_overview_rows(q="", branch="", sem="", payment_state="", sort_by="name", forced_department=""):
    ensure_students_college_reg_no_column()
    ensure_students_admission_year_column()
    ensure_students_year_sem_column()
    ensure_fee_module_tables()

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT
            s.admission_id,
            s.student_name,
            s.branch,
            s.mobile,
            s.year_sem,
            s.admission_year,
            COALESCE(s.college_reg_no, '') AS college_reg_no
        FROM students s
        WHERE 1=1
    """
    params = []

    if q:
        query += """
            AND (
                s.student_name LIKE %s
                OR s.admission_id LIKE %s
                OR s.college_reg_no LIKE %s
            )
        """
        like_q = f"%{q}%"
        params.extend([like_q, like_q, like_q])

    if forced_department:
        query += " AND s.branch=%s"
        params.append(forced_department)
    elif branch:
        query += " AND s.branch=%s"
        params.append(branch)

    cur.execute(query, tuple(params))
    raw_rows = cur.fetchall()

    cur.execute("SELECT DISTINCT branch FROM students ORDER BY branch ASC")
    branches = [row["branch"] for row in cur.fetchall() if row.get("branch")]

    ay = current_academic_year()
    cur.execute("""
        SELECT
            branch,
            semester_no,
            admission_fee_due,
            tuition_fee_due,
            management_fee_due,
            exam_fee_due
        FROM fee_structure_master
        WHERE academic_year=%s
    """, (ay,))
    structure_rows = cur.fetchall()
    fee_map = {}
    for fr in structure_rows:
        fee_map[(fr["branch"], int(fr["semester_no"]))] = {
            "ADMISSION": float(fr.get("admission_fee_due") or 0),
            "TUITION": float(fr.get("tuition_fee_due") or 0),
            "MANAGEMENT": float(fr.get("management_fee_due") or 0),
            "EXAM": float(fr.get("exam_fee_due") or 0),
        }

    cur.execute("""
        SELECT
            admission_id,
            fee_type,
            semester_no,
            COALESCE(SUM(amount), 0) AS total_amount
        FROM fee_payments
        WHERE academic_year=%s
        GROUP BY admission_id, fee_type, semester_no
    """, (ay,))
    payment_rows = cur.fetchall()
    payment_map = {}
    for pr in payment_rows:
        sem_no = int(pr["semester_no"]) if pr.get("semester_no") else None
        if sem_no is None:
            continue
        payment_map[(pr["admission_id"], sem_no, pr["fee_type"])] = float(pr.get("total_amount") or 0)

    cur.close()
    db.close()

    rows = []
    for row in raw_rows:
        current_sem = infer_current_sem(row.get("admission_year"), row.get("year_sem"))
        due = fee_map.get((row.get("branch"), current_sem), {})
        paid = {
            "ADMISSION": payment_map.get((row.get("admission_id"), current_sem, "ADMISSION"), 0),
            "TUITION": payment_map.get((row.get("admission_id"), current_sem, "TUITION"), 0),
            "MANAGEMENT": payment_map.get((row.get("admission_id"), current_sem, "MANAGEMENT"), 0),
            "EXAM": payment_map.get((row.get("admission_id"), current_sem, "EXAM"), 0),
        }
        rows.append(fee_summary_from_row(row, due=due, paid=paid))

    if sem:
        try:
            sem_value = int(sem)
            rows = [row for row in rows if row["current_sem"] == sem_value]
        except ValueError:
            pass

    if payment_state:
        state = payment_state.strip().upper()
        rows = [row for row in rows if row["payment_state"] == state]

    sort_key_map = {
        "name": lambda row: (row.get("student_name") or "").lower(),
        "admission_id": lambda row: (row.get("admission_id") or "").lower(),
        "branch": lambda row: (row.get("branch") or "").lower(),
        "sem_asc": lambda row: row.get("current_sem") or 0,
        "sem_desc": lambda row: -(row.get("current_sem") or 0),
        "balance_desc": lambda row: -(row.get("balance") or 0),
        "balance_asc": lambda row: row.get("balance") or 0,
    }
    rows.sort(key=sort_key_map.get(sort_by, sort_key_map["name"]))
    return rows, branches


@app.route("/admin/fees", methods=["GET", "POST"])
def admin_fees():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    can_manage = can_edit_fees(scope)
    is_management_staff = scope["is_staff"] and "management" in (scope.get("department") or "").strip().lower()

    ensure_students_college_reg_no_column()
    ensure_students_admission_year_column()
    ensure_students_year_sem_column()
    ensure_fee_module_tables()

    msg = request.args.get("msg", "")
    ay = current_academic_year()

    if request.method == "POST":
        if not can_manage:
            return "Forbidden: Only Admin, HOD, or Management staff can add/edit fee details.", 403
        fee_branch = request.form.get("fee_branch", "").strip()
        if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower():
            fee_branch = scope["department"]
        semester_no = parse_int_prefix(request.form.get("semester_no", "").strip())
        academic_year = request.form.get("academic_year", "").strip() or ay
        admission_fee_due = float(request.form.get("admission_fee_due") or 0)
        tuition_fee_due = float(request.form.get("tuition_fee_due") or 0)
        management_fee_due = float(request.form.get("management_fee_due") or 0)
        exam_fee_due = float(request.form.get("exam_fee_due") or 0)

        if not fee_branch:
            return redirect(url_for("admin_fees", msg="Branch is required for fee setup"))
        if semester_no is None or semester_no < 1 or semester_no > 6:
            return redirect(url_for("admin_fees", msg="Semester must be between 1 and 6"))

        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("""
            INSERT INTO fee_structure_master (
                branch,
                semester_no,
                academic_year,
                admission_fee_due,
                tuition_fee_due,
                management_fee_due,
                exam_fee_due
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                admission_fee_due=VALUES(admission_fee_due),
                tuition_fee_due=VALUES(tuition_fee_due),
                management_fee_due=VALUES(management_fee_due),
                exam_fee_due=VALUES(exam_fee_due)
        """, (
            fee_branch,
            semester_no,
            academic_year,
            admission_fee_due,
            tuition_fee_due,
            management_fee_due,
            exam_fee_due
        ))

        db.commit()
        cur.close()
        db.close()
        return redirect(url_for("admin_fees", msg="Fee structure saved for branch/semester"))

    q = request.args.get("q", "").strip()
    branch = request.args.get("branch", "").strip()
    if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower():
        branch = scope["department"]
    sem = request.args.get("sem", "").strip()
    payment_state = request.args.get("payment_state", "").strip().upper()
    sort_by = request.args.get("sort_by", "name").strip()

    students, branches = fetch_fee_overview_rows(
        q=q,
        branch=branch,
        sem=sem,
        payment_state=payment_state,
        sort_by=sort_by,
        forced_department=scope["department"] if (scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower()) else ""
    )

    return render_template(
        "admin_fees.html",
        students=students,
        branches=branches,
        filters={
            "q": q,
            "branch": branch,
            "sem": sem,
            "payment_state": payment_state,
            "sort_by": sort_by
        },
        default_academic_year=ay,
        is_staff=scope["is_staff"],
        is_management_staff=is_management_staff,
        staff_department=scope["department"],
        can_manage_fees=can_manage,
        msg=msg
    )


@app.route("/admin/fees/student/<admission_id>/delete", methods=["POST"])
def admin_fees_delete_student(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if not can_edit_fees(scope):
        return "Forbidden: Only Admin, HOD, or Management staff can delete students from fees.", 403
    if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower() and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can delete only your department students.", 403

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT admission_id FROM students WHERE admission_id=%s", (admission_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        db.close()
        return redirect(url_for("admin_fees", msg="Student not found"))

    delete_cur = db.cursor()
    try:
        delete_cur.execute("DELETE FROM fee_payments WHERE admission_id=%s", (admission_id,))
        delete_cur.execute("DELETE FROM student_fee_structure WHERE admission_id=%s", (admission_id,))
        delete_cur.execute("DELETE FROM fees WHERE admission_id=%s", (admission_id,))
        delete_cur.execute("DELETE FROM education_details WHERE admission_id=%s", (admission_id,))
        delete_cur.execute("DELETE FROM student_documents WHERE admission_id=%s", (admission_id,))
        delete_cur.execute("DELETE FROM student_personal_details WHERE admission_id=%s", (admission_id,))
        delete_cur.execute("DELETE FROM students WHERE admission_id=%s", (admission_id,))
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        delete_cur.close()
        cur.close()
        db.close()

    return redirect(url_for("admin_fees", msg="Student deleted successfully"))


@app.route("/admin/fees/export")
def admin_fees_export():
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    q = request.args.get("q", "").strip()
    branch = request.args.get("branch", "").strip()
    if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower():
        branch = scope["department"]
    sem = request.args.get("sem", "").strip()
    payment_state = request.args.get("payment_state", "").strip().upper()
    sort_by = request.args.get("sort_by", "name").strip()

    rows, _ = fetch_fee_overview_rows(
        q=q,
        branch=branch,
        sem=sem,
        payment_state=payment_state,
        sort_by=sort_by,
        forced_department=scope["department"] if (scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower()) else ""
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Admission ID",
        "Reg No",
        "Student Name",
        "Branch",
        "Current Sem",
        "Academic Year",
        "Admission Due",
        "Tuition Due",
        "Management Due",
        "Exam Due",
        "Total Due",
        "Admission Paid",
        "Tuition Paid",
        "Management Paid",
        "Exam Paid",
        "Total Paid",
        "Balance",
        "Payment State"
    ])

    for row in rows:
        writer.writerow([
            row.get("admission_id", ""),
            row.get("college_reg_no", ""),
            row.get("student_name", ""),
            row.get("branch", ""),
            row.get("current_sem", ""),
            row.get("academic_year", ""),
            row.get("admission_due_total", 0),
            row.get("tuition_due_total", 0),
            row.get("management_due_total", 0),
            row.get("exam_due_total", 0),
            row.get("total_due", 0),
            row.get("admission_paid", 0),
            row.get("tuition_paid", 0),
            row.get("management_paid", 0),
            row.get("exam_paid", 0),
            row.get("total_paid", 0),
            row.get("balance", 0),
            row.get("payment_state", ""),
        ])

    csv_bytes = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    csv_bytes.seek(0)
    filename = f"fees_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(
        csv_bytes,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename
    )


@app.route("/admin/fees/student/<admission_id>")
def admin_fees_student_history(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower() and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can view only your department students.", 403

    ensure_students_college_reg_no_column()
    ensure_students_admission_year_column()
    ensure_students_year_sem_column()
    ensure_fee_module_tables()

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT
            s.admission_id,
            s.student_name,
            s.branch,
            s.mobile,
            s.year_sem,
            s.admission_year,
            COALESCE(s.college_reg_no, '') AS college_reg_no
        FROM students s
        WHERE s.admission_id=%s
    """, (admission_id,))
    student = cur.fetchone()

    if not student:
        cur.close()
        db.close()
        return "Student not found", 404

    current_sem = infer_current_sem(student.get("admission_year"), student.get("year_sem"))
    ay = current_academic_year()
    cur.execute("""
        SELECT
            admission_fee_due,
            tuition_fee_due,
            management_fee_due,
            exam_fee_due
        FROM fee_structure_master
        WHERE branch=%s AND semester_no=%s AND academic_year=%s
    """, (student.get("branch"), current_sem, ay))
    structure = cur.fetchone() or {}

    due = {
        "ADMISSION": float(structure.get("admission_fee_due") or 0),
        "TUITION": float(structure.get("tuition_fee_due") or 0),
        "MANAGEMENT": float(structure.get("management_fee_due") or 0),
        "EXAM": float(structure.get("exam_fee_due") or 0),
    }

    cur.execute("""
        SELECT
            fee_type,
            COALESCE(SUM(amount), 0) AS total_amount
        FROM fee_payments
        WHERE admission_id=%s AND academic_year=%s AND semester_no=%s
        GROUP BY fee_type
    """, (admission_id, ay, current_sem))
    paid_rows = cur.fetchall()
    paid = {"ADMISSION": 0, "TUITION": 0, "MANAGEMENT": 0, "EXAM": 0}
    for pr in paid_rows:
        paid[pr["fee_type"]] = float(pr.get("total_amount") or 0)

    summary = fee_summary_from_row(student, due=due, paid=paid)

    cur.execute("""
        SELECT
            id,
            admission_id,
            fee_type,
            academic_year,
            semester_no,
            amount,
            payment_date,
            receipt_no,
            remarks
        FROM fee_payments
        WHERE admission_id=%s
        ORDER BY payment_date DESC, id DESC
    """, (admission_id,))
    payments = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "admin_fee_history.html",
        student=summary,
        payments=payments,
        default_academic_year=ay,
        current_sem=current_sem,
        today_iso=datetime.today().strftime("%Y-%m-%d"),
        can_manage_fees=can_edit_fees(scope),
        msg=request.args.get("msg", "")
    )


@app.route("/admin/fees/student/<admission_id>/add-payment", methods=["POST"])
def admin_fees_add_payment(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if not can_edit_fees(scope):
        return "Forbidden: Only Admin, HOD, or Management staff can add/edit fee details.", 403
    if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower() and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can add payment only for your department students.", 403

    ensure_fee_module_tables()
    ensure_students_year_sem_column()

    fee_type = request.form.get("fee_type", "").strip().upper()
    if fee_type not in {"ADMISSION", "TUITION", "MANAGEMENT", "EXAM"}:
        return redirect(url_for("admin_fees_student_history", admission_id=admission_id, msg="Invalid fee type"))

    amount = float(request.form.get("amount") or 0)
    if amount <= 0:
        return redirect(url_for("admin_fees_student_history", admission_id=admission_id, msg="Amount must be greater than 0"))

    academic_year = request.form.get("academic_year", "").strip() or current_academic_year()
    semester_raw = request.form.get("semester_no", "").strip()
    semester_no = parse_int_prefix(semester_raw)
    if semester_no is None:
        cur_date = datetime.today().date()
        db_tmp = get_db()
        cur_tmp = db_tmp.cursor(dictionary=True)
        cur_tmp.execute("SELECT admission_year, year_sem FROM students WHERE admission_id=%s", (admission_id,))
        sem_src = cur_tmp.fetchone() or {}
        cur_tmp.close()
        db_tmp.close()
        semester_no = infer_current_sem(sem_src.get("admission_year"), sem_src.get("year_sem"), today=cur_date)
    semester_no = min(max(semester_no, 1), 6)

    payment_date = request.form.get("payment_date", "").strip() or datetime.today().strftime("%Y-%m-%d")
    remarks = request.form.get("remarks", "").strip()

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT admission_id FROM students WHERE admission_id=%s", (admission_id,))
    student = cur.fetchone()
    if not student:
        cur.close()
        db.close()
        return redirect(url_for("admin_fees", msg="Student not found"))

    receipt_no = f"SVP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{random.randint(100, 999)}"
    cur.execute("""
        INSERT INTO fee_payments (
            admission_id,
            fee_type,
            academic_year,
            semester_no,
            amount,
            payment_date,
            receipt_no,
            remarks
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        admission_id,
        fee_type,
        academic_year,
        semester_no,
        amount,
        payment_date,
        receipt_no,
        remarks
    ))
    db.commit()
    cur.close()
    db.close()
    return redirect(url_for("admin_fees_student_history", admission_id=admission_id, msg="Payment recorded successfully"))


@app.route("/admin/fees/payment/<int:payment_id>/receipt")
def admin_fees_payment_receipt(payment_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")

    ensure_fee_module_tables()

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT
            fp.*,
            s.student_name,
            s.branch
        FROM fee_payments fp
        JOIN students s ON s.admission_id = fp.admission_id
        WHERE fp.id=%s
    """, (payment_id,))
    payment = cur.fetchone()
    cur.close()
    db.close()

    if not payment:
        return "Payment not found", 404
    if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower() and (payment.get("branch") or "").strip().lower() != (scope["department"] or "").strip().lower():
        return "Forbidden: You can access only your department receipts.", 403

    receipt_path = generate_fee_receipt(
        {
            "admission_id": payment["admission_id"],
            "student_name": payment["student_name"],
            "branch": payment["branch"],
        },
        {
            "admission_fee": payment["amount"] if payment["fee_type"] == "ADMISSION" else 0,
            "tuition_fee": payment["amount"] if payment["fee_type"] == "TUITION" else 0,
            "management_fee": payment["amount"] if payment["fee_type"] == "MANAGEMENT" else 0,
            "exam_fee": payment["amount"] if payment["fee_type"] == "EXAM" else 0,
            "payment_type": payment["fee_type"],
            "payment_status": payment["fee_type"],
            "receipt_no": payment["receipt_no"],
            "payment_date": payment["payment_date"],
            "academic_year": payment.get("academic_year"),
            "semester_no": payment.get("semester_no"),
        }
    )
    return send_file(
        receipt_path,
        as_attachment=True,
        download_name=f"receipt_{payment['receipt_no']}.pdf"
    )


@app.route("/admin/fees/payment/<int:payment_id>/delete", methods=["POST"])
def admin_fees_delete_payment(payment_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if not can_edit_fees(scope):
        return "Forbidden: Only Admin, HOD, or Management staff can delete fee entries.", 403

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT fp.id, fp.admission_id, s.branch
        FROM fee_payments fp
        JOIN students s ON s.admission_id = fp.admission_id
        WHERE fp.id=%s
    """, (payment_id,))
    payment = cur.fetchone()
    if not payment:
        cur.close()
        db.close()
        return redirect(url_for("admin_fees", msg="Payment record not found"))

    if scope["is_staff"] and "management" not in (scope.get("department") or "").strip().lower() and (payment.get("branch") or "").strip().lower() != (scope.get("department") or "").strip().lower():
        cur.close()
        db.close()
        return "Forbidden: You can delete only your department fee entries.", 403

    admission_id = payment.get("admission_id")
    cur.execute("DELETE FROM fee_payments WHERE id=%s", (payment_id,))
    db.commit()
    cur.close()
    db.close()
    return redirect(url_for("admin_fees_student_history", admission_id=admission_id, msg="Payment deleted successfully"))


# =========================
# STUDENT DASHBOARD
# =========================
@app.route("/student")
def student_dashboard():
    if "student" not in session:
        return redirect("/")

    admission_id = session["student"]
    ensure_academic_module_tables()
    ensure_fee_module_tables()
    ensure_student_attendance_table()
    ensure_subject_master_table()
    ensure_students_admission_year_column()

    student, personal, education, docs = fetch_student_full_bundle(admission_id)
    if not student:
        session.clear()
        return redirect("/login/student")

    student_view = dict(student)
    student_view.update({
        "student_email": (personal or {}).get("student_email"),
        "dob": (personal or {}).get("dob"),
        "gender": (personal or {}).get("gender"),
        "student_mobile": (personal or {}).get("student_mobile"),
        "caste_category": (personal or {}).get("caste_category"),
        "alloted_category": (personal or {}).get("alloted_category"),
        "admission_quota": (personal or {}).get("admission_quota"),
        "residential_address": (personal or {}).get("residential_address"),
        "qualifying_exam": (personal or {}).get("qualifying_exam"),
        "year_of_passing": (personal or {}).get("year_of_passing"),
        "register_number": (personal or {}).get("register_number"),
        "total_max_marks": (education or {}).get("total_max_marks"),
        "total_marks_obtained": (education or {}).get("total_marks_obtained"),
        "science_marks_obtained": (education or {}).get("science_marks_obtained"),
        "maths_marks_obtained": (education or {}).get("maths_marks_obtained"),
        "student_photo": (docs or {}).get("student_photo"),
    })

    current_sem = infer_current_sem(student.get("admission_year"), student.get("year_sem"))

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT
            ss.subject_code,
            ss.subject_name,
            ss.internal_max,
            ss.external_max,
            COALESCE(sm.internal_marks, 0) AS internal_marks,
            COALESCE(sm.external_marks, 0) AS external_marks,
            COALESCE(sa.total_classes, 0) AS total_classes,
            COALESCE(sa.present_classes, 0) AS present_classes
        FROM student_subjects ss
        LEFT JOIN student_internal_marks sm
            ON sm.admission_id=ss.admission_id
            AND sm.semester_no=ss.semester_no
            AND sm.subject_code=ss.subject_code
        LEFT JOIN student_attendance sa
            ON sa.admission_id=ss.admission_id
            AND sa.semester_no=ss.semester_no
            AND sa.subject_code=ss.subject_code
        WHERE ss.admission_id=%s AND ss.semester_no=%s
        ORDER BY ss.subject_name ASC
    """, (admission_id, current_sem))
    records = cur.fetchall()

    cur.execute("""
        SELECT
            id,
            payment_date,
            fee_type,
            amount,
            receipt_no
        FROM fee_payments
        WHERE admission_id=%s
        ORDER BY payment_date DESC, id DESC
    """, (admission_id,))
    payment_history = cur.fetchall()

    cur.execute("""
        SELECT subject_name,
               SUM(CASE WHEN status IN ('PRESENT','LATE') THEN 1 ELSE 0 END) AS present_classes,
               SUM(CASE WHEN status='ABSENT' THEN 1 ELSE 0 END) AS absent_classes,
               SUM(CASE WHEN status='LEAVE' THEN 1 ELSE 0 END) AS leave_classes,
               COUNT(*) AS total_classes
        FROM student_daily_attendance
        WHERE admission_id=%s AND semester_no=%s
        GROUP BY subject_name
        ORDER BY subject_name ASC
    """, (admission_id, current_sem))
    daily_subject_attendance = cur.fetchall()

    cur.execute("""
        SELECT attendance_date, subject_name, period_no, status, remarks,
               COALESCE(NULLIF(marked_by_name, ''), marked_by) AS marked_by
        FROM student_daily_attendance
        WHERE admission_id=%s AND semester_no=%s
        ORDER BY attendance_date DESC, period_no DESC, id DESC
        LIMIT 40
    """, (admission_id, current_sem))
    daily_attendance_log = cur.fetchall()

    cur.execute("""
        SELECT subject_name, course_code
        FROM subjects
        WHERE branch=%s AND semester=%s
    """, (student.get("branch"), current_sem))
    subject_master_rows = cur.fetchall()
    cur.close()
    db.close()

    subject_rows = []
    total_scored = 0.0
    total_maximum = 0.0
    total_classes = 0
    total_present = 0
    for item in records:
        internal_max = float(item.get("internal_max") or 25)
        external_max = float(item.get("external_max") or 75)
        internal_marks = float(item.get("internal_marks") or 0)
        external_marks = float(item.get("external_marks") or 0)
        subject_total = internal_marks + external_marks
        max_total = internal_max + external_max
        attendance_total = int(item.get("total_classes") or 0)
        attendance_present = int(item.get("present_classes") or 0)
        attendance_pct = round((attendance_present / attendance_total) * 100, 1) if attendance_total > 0 else 0.0

        subject_rows.append({
            **item,
            "total_marks": round(subject_total, 2),
            "max_total": round(max_total, 2),
            "grade": marks_grade((subject_total / max_total) * 100 if max_total > 0 else 0),
            "attendance_pct": attendance_pct,
        })
        total_scored += subject_total
        total_maximum += max_total
        total_classes += attendance_total
        total_present += attendance_present

    subject_code_map = {}
    for srow in subject_master_rows:
        key = normalize_subject_name(srow.get("subject_name", "")).lower()
        if key and key not in subject_code_map:
            subject_code_map[key] = (srow.get("course_code") or "").strip().upper()

    attendance_rows = []
    daily_total_classes = 0
    daily_total_present = 0
    daily_total_absent = 0
    daily_total_leave = 0
    for row in daily_subject_attendance:
        total_cls = int(row.get("total_classes") or 0)
        present_cls = int(row.get("present_classes") or 0)
        absent_cls = int(row.get("absent_classes") or 0)
        leave_cls = int(row.get("leave_classes") or 0)
        attendance_pct = round((present_cls / total_cls) * 100, 1) if total_cls > 0 else 0.0
        subject_name = row.get("subject_name") or "-"
        key = normalize_subject_name(subject_name).lower()
        attendance_rows.append({
            "subject_name": subject_name,
            "subject_code": subject_code_map.get(key, "-"),
            "present_classes": present_cls,
            "absent_classes": absent_cls,
            "leave_classes": leave_cls,
            "total_classes": total_cls,
            "attendance_pct": attendance_pct,
        })
        daily_total_classes += total_cls
        daily_total_present += present_cls
        daily_total_absent += absent_cls
        daily_total_leave += leave_cls

    marks_percentage = round((total_scored / total_maximum) * 100, 1) if total_maximum > 0 else 0.0
    overall_attendance = round((daily_total_present / daily_total_classes) * 100, 1) if daily_total_classes > 0 else round((total_present / total_classes) * 100, 1) if total_classes > 0 else 0.0
    cgpa = round(min(marks_percentage / 10.0, 10.0), 2)

    fee_calc = fetch_student_fee_summary(admission_id, student) or {}
    paid_total = round(float(fee_calc.get("total_paid") or 0), 2)
    due_total = round(float(fee_calc.get("total_due") or 0), 2)
    pending_total = round(max(due_total - paid_total, 0), 2)

    for row in payment_history:
        row["status"] = "Paid"

    return render_template(
        "student_dashboard.html",
        student=student_view,
        current_sem=current_sem,
        subject_rows=subject_rows,
        attendance_rows=attendance_rows,
        attendance_log_rows=daily_attendance_log,
        marks_summary={
            "total_scored": round(total_scored, 2),
            "total_maximum": round(total_maximum, 2),
            "percentage": marks_percentage,
            "cgpa": cgpa,
            "grade": marks_grade(marks_percentage),
        },
        attendance_summary={
            "overall_pct": overall_attendance,
            "total_classes": daily_total_classes if daily_total_classes > 0 else total_classes,
            "present_classes": daily_total_present if daily_total_classes > 0 else total_present,
            "absent_classes": daily_total_absent if daily_total_classes > 0 else max(total_classes - total_present, 0),
            "leave_classes": daily_total_leave if daily_total_classes > 0 else 0,
        },
        fee_summary={
            "paid_total": paid_total,
            "pending_total": pending_total,
            "due_total": due_total,
            "academic_year": fee_calc.get("academic_year") or current_academic_year(),
        },
        payment_history=payment_history,
        education=education,
    )


# =========================
# PDF: ADMISSION LETTER
# =========================
@app.route("/student/admission-letter")
def admission_letter():
    if "student" not in session:
        return redirect("/")

    admission_id = session["student"]
    student, personal, education, docs = fetch_student_full_bundle(admission_id)
    if not student:
        return "Student not found", 404
    fee_summary = fetch_student_fee_summary(admission_id, student)
    pdf = generate_admission_letter(student, personal, education, docs, fee_summary)
    return send_file(pdf, as_attachment=True)


# =========================
# PDF: FEE RECEIPT
# =========================
@app.route("/student/fee-receipt")
def fee_receipt():
    if "student" not in session:
        return redirect("/")

    admission_id = session["student"]
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT * FROM students WHERE admission_id=%s", (admission_id,))
    student = cur.fetchone()

    cur.execute("SELECT * FROM fees WHERE admission_id=%s", (admission_id,))
    fees = cur.fetchone()

    pdf = generate_fee_receipt(student, fees)
    return send_file(pdf, as_attachment=True)


@app.route("/student/fees/payment/<int:payment_id>/receipt")
def student_fee_payment_receipt(payment_id):
    if "student" not in session:
        return redirect("/")

    admission_id = session["student"]
    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT
            fp.*,
            s.student_name,
            s.branch
        FROM fee_payments fp
        JOIN students s ON s.admission_id = fp.admission_id
        WHERE fp.id=%s AND fp.admission_id=%s
    """, (payment_id, admission_id))
    payment = cur.fetchone()
    cur.close()
    db.close()

    if not payment:
        return "Payment not found", 404

    receipt_path = generate_fee_receipt(
        {
            "admission_id": payment["admission_id"],
            "student_name": payment["student_name"],
            "branch": payment["branch"],
        },
        {
            "admission_fee": payment["amount"] if payment["fee_type"] == "ADMISSION" else 0,
            "tuition_fee": payment["amount"] if payment["fee_type"] == "TUITION" else 0,
            "management_fee": payment["amount"] if payment["fee_type"] == "MANAGEMENT" else 0,
            "exam_fee": payment["amount"] if payment["fee_type"] == "EXAM" else 0,
            "payment_type": payment["fee_type"],
            "payment_status": payment["fee_type"],
            "receipt_no": payment["receipt_no"],
            "payment_date": payment["payment_date"],
            "academic_year": payment.get("academic_year"),
            "semester_no": payment.get("semester_no"),
        }
    )
    return send_file(
        receipt_path,
        as_attachment=True,
        download_name=f"receipt_{payment['receipt_no']}.pdf"
    )


@app.route("/admin/student/<admission_id>")
def admin_view_student(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"] and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can view only your department students.", 403

    student, personal, education, docs = fetch_student_full_bundle(admission_id)
    if not student:
        return "Student not found", 404
    fee_summary = fetch_student_fee_summary(admission_id, student)

    return render_template(
        "admin_view_student.html",
        student=student,
        personal=personal,
        education=education,
        docs=docs,
        fee_summary=fee_summary,
        print_date=datetime.today().strftime("%d-%m-%Y")
    )


@app.route("/admin/student/<admission_id>/admission-pdf")
def admin_student_admission_pdf(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"] and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can download only your department students admission PDF.", 403

    student, personal, education, docs = fetch_student_full_bundle(admission_id)
    if not student:
        return "Student not found", 404
    fee_summary = fetch_student_fee_summary(admission_id, student)
    pdf = generate_admission_letter(student, personal, education, docs, fee_summary)
    return send_file(pdf, as_attachment=True)


@app.route("/admin/student/<admission_id>/upload-docs", methods=["POST"])
def admin_upload_student_docs(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"] and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can upload documents only for your department students.", 403

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT admission_id FROM students WHERE admission_id=%s", (admission_id,))
    if not cur.fetchone():
        cur.close()
        db.close()
        return "Student not found", 404

    def save_admin_doc(file_obj, prefix):
        if not file_obj or file_obj.filename == "":
            return None
        if not allowed_file(file_obj.filename):
            return "INVALID_FILE"
        filename = f"{admission_id}_{prefix}_{secure_filename(file_obj.filename)}"
        file_obj.save(os.path.join(UPLOAD_FOLDER, filename))
        return filename

    photo_name = save_admin_doc(request.files.get("student_photo"), "photo")
    aadhaar_name = save_admin_doc(request.files.get("aadhaar_file"), "aadhaar")
    caste_name = save_admin_doc(request.files.get("caste_file"), "caste")
    income_name = save_admin_doc(request.files.get("income_file"), "income")
    marks_name = save_admin_doc(request.files.get("marks_card_file"), "marks")

    if "INVALID_FILE" in [photo_name, aadhaar_name, caste_name, income_name, marks_name]:
        cur.close()
        db.close()
        return "Invalid file type. Allowed: pdf, jpg, jpeg, png", 400

    cur.execute("SELECT id FROM student_documents WHERE admission_id=%s", (admission_id,))
    existing = cur.fetchone()
    write_cur = db.cursor()

    if existing:
        write_cur.execute("""
            UPDATE student_documents
            SET student_photo=COALESCE(%s, student_photo),
                aadhaar_file=COALESCE(%s, aadhaar_file),
                caste_file=COALESCE(%s, caste_file),
                income_file=COALESCE(%s, income_file),
                marks_card_file=COALESCE(%s, marks_card_file)
            WHERE admission_id=%s
        """, (photo_name, aadhaar_name, caste_name, income_name, marks_name, admission_id))
    else:
        write_cur.execute("""
            INSERT INTO student_documents
            (admission_id, student_photo, aadhaar_file, caste_file, income_file, marks_card_file)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (admission_id, photo_name, aadhaar_name, caste_name, income_name, marks_name))

    db.commit()
    write_cur.close()
    cur.close()
    db.close()
    return redirect(url_for("admin_view_student", admission_id=admission_id))

import zipfile
from io import BytesIO

import os
import zipfile
from flask import send_file
@app.route("/admin/download-all/<admission_id>")
def download_all_docs(admission_id):
    scope = get_access_scope()
    if not scope["allowed"]:
        return redirect("/")
    if scope["is_staff"] and not student_in_department(admission_id, scope["department"]):
        return "Forbidden: You can download only your department students documents.", 403

    student, personal, education, docs = fetch_student_full_bundle(admission_id)
    if not student:
        return "Student not found", 404
    fee_summary = fetch_student_fee_summary(admission_id, student)

    if not any(docs.values()):
        return "No documents found for this admission ID", 404

    zip_path = f"static/uploads/{admission_id}_documents.zip"
    admission_pdf = generate_admission_letter(student, personal, education, docs, fee_summary)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        if os.path.exists(admission_pdf):
            z.write(admission_pdf, os.path.basename(admission_pdf))
        for filename in docs.values():
            if filename:
                file_path = os.path.join("static/uploads", filename)
                if os.path.exists(file_path):
                    z.write(file_path, filename)

    return send_file(zip_path, as_attachment=True)
# imports
import os
import random
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# ✅ save_file MUST be here
def save_file(file, prefix, admission_id):
    if not file or file.filename == "":
        return None

    filename = secure_filename(file.filename)
    ext = filename.rsplit(".", 1)[-1]

    unique_name = f"{admission_id}_{prefix}_{random.randint(100000,999999)}.{ext}"
    file_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)

    file.save(file_path)
    return unique_name


# ✅ THEN your upload route


@app.route("/admission/upload-documents", methods=["POST"])
def upload_documents():

    # 1️⃣ admission id
    admission_id = request.form.get("admission_id")

    # 2️⃣ save files
    photo = save_file(request.files.get("student_photo"), "photo", admission_id)
    aadhaar = save_file(request.files.get("aadhaar_file"), "aadhaar", admission_id)
    caste = save_file(request.files.get("caste_file"), "caste", admission_id)
    income = save_file(request.files.get("income_file"), "income", admission_id)
    marks = save_file(request.files.get("marks_card_file"), "marks", admission_id)
    aadhaar_number = (request.form.get("aadhaar_number") or "").strip().upper()
    caste_rd_number = (request.form.get("caste_rd_number") or "").strip().upper()
    income_rd_number = (request.form.get("income_rd_number") or "").strip().upper()

    # 3️⃣ student_documents table
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        "SELECT id FROM student_documents WHERE admission_id=%s",
        (admission_id,)
    )
    row = cursor.fetchone()

    if row:
        cursor.execute("""
            UPDATE student_documents SET
                student_photo=%s,
                aadhaar_file=%s,
                caste_file=%s,
                income_file=%s,
                marks_card_file=%s
            WHERE admission_id=%s
        """, (photo, aadhaar, caste, income, marks, admission_id))
    else:
        cursor.execute("""
            INSERT INTO student_documents
            (admission_id, student_photo, aadhaar_file, caste_file, income_file, marks_card_file)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (admission_id, photo, aadhaar, caste, income, marks))

    conn.commit()
    cursor.close()
    conn.close()

    # Mirror document details in personal table as well.
    db_docs = get_db()
    cur_docs = db_docs.cursor()
    cur_docs.execute("""
        UPDATE student_personal_details
        SET photo_file=COALESCE(%s, photo_file),
            caste_certificate_file=COALESCE(%s, caste_certificate_file),
            income_certificate_file=COALESCE(%s, income_certificate_file),
            marks_card_file=COALESCE(%s, marks_card_file),
            aadhaar_number=COALESCE(NULLIF(%s, ''), aadhaar_number),
            caste_rd_number=COALESCE(NULLIF(%s, ''), caste_rd_number),
            income_rd_number=COALESCE(NULLIF(%s, ''), income_rd_number)
        WHERE admission_id=%s
    """, (
        photo,
        caste,
        income,
        marks,
        aadhaar_number,
        caste_rd_number,
        income_rd_number,
        admission_id
    ))
    db_docs.commit()
    cur_docs.close()
    db_docs.close()

    # 4️⃣ get admission data from session
    admission = session.get("admission")
    if not admission:
        return redirect("/admission/step-1")

    # 5️⃣ students table (login)
    db = get_db()
    cur = db.cursor()
    ensure_students_admission_year_column()

    cur.execute(
        "SELECT admission_id FROM students WHERE admission_id=%s",
        (admission_id,)
    )

    if not cur.fetchone():
        cur.execute("""
            INSERT INTO students
            (admission_id, student_name, branch, admission_year, mobile, password_hash, status)
            VALUES (%s,%s,%s,%s,%s,%s,'INACTIVE')
        """, (
            admission_id,
            admission["student_name"],
            admission["branch"],
            parse_int_prefix(admission.get("admission_year")) or datetime.today().year,
            admission["student_mobile"],
            generate_password_hash(admission["password"])

        ))

    db.commit()
    cur.close()
    db.close()

    # 🔥🔥🔥 ADD YOUR CODE HERE 🔥🔥🔥
    # ===============================
    # INSERT PERSONAL DETAILS (STEP 1 + 2)
    # ===============================
    db = get_db()
    cur = db.cursor()

    cur.execute(
        "SELECT id FROM student_personal_details WHERE admission_id=%s",
        (admission_id,)
    )

    if not cur.fetchone():
        cur.execute("""
            INSERT INTO student_personal_details (
                admission_id,
                student_mobile, student_email, disability,
                dob, gender, indian_nationality, religion,
                caste_category, alloted_category,
                qualifying_exam, year_of_passing, register_number,
                admission_quota,
                father_name, father_mobile,
                mother_name, mother_mobile,
                residential_address, permanent_address
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            admission_id,
            admission.get("student_mobile"),
            admission.get("student_email"),
            admission.get("disability", "NO"),
            admission["dob"],
            admission["gender"],
            admission["indian_nationality"],
            admission["religion"],
            admission["caste_category"],
            admission["alloted_category"],
            admission["qualifying_exam"],
            admission["year_of_passing"],
            admission["register_number"],
            admission["admission_quota"],
            admission["father_name"],
            admission["father_mobile"],
            admission["mother_name"],
            admission["mother_mobile"],
            admission["residential_address"],
            admission["permanent_address"]
        ))

    db.commit()
    cur.close()
    db.close()

    # 6️⃣ clear session & show success
    session.pop("admission", None)

    return render_template(
        "admission_success.html",
        admission_id=admission_id
    )


@app.route("/student/reupload", methods=["GET", "POST"])
def student_reupload():
    if "student" not in session:
        return redirect("/login")

    admission_id = session["student"]

    if request.method == "POST":

        photo = request.files.get("photo")
        marks = request.files.get("marks_card")
        caste = request.files.get("caste_certificate")
        income = request.files.get("income_certificate")

        def save_file(file, prefix):
            if file and file.filename:
                if not allowed_file(file.filename):
                    return None, "❌ Invalid file type"
                name = f"{admission_id}_{prefix}_{secure_filename(file.filename)}"
                file.save(os.path.join(UPLOAD_FOLDER, name))
                return name, None
            return None, None

        photo_name, err = save_file(photo, "photo")
        if err: return err

        marks_name, err = save_file(marks, "marks")
        if err: return err

        caste_name, err = save_file(caste, "caste")
        if err: return err

        income_name, err = save_file(income, "income")
        if err: return err

        db = get_db()
        cur = db.cursor()

        # 🔁 UPDATE ONLY UPLOADED FILES
        if photo_name:
            cur.execute(
                "UPDATE student_personal_details SET photo_file=%s WHERE admission_id=%s",
                (photo_name, admission_id)
            )
        if marks_name:
            cur.execute(
                "UPDATE student_personal_details SET marks_card_file=%s WHERE admission_id=%s",
                (marks_name, admission_id)
            )
        if caste_name:
            cur.execute(
                "UPDATE student_personal_details SET caste_certificate_file=%s WHERE admission_id=%s",
                (caste_name, admission_id)
            )
        if income_name:
            cur.execute(
                "UPDATE student_personal_details SET income_certificate_file=%s WHERE admission_id=%s",
                (income_name, admission_id)
            )

        # 🔁 RESET STATUS FOR REVIEW
        cur.execute("""
            UPDATE students
            SET status='INACTIVE', rejection_reason=NULL
            WHERE admission_id=%s
        """, (admission_id,))

        db.commit()

        return """
        <h3>✅ Documents Re-uploaded Successfully</h3>
        <p>Your application is sent back for admin review.</p>
        <a href="/login">Back to Login</a>
        """

    return render_template("student_reupload.html")

# =========================
# RUN SERVER
# =========================
if __name__ == "__main__":
    app.run(debug=True)
